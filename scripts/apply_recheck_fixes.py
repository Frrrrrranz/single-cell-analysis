"""豆包交叉核对修正（2026-09-02）：PMID_35115729 四项补录/修正。

来源：豆包教学范例交叉核对 + review_md 原文复核。
- 补录 Slit2（mSC，author_declared，Results p.6）
- 补录 epineurial fibroblasts + Pdgfra
- Itgb4/Slc2a1 evidence_type 升级 annotation_marker → author_declared
- M01532-M01534（Schwann cell/myelinating-SC）语义重复行移入 audit_exclusions
"""

from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(r"d:\OneDrive\Desktop\组\marker提取")
XLSX = ROOT / "表单" / "our_markers.xlsx"
BACKUP = ROOT / "audited-extraction" / "recovery" / "our_markers_pre_recheck_backup_2026-09-02.xlsx"

PAPER = "PMID_35115729"
TASK_NO = 38
DATASET = "GSE182098"
TODAY = datetime(2026, 9, 2)
REVIEW_METHOD = "doubao_crosscheck_2026-09-02"
SOURCE_MD = "marker提取/review_md/PMID_35115729.md"


def new_row(marker_id: int) -> dict:
    return {
        "marker_id": f"M{marker_id:05d}",
        "task_no": TASK_NO,
        "dataset_id": DATASET,
        "paper_id": PAPER,
        "document_id": PAPER,
        "document_role": "primary",
        "ct_id": None,
        "subtype_id": None,
        "species": "mouse",
        "candidate_class": "formal_candidate",
        "marker_polarity": "positive",
        "review_status": "approved",
        "review_method": REVIEW_METHOD,
        "source_file": SOURCE_MD,
        "imported_at": TODAY,
        "audit_status": "recheck_include",
        "normalization_status": "exact",
        "citation_verified": True,
        "audit_model": "doubao_crosscheck_manual_verified",
        "recovery_source": "doubao_crosscheck",
    }


ADD_ROWS = [
    {
        **new_row(1884),
        "cell_type": "myelinating Schwann cell (mSC)",
        "subtype": "Slit2-high population (global Cluster 8)",
        "is_pns_cell": "true",
        "gene_symbol": "Slit2",
        "original_symbol": "Slit2",
        "evidence_type": "author_declared",
        "source_locator": "Results p.6, Fig. 4a-c",
        "source_context": (
            "the SC subpopulation marker genes identified in our snRNAseq analysis ... several of the unique "
            "mSC genes, including Slit2, Col23a1, Adamtsl1, Cldn14, and Pmp2; The Slit2-high population from "
            "global analysis further separated into two clusters"
        ),
        "notes": (
            "豆包交叉核对发现漏提：作者将 Slit2 与已收录的 Col23a1/Adamtsl1/Cldn14/Pmp2 并列为 "
            "SC subpopulation marker genes，且用 Slit2-high 命名亚群；此前三轮（v2/审计/恢复）均未入池"
        ),
        "audit_notes": "复核轮 2026-09-02 补录",
        "four_layer_category": "L2",
    },
    {
        **new_row(1885),
        "cell_type": "epineurial fibroblasts",
        "subtype": None,
        "is_pns_cell": "false",
        "gene_symbol": "Pdgfra",
        "original_symbol": "Pdgfra",
        "evidence_type": "annotation_marker",
        "source_locator": "Results p.4, Fig. 1e-f",
        "source_context": (
            "Epineurial fibroblasts, which surround the outermost layer of the nerve, express Pdgfra and Pcolce"
        ),
        "notes": (
            "豆包交叉核对发现：Pdgfra 此前仅记录于 endoneurial fibroblasts (EFs) 行；原文对 epineurial "
            "fibroblasts 同样明确列出 Pdgfra，与 Pcolce 行对称，补录外膜行"
        ),
        "audit_notes": "复核轮 2026-09-02 补录",
        "four_layer_category": "L3",
    },
]

EVIDENCE_UPGRADES = {
    "M01528": {
        "source_context": (
            "Perineurial fibroblasts (Cluster 10), which ensheath nerve fascicles, were notable for their lack "
            "of Pdgfra expression and their expression of markers Itgb4 and Slc2a1 (also known as Glut1)"
        ),
        "notes": (
            "复核修正 2026-09-02：原文 'expression of markers Itgb4 and Slc2a1' 为明确 marker 措辞，"
            "evidence_type 由 annotation_marker 升级为 author_declared"
        ),
    },
    "M01529": {
        "source_context": (
            "Perineurial fibroblasts (Cluster 10), which ensheath nerve fascicles, were notable for their lack "
            "of Pdgfra expression and their expression of markers Itgb4 and Slc2a1 (also known as Glut1)"
        ),
        "notes": (
            "复核修正 2026-09-02：原文 'expression of markers Itgb4 and Slc2a1' 为明确 marker 措辞，"
            "evidence_type 由 annotation_marker 升级为 author_declared"
        ),
    },
}

REMOVE_IDS = ["M01532", "M01533", "M01534"]
REMOVE_REASON = (
    "语义重复：与 M00082-M00084（myelinating Schwann cell (mSC) / Cldn14+ Adamtsl1+ Pmp2+ (motor-associated)）"
    "指向同一证据（Fig. 4a cluster 3，Abstract 与 Results 重复陈述），因 cell_type 写法 'Schwann cell/myelinating-SC' "
    "不同绕过唯一键去重；保留更具体的 author_declared 行"
)


def main() -> None:
    if not BACKUP.exists():
        shutil.copy2(XLSX, BACKUP)
        print(f"backup -> {BACKUP.name}")
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["markers"]
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}

    # 1) 追加 2 行
    for row_data in ADD_ROWS:
        exists = any(
            r[idx["paper_id"]] == row_data["paper_id"]
            and r[idx["cell_type"]] == row_data["cell_type"]
            and r[idx["gene_symbol"]] == row_data["gene_symbol"]
            for r in ws.iter_rows(min_row=2, values_only=True)
        )
        if exists:
            print(f"skip existing: {row_data['cell_type']} + {row_data['gene_symbol']}")
            continue
        ws.append([row_data.get(h) for h in header])
        print(f"added {row_data['marker_id']}: {row_data['cell_type']} + {row_data['gene_symbol']}")

    # 2) evidence_type 升级
    for row in ws.iter_rows(min_row=2):
        mid = row[idx["marker_id"]].value
        if mid in EVIDENCE_UPGRADES:
            up = EVIDENCE_UPGRADES[mid]
            gene = row[idx["gene_symbol"]].value
            row[idx["evidence_type"]].value = "author_declared"
            row[idx["source_context"]].value = up["source_context"]
            row[idx["notes"]].value = up["notes"]
            old = row[idx["audit_notes"]].value or ""
            row[idx["audit_notes"]].value = f"{old}; 复核轮 2026-09-02 evidence_type 升级".lstrip("; ")
            rm = row[idx["review_method"]].value or ""
            row[idx["review_method"]].value = f"{rm}; recheck_2026-09-02"
            print(f"upgraded {mid}: {gene} -> author_declared")

    # 3) 语义重复行移入 audit_exclusions 并从 markers 删除
    ex = wb["audit_exclusions"]
    ex_header = [c.value for c in ex[1]]
    ex_idx = {h: i for i, h in enumerate(ex_header)}
    removed_rows = []
    for row in list(ws.iter_rows(min_row=2)):
        mid = row[idx["marker_id"]].value
        if mid in REMOVE_IDS:
            values = {h: row[idx[h]].value for h in header}
            removed_rows.append(values)
    for values in removed_rows:
        ex.append(
            [
                values["paper_id"],
                values["task_no"],
                values["cell_type"],
                values["subtype"],
                values["species"],
                values["original_symbol"],
                values["gene_symbol"],
                values["normalization_status"],
                "author_declared",
                values["marker_polarity"],
                "exclude",
                REMOVE_REASON,
                values["source_locator"],
                values["source_context"],
                None,
                values["citation_verified"],
                "doubao_crosscheck_manual_verified",
                values["marker_id"],
                "dedup_removed_2026-09-02",
            ]
        )
    for row in list(ws.iter_rows(min_row=2)):
        if row[idx["marker_id"]].value in REMOVE_IDS:
            ws.delete_rows(row[0].row)
            print(f"removed {row[idx['marker_id']].value} -> audit_exclusions")

    # 4) import_log 记录
    log = wb["import_log"]
    log_header = [c.value for c in log[1]]
    log_row = {h: None for h in log_header}
    log_row.update(
        {
            "batch_id": "B20260902-RECHECK",
            "task_no": TASK_NO,
            "paper_id": PAPER,
            "document_id": PAPER,
            "paper_title": "Disentangling glial diversity in peripheral nerves at single-nuclei resolution",
            "review_method": "doubao_crosscheck",
            "formal_candidates": 2,
            "quick_approved": 2,
            "imported_count": 2,
            "source_file": SOURCE_MD,
            "imported_at": TODAY,
            "notes": (
                "豆包交叉核对修正：补录 Slit2（mSC）与 epineurial Pdgfra；Itgb4/Slc2a1 升级 author_declared；"
                "移除 3 条语义重复行（M01532-34 入 audit_exclusions）；净变化 -1 行（1883→1882）"
            ),
        }
    )
    log.append([log_row.get(h) for h in log_header])

    # 5) 说明与统计追加复核轮说明
    st = wb["说明与统计"]
    st.append(["复核轮说明（doubao crosscheck 2026-09-02）"])
    st.append(
        [
            "以豆包教学范例交叉核对 PMID_35115729：补录 2 条（Slit2、epineurial Pdgfra）；Itgb4/Slc2a1 "
            "升级 author_declared；3 条语义重复行（M01532-34）移入 audit_exclusions；总行数 1883→1882。"
        ]
    )

    wb.save(XLSX)
    print("saved.")


if __name__ == "__main__":
    main()
