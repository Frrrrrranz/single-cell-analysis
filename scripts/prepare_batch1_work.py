import openpyxl, os, io, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

BASE = r'D:\OneDrive\Desktop\组'
WORK = os.path.join(BASE, r'marker提取\audits\recheck-2026-09-02\batch1_work')
os.makedirs(WORK, exist_ok=True)

PAPERS = {
    'DOI_10.1038_s41586-020-2922-4': '人肺图谱',
    'DOI_10.1038_s41591-024-03215-z': '乳腺癌多模态',
    'DOI_10.1038_s41588-022-01243-4': '人肺空间图谱',
    'DOI_10.7554_elife.71752': '人DRG',
    'DOI_10.1101_2025.09.26.678707': '肾损伤',
}

plan = openpyxl.load_workbook(os.path.join(BASE, r'marker提取\audits\recheck-2026-09-02\recheck_plan.xlsx'), read_only=True, data_only=True)
total = openpyxl.load_workbook(os.path.join(BASE, r'marker提取\表单\our_markers.xlsx'), read_only=True, data_only=True)

def cell(v, n=400):
    if v is None:
        return ''
    s = str(v).replace('\n', ' ')
    return s[:n] + ('...' if len(s) > n else '')

# 候选导出（A/B/B2/C + 试点遗留）
sheets = [('A_漏提候选', 'paper_id'), ('B_证据类型候选', 'paper_id'), ('B2_补充候选', 'paper_id'), ('C_语义重复候选', 'paper_id')]
for pid, label in PAPERS.items():
    lines = [f'# {pid} ({label}) 候选材料']
    for sn, pcol in sheets:
        ws = plan[sn]
        rows = list(ws.iter_rows(values_only=True))
        h = list(rows[0])
        pi = h.index(pcol)
        n = 0
        sec = [f'\n## {sn}\n']
        for r in rows[1:]:
            if not r or all(v is None for v in r):
                continue
            if str(r[pi] or '') != pid:
                continue
            n += 1
            sec.append(f'### 候选{n}')
            for i, v in enumerate(r):
                if v is not None:
                    sec.append(f'- {h[i]}: {cell(v)}')
        lines.extend(sec)
        lines.append(f'\n({sn} 共 {n} 条)')
    with open(os.path.join(WORK, f'{pid}_candidates.md'), 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

# 试点遗留单独
ws = plan['B_证据类型候选']
rows = list(ws.iter_rows(values_only=True))
h = list(rows[0])
lines = ['# 试点 PMID_35115729 遗留 B 候选（并入 Batch1 处置）']
for r in rows[1:]:
    if r and r[0] in ('M01516', 'M01517'):
        lines.append('## ' + str(r[0]))
        for i, v in enumerate(r):
            if v is not None:
                lines.append(f'- {h[i]}: {cell(v)}')
with open(os.path.join(WORK, 'PMID_35115729_leftover.md'), 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

# 总表现有行导出
ws = total['markers']
rows = list(ws.iter_rows(values_only=True))
h = list(rows[0])
KEEP = ['marker_id', 'cell_type', 'subtype', 'species', 'is_pns_cell', 'gene_symbol',
        'evidence_type', 'marker_polarity', 'source_locator', 'source_context',
        'audit_status', 'review_method', 'four_layer_category']
keep_idx = [h.index(k) for k in KEEP]
for pid, label in PAPERS.items():
    lines = [f'# {pid} ({label}) 总表现有行']
    n = 0
    pi = h.index('paper_id')
    for r in rows[1:]:
        if not r:
            continue
        if str(r[pi] or '') != pid:
            continue
        n += 1
        parts = [f'{k}={cell(r[j], 350)}' for k, j in zip(KEEP, keep_idx) if j < len(r) and r[j] is not None]
        lines.append(f'## 行{n}: ' + ' | '.join(parts))
    lines.append(f'\n共 {n} 行')
    with open(os.path.join(WORK, f'{pid}_existing.md'), 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

# 试点现有行（供遗留候选判定参考）
lines = ['# 试点 PMID_35115729 总表现有行']
n = 0
for r in rows[1:]:
    if not r:
        continue
    if str(r[h.index('paper_id')] or '') != 'PMID_35115729':
        continue
    n += 1
    parts = [f'{k}={cell(r[j], 350)}' for k, j in zip(KEEP, keep_idx) if j < len(r) and r[j] is not None]
    lines.append(f'## 行{n}: ' + ' | '.join(parts))
lines.append(f'\n共 {n} 行')
with open(os.path.join(WORK, 'PMID_35115729_existing.md'), 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

plan.close()
total.close()
print('work files written to', WORK)
for fn in sorted(os.listdir(WORK)):
    print(' ', fn, os.path.getsize(os.path.join(WORK, fn)))
