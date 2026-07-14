"""
migrate_old_markers.py — Task 6: 迁移已有数据

将 cell_types 和 cell_subtypes 表中已有的 markers 数据迁移到新的 markers sheet。

迁移逻辑：
1. 解析 cell_types.markers 分号分隔的列表 → 每行一个 marker
2. 解析 cell_subtypes.markers 同样处理
3. evidence_level 统一标记为 'imported'（旧数据无法追溯原始证据）
4. provenance 列中的来源信息写入 source_section
5. review_status 统一标记为 'pending'（需要复核确认）

用法：
    python migrate_old_markers.py [--db PATH]
"""
import argparse
import logging
import re
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

DB_PATH = Path(r"D:\OneDrive\Desktop\组\db\pns-scrna.xlsx")

# 从 cell_types/markers 拆分时，识别分号但忽略括号内的分号
MARKER_SPLIT_PATTERN = re.compile(r";(?![^(]*\))")


def parse_markers(markers_str: Optional[str]) -> list[str]:
    """解析 markers 字符串（分号分隔），返回基因列表"""
    if not markers_str or str(markers_str).strip() in ("", "NA", "null"):
        return []
    markers = [m.strip() for m in MARKER_SPLIT_PATTERN.split(str(markers_str))]
    return [m for m in markers if m and m not in ("NA", "null", "")]


def get_next_marker_id(ws) -> str:
    """读取 markers sheet 当前最大 marker_id 并自增"""
    max_num = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        mid = row[0]
        if mid:
            m = re.match(r"M(\d+)", str(mid))
            if m:
                num = int(m.group(1))
                max_num = max(max_num, num)
    return f"M{max_num + 1:05d}"


def find_existing_entry(ws_markers, ct_id: str, subtype_id: Optional[str],
                         gene: str) -> bool:
    """检查 markers sheet 中是否已存在相同记录"""
    headers = [c.value for c in ws_markers[1]]
    col_map = {h: i + 1 for i, h in enumerate(headers)}

    for row in ws_markers.iter_rows(min_row=2, max_row=ws_markers.max_row,
                                    values_only=True):
        if (row[col_map.get("ct_id", 2) - 1] == ct_id and
                row[col_map.get("subtype_id", 3) - 1] == subtype_id and
                str(row[col_map.get("gene_symbol", 4) - 1] or "").upper() == gene.upper()):
            return True
    return False


def migrate_cell_types(ws_cell_types, ws_markers) -> int:
    """迁移 cell_types 表的 markers 到 markers sheet，返回迁移条数"""
    headers_ct = [c.value for c in ws_cell_types[1]]
    col_map_ct = {h: i + 1 for i, h in enumerate(headers_ct)}

    marker_headers = [c.value for c in ws_markers[1]]
    marker_col_map = {h: i + 1 for i, h in enumerate(marker_headers)}

    next_id = get_next_marker_id(ws_markers)
    count = 0

    for row in ws_cell_types.iter_rows(min_row=2, max_row=ws_cell_types.max_row,
                                       values_only=False):
        paper_id = row[col_map_ct["paper_id"] - 1].value
        ct_id = row[col_map_ct["ct_id"] - 1].value
        cell_type = row[col_map_ct["cell_type"] - 1].value
        markers_str = row[col_map_ct["markers"] - 1].value
        provenance = row[col_map_ct["provenance"] - 1].value

        genes = parse_markers(markers_str)
        if not genes:
            continue

        source_section = str(provenance).strip() if provenance and provenance != "NA" else ""

        for gene in genes:
            # 去重检查
            if find_existing_entry(ws_markers, ct_id, None, gene):
                continue

            new_row = ws_markers.max_row + 1
            row_data = {
                "marker_id": next_id,
                "ct_id": ct_id,
                "gene_symbol": gene,
                "original_symbol": gene,
                "evidence_level": "imported",
                "source_section": source_section,
                "review_status": "pending",
                "notes": f"从 cell_types 迁移（paper={paper_id}, cell_type={cell_type}）",
            }
            for col_name, value in row_data.items():
                col_idx = marker_col_map.get(col_name)
                if col_idx:
                    ws_markers.cell(row=new_row, column=col_idx, value=value)

            next_id = f"M{int(next_id[1:]) + 1:05d}"
            count += 1

    logger.info(f"  cell_types → markers: {count} 条")
    return count


def migrate_cell_subtypes(ws_subtypes, ws_markers) -> int:
    """迁移 cell_subtypes 表的 markers 到 markers sheet"""
    headers_st = [c.value for c in ws_subtypes[1]]
    col_map_st = {h: i + 1 for i, h in enumerate(headers_st)}

    marker_headers = [c.value for c in ws_markers[1]]
    marker_col_map = {h: i + 1 for i, h in enumerate(marker_headers)}

    next_id = get_next_marker_id(ws_markers)
    count = 0

    for row in ws_subtypes.iter_rows(min_row=2, max_row=ws_subtypes.max_row,
                                     values_only=False):
        paper_id = row[col_map_st["paper_id"] - 1].value
        subtype_id = row[col_map_st["subtype_id"] - 1].value
        parent_ct = row[col_map_st["parent_cell_type"] - 1].value
        subtype = row[col_map_st["subtype"] - 1].value
        markers_str = row[col_map_st["markers"] - 1].value
        provenance = row[col_map_st["provenance"] - 1].value

        genes = parse_markers(markers_str)
        if not genes:
            continue

        source_section = str(provenance).strip() if provenance and provenance != "NA" else ""

        for gene in genes:
            if find_existing_entry(ws_markers, None, subtype_id, gene):
                continue

            new_row = ws_markers.max_row + 1
            row_data = {
                "marker_id": next_id,
                "subtype_id": subtype_id,
                "gene_symbol": gene,
                "original_symbol": gene,
                "evidence_level": "imported",
                "source_section": source_section,
                "review_status": "pending",
                "notes": f"从 cell_subtypes 迁移（paper={paper_id}, subtype={parent_ct}/{subtype}）",
            }
            for col_name, value in row_data.items():
                col_idx = marker_col_map.get(col_name)
                if col_idx:
                    ws_markers.cell(row=new_row, column=col_idx, value=value)

            next_id = f"M{int(next_id[1:]) + 1:05d}"
            count += 1

    logger.info(f"  cell_subtypes → markers: {count} 条")
    return count


def main() -> None:
    parser = argparse.ArgumentParser(description="迁移旧数据到 markers sheet")
    parser.add_argument("--db", default=str(DB_PATH),
                        help=f"数据库路径 (默认: {DB_PATH})")
    args = parser.parse_args()

    db_path = Path(args.db)
    if not db_path.exists():
        logger.error(f"数据库文件不存在: {db_path}")
        return

    wb = load_workbook(db_path)
    ws_markers = wb["markers"]
    ws_cell_types = wb["cell_types"]
    ws_subtypes = wb["cell_subtypes"]

    logger.info("开始迁移已有 marker 数据...")

    # 先迁移 subtype（不占用 ct_id 关联），再迁移 cell_type
    subtype_count = migrate_cell_subtypes(ws_subtypes, ws_markers)
    ct_count = migrate_cell_types(ws_cell_types, ws_markers)

    # 更新 cell_types.mark_status（只有 ct_count > 0 的 paper 需要更新）
    if ct_count > 0:
        headers_ct = [c.value for c in ws_cell_types[1]]
        col_map_ct = {h: i + 1 for i, h in enumerate(headers_ct)}
        if "mark_status" in col_map_ct:
            updated = 0
            for row in ws_cell_types.iter_rows(min_row=2, max_row=ws_cell_types.max_row,
                                               values_only=False):
                markers_str = row[col_map_ct["markers"] - 1].value
                if markers_str and str(markers_str).strip() not in ("", "NA"):
                    current = row[col_map_ct["mark_status"] - 1].value
                    if current is None or str(current).strip() == "":
                        row[col_map_ct["mark_status"] - 1].value = "old"
                        updated += 1
            logger.info(f"  更新 mark_status: {updated} 行标记为 'old'")

    wb.save(db_path)
    wb.close()

    total = ct_count + subtype_count
    logger.info(f"\n✅ 迁移完成: 共 {total} 条 marker 写入 markers sheet")
    if total > 0:
        logger.info("注意: 所有迁移数据的 review_status = pending，需要人工复核后方可使用。")


if __name__ == "__main__":
    main()
