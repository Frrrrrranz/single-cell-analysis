"""从旧版逐篇复核 CSV 生成保守的 PNS marker 快速复核候选。

本脚本只生成 JSON 中间结果，不写 Excel。旧结果仅作为候选来源；
只保留细胞类型明确属于 PNS 且作者直接给出标记/注释语境的记录。
"""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
TASK_WORKBOOK = PROJECT_ROOT / "db" / "cellxgene" / "our_marker_papers.xlsx"
LEGACY_DIR = Path(__file__).resolve().parent / "markers_output"
OUTPUT_DIR = Path(__file__).resolve().parent / "review_md"
OUTPUT_PATH = OUTPUT_DIR / "quick_review_candidates.json"

INVALID_LEGACY_PAPER_IDS = {
    "DOI_10.1016_j.jcf.2025.01.016",
    "DOI_10.1038_s41586-020-2922-4",
    "DOI_10.1038_s41586-021-03929-x",
    "DOI_10.1038_s44318-024-00328-6",
    "DOI_10.1101_2025.09.26.678707",
}

CUSTOM_REVIEW_FILES = {
    "DOI_10.1016_j.stem.2022.11.013": "Organoid_modeling_of_human_fetal_lung_alveolar_development_r_review.csv",
    "DOI_10.1038_s41586-020-2496-1": "NATURE.587.619.2020_review.csv",
}

PNS_CELL_PATTERN = re.compile(
    r"schwann|satellite glia|sensory neuron|nociceptor|neuroendocrine|"
    r"enteroendocrine|enteric neuron|peripheral nervous|cardiac neuron|"
    r"sympathetic|parasympathetic|efferent neuron",
    re.IGNORECASE,
)

DIRECT_CONTEXT_PATTERN = re.compile(
    r"markers?|annotat|defined|identified|characteri[sz]ed|marker list|"
    r"\b[A-Za-z][A-Za-z0-9.-]{1,15}\+",
    re.IGNORECASE,
)

MANUAL_ALLOW = {
    (36, "nociceptors", "SST"),
    (42, "Myelinating Schwann cells", "Tgfb2"),
}

# 任务 5 的正文直接以“identified ... clusters”列出 Schwann 亚型及对应基因；
# 单行上下文未重复 marker 一词，但已由 Markdown 原文抽样核对。
MANUAL_ALLOW_ALL_EXPLICIT_TASKS = {5}

INVALID_OR_INCOMPLETE_SYMBOLS = {"CADM", "NRXN"}


def normalize_species(value: str) -> str:
    species = value.strip().lower()
    if species in {"human", "homo sapiens"}:
        return "human"
    if species in {"mouse", "mus musculus"}:
        return "mouse"
    if species in {"rat", "rattus norvegicus"}:
        return "rat"
    return species or "unknown"


def normalize_gene(value: str, species: str) -> str:
    gene = value.strip()
    if species == "human":
        return gene.upper()
    if species in {"mouse", "rat"}:
        return gene[:1].upper() + gene[1:].lower()
    return gene


def classify_evidence(locator: str, context: str) -> str:
    combined = f"{locator} {context}".lower()
    if "annotat" in combined or "marker list" in combined or "cell annotation" in combined:
        return "annotation_marker"
    if "fig" in locator.lower() and "marker" in context.lower():
        return "figure_labeled"
    if "marker" in combined:
        return "author_declared"
    return "annotation_marker"


def load_tasks() -> list[dict[str, object]]:
    workbook = load_workbook(TASK_WORKBOOK, read_only=True, data_only=False)
    sheet = workbook["我方Marker文章"]
    headers = [cell.value for cell in sheet[1]]
    tasks: list[dict[str, object]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        tasks.append(
            {
                "task_no": int(row["序号"]),
                "dataset_id": row["代表Dataset ID"],
                "paper_id": row["paper_id"],
                "paper_title": row["论文标题"],
                "task_species": row["物种"],
                "pdf_status": row["PDF状态"],
                "marker_status": row["Marker状态"],
            }
        )
    workbook.close()
    return tasks


def review_file_index() -> dict[str, Path]:
    index: dict[str, Path] = {}
    for path in LEGACY_DIR.glob("*_review.csv"):
        document_id = path.name[: -len("_review.csv")]
        index[normalized_file_key(document_id)] = path
    return index


def normalized_file_key(value: str) -> str:
    """兼容旧产物把 DOI 中斜杠、连字符和句点统一改成下划线的命名。"""
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def main() -> None:
    tasks = load_tasks()
    review_index = review_file_index()
    candidates: list[dict[str, object]] = []
    audits: list[dict[str, object]] = []

    for task in tasks:
        task_no = int(task["task_no"])
        paper_id = str(task["paper_id"] or "")
        if task["pdf_status"] != "已核验" or task["marker_status"] == "快速复核已录入":
            continue

        review_path: Path | None = None
        reason = ""
        if paper_id in INVALID_LEGACY_PAPER_IDS:
            reason = "旧结果存在已确认的论文映射污染，未复用"
        elif paper_id in CUSTOM_REVIEW_FILES:
            candidate_path = LEGACY_DIR / CUSTOM_REVIEW_FILES[paper_id]
            if candidate_path.exists():
                review_path = candidate_path
        else:
            review_path = review_index.get(normalized_file_key(paper_id))

        if review_path is None:
            audits.append(
                {
                    **task,
                    "source_file": None,
                    "legacy_rows": 0,
                    "pns_relevant_rows": 0,
                    "approved_rows": 0,
                    "status": "needs_reextract",
                    "notes": reason or "未找到可复用的旧复核 CSV",
                }
            )
            continue

        with review_path.open("r", encoding="utf-8-sig", newline="") as stream:
            rows = list(csv.DictReader(stream))

        relevant_count = 0
        approved_count = 0
        seen: set[tuple[str, str, str]] = set()
        for row in rows:
            cell_type = (row.get("cell_type") or "").strip()
            subtype = (row.get("subtype") or "").strip()
            gene_raw = (row.get("gene_symbol") or "").strip()
            if not PNS_CELL_PATTERN.search(f"{cell_type} {subtype}"):
                continue
            relevant_count += 1
            if gene_raw.upper() in INVALID_OR_INCOMPLETE_SYMBOLS:
                continue

            locator = (row.get("source_section") or "").strip()
            context = (row.get("source_context") or "").strip()
            evidence_level = (row.get("evidence_level") or "").strip().lower()
            allowed_manually = (
                (task_no, cell_type, gene_raw) in MANUAL_ALLOW
                or (task_no in MANUAL_ALLOW_ALL_EXPLICIT_TASKS and evidence_level == "explicit")
            )
            if evidence_level != "explicit" and not allowed_manually:
                continue
            if not DIRECT_CONTEXT_PATTERN.search(f"{locator} {context}") and not allowed_manually:
                continue

            species = normalize_species(row.get("species") or str(task["task_species"] or ""))
            gene = normalize_gene(gene_raw, species)
            dedupe_key = (cell_type.lower(), subtype.lower(), gene.lower())
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)

            candidates.append(
                {
                    "task_no": task_no,
                    "dataset_id": task["dataset_id"],
                    "paper_id": paper_id,
                    "document_id": paper_id,
                    "document_role": "primary",
                    "ct_id": None,
                    "subtype_id": None,
                    "cell_type": cell_type,
                    "subtype": subtype or None,
                    "species": species,
                    "is_pns_cell": "true",
                    "gene_symbol": gene,
                    "original_symbol": gene_raw,
                    "evidence_type": classify_evidence(locator, context),
                    "marker_polarity": "positive",
                    "candidate_class": "formal_candidate",
                    "source_locator": locator,
                    "source_context": context,
                    "review_status": "approved",
                    "review_method": "quick_consistency",
                    "notes": "快速一致性复核：作者直接给出PNS细胞标记/注释；抽样核对原文，未做逐句全量复核",
                    "source_file": str(review_path.relative_to(PROJECT_ROOT)).replace("\\", "/"),
                    "paper_title": task["paper_title"],
                }
            )
            approved_count += 1

        audits.append(
            {
                **task,
                "source_file": str(review_path.relative_to(PROJECT_ROOT)).replace("\\", "/"),
                "legacy_rows": len(rows),
                "pns_relevant_rows": relevant_count,
                "approved_rows": approved_count,
                "status": "ready_to_import" if approved_count else "reviewed_no_formal_pns_marker",
                "notes": "仅复用明确PNS细胞且作者直接给出marker/注释语境的旧候选",
            }
        )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps({"candidates": candidates, "audits": audits}, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps({
        "candidate_count": len(candidates),
        "paper_count": len({row["paper_id"] for row in candidates}),
        "audit_count": len(audits),
        "status_counts": {
            status: sum(1 for row in audits if row["status"] == status)
            for status in sorted({str(row["status"]) for row in audits})
        },
        "output": str(OUTPUT_PATH),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
