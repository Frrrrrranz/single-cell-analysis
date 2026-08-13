"""
import_markers.py — 复核后数据入库

功能：
1. 读取审核通过的 CSV（review_status=approved/modified）
2. 在 markers sheet 中分配新的 marker_id
3. 关联 ct_id（按 paper_id + cell_type 匹配）
4. 写入 markers sheet
5. 更新 cell_types 对应行的 mark_status

用法：
    python import_markers.py <review_csv_path> [--db PATH]

输入：{paper_id}_review.csv（仅 review_status 为 approved / modified 的行）
输出：pns-scrna.xlsx → markers sheet（追加写入）
"""
import argparse
import csv
import logging
import re
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from marker_schema import FORMAL_EVIDENCE_TYPES, MARKER_POLARITIES

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

DB_PATH = Path(r"D:\OneDrive\Desktop\组\db\pns-scrna.xlsx")

# 允许导入的 review_status 值
ALLOWED_STATUSES = {"approved", "modified"}
REQUIRED_REVIEW_COLUMNS = {
    "paper_id", "document_id", "document_role", "cell_type", "subtype", "species",
    "gene_symbol", "evidence_type", "marker_polarity", "candidate_class",
    "source_locator", "source_context", "review_status",
}
REQUIRED_MARKER_COLUMNS = {
    "marker_id", "paper_id", "document_id", "document_role", "ct_id", "subtype_id",
    "cell_type", "species", "gene_symbol", "original_symbol", "evidence_type",
    "marker_polarity", "candidate_class", "source_locator", "source_context",
    "review_status", "notes",
}


def get_next_marker_id(ws) -> str:
    """读取 markers sheet 当前最大 marker_id 并自增"""
    max_num = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        mid = row[0]
        if mid:
            m = re.match(r"M(\d+)", str(mid))
            if m:
                num = int(m.group(1))
                max_num = max(max_num, num)
    return f"M{max_num + 1:05d}"


def match_ct_id(ws_cell_types, paper_id: str, cell_type: str) -> Optional[str]:
    """在 cell_types sheet 中按 paper_id + cell_type 匹配 ct_id

    当前阶段做论文内的精确匹配（大小写不敏感）。
    后续可扩展为模糊匹配或同义词表。
    """
    headers = [c.value for c in ws_cell_types[1]]
    col_map = {h: i + 1 for i, h in enumerate(headers)}

    # cell_types 的 paper_id 列可能叫 paper_id
    pid_col = col_map.get("paper_id")
    ct_col = col_map.get("cell_type")
    if pid_col is None or ct_col is None:
        return None

    for row in ws_cell_types.iter_rows(min_row=2, max_row=ws_cell_types.max_row,
                                       values_only=False):
        row_paper = row[pid_col - 1].value
        row_ct = row[ct_col - 1].value
        if row_paper == paper_id and row_ct and row_ct.lower() == cell_type.lower():
            return row[col_map["ct_id"] - 1].value
    return None


def normalize_gene_casing(gene: str, species: str) -> str:
    """按物种决定基因名大小写：human 全大写；mouse/rat 首字母大写其余小写。

    species 为空或非 human/mouse/rat 时原样返回（交人工）。
    """
    if not gene:
        return gene
    sp = (species or "").strip().lower()
    if sp == "human":
        return gene.upper()
    if sp in ("mouse", "rat"):
        return gene[:1].upper() + gene[1:].lower()
    return gene


def update_mark_status(ws_cell_types, paper_id: str):
    """更新 cell_types 中 paper_id 对应行的 mark_status"""
    headers = [c.value for c in ws_cell_types[1]]
    col_map = {h: i + 1 for i, h in enumerate(headers)}

    if "mark_status" not in col_map:
        logger.warning("cell_types sheet 缺少 mark_status 列")
        return

    updated = 0
    for row in ws_cell_types.iter_rows(min_row=2, max_row=ws_cell_types.max_row,
                                       values_only=False):
        row_paper = row[col_map["paper_id"] - 1].value
        if row_paper == paper_id:
            current_status = row[col_map["mark_status"] - 1].value
            if current_status in ("old", None):
                row[col_map["mark_status"] - 1].value = "migrated"
                updated += 1
            elif current_status == "migrated":
                pass
            else:
                row[col_map["mark_status"] - 1].value = "mixed"

    logger.info(f"  更新了 {updated} 行的 mark_status → migrated")


def import_review_csv(csv_path: Path, db_path: Path) -> None:
    """将复核通过的 CSV 数据导入 markers sheet"""
    logger.info(f"=" * 60)
    logger.info(f"导入: {csv_path.name}")

    # 读取 CSV
    rows_to_import: list[dict] = []
    try:
        with open(csv_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            missing_columns = REQUIRED_REVIEW_COLUMNS - set(reader.fieldnames or [])
            if missing_columns:
                raise ValueError(f"复核 CSV 缺少新 schema 列: {sorted(missing_columns)}")
            for row in reader:
                status = row.get("review_status", "").strip().lower()
                if status in ALLOWED_STATUSES:
                    evidence_type = row.get("evidence_type", "").strip()
                    candidate_class = row.get("candidate_class", "").strip()
                    marker_polarity = row.get("marker_polarity", "").strip()
                    if evidence_type not in FORMAL_EVIDENCE_TYPES or candidate_class != "formal_candidate":
                        raise ValueError(
                            f"非正式 marker 证据不可批准入库: {row.get('gene_symbol')} ({evidence_type})"
                        )
                    if marker_polarity not in MARKER_POLARITIES:
                        raise ValueError(f"无效 marker_polarity: {marker_polarity!r}")
                    rows_to_import.append(row)
    except FileNotFoundError:
        logger.error(f"文件不存在: {csv_path}")
        return

    if not rows_to_import:
        logger.info("  没有待导入的行（所有行状态都不是 approved/modified）")
        return

    paper_ids = {row["paper_id"].strip() for row in rows_to_import}
    document_ids = {row["document_id"].strip() for row in rows_to_import}
    if len(paper_ids) != 1 or "" in paper_ids:
        raise ValueError(f"单个复核 CSV 必须且只能包含一个 paper_id: {sorted(paper_ids)}")
    if len(document_ids) != 1 or "" in document_ids:
        raise ValueError(f"单个复核 CSV 必须且只能包含一个 document_id: {sorted(document_ids)}")
    paper_id = next(iter(paper_ids))
    document_id = next(iter(document_ids))
    logger.info("  paper_id: %s", paper_id)
    logger.info("  document_id: %s", document_id)

    logger.info(f"  待导入: {len(rows_to_import)} 行")

    # 打开 Excel
    wb = load_workbook(db_path)
    ws_markers = wb["markers"]
    ws_cell_types = wb["cell_types"]

    # 获取下一组 marker_id
    next_marker_id = get_next_marker_id(ws_markers)

    # headers for markers sheet
    marker_headers = [c.value for c in ws_markers[1]]
    marker_col_map = {h: i + 1 for i, h in enumerate(marker_headers)}
    missing_db_columns = REQUIRED_MARKER_COLUMNS - set(marker_col_map)
    if missing_db_columns:
        wb.close()
        raise ValueError(
            "markers sheet 尚未升级到 marker schema v2，缺少列: "
            f"{sorted(missing_db_columns)}。为防止证据丢失，本次拒绝导入。"
        )

    # 幂等去重：同一论文、细胞、亚型、基因和极性只保留一条正式 marker。
    pid_col_idx = marker_col_map.get("paper_id")
    ct_col_idx = marker_col_map.get("cell_type")
    gene_col_idx = marker_col_map.get("gene_symbol")
    subtype_col_idx = marker_col_map.get("subtype_id")
    polarity_col_idx = marker_col_map.get("marker_polarity")
    existing_keys: set[tuple] = set()
    if pid_col_idx and ct_col_idx and gene_col_idx:
        for row in ws_markers.iter_rows(min_row=2, max_row=ws_markers.max_row,
                                        values_only=True):
            row_pid = row[pid_col_idx - 1]
            row_ct = row[ct_col_idx - 1]
            row_gene = row[gene_col_idx - 1]
            if row_pid == paper_id and row_ct and row_gene:
                existing_keys.add((str(row_ct).strip().lower(),
                                   str(row[subtype_col_idx - 1] or "").strip().lower(),
                                   str(row_gene).strip().lower(),
                                   str(row[polarity_col_idx - 1] or "unknown").strip().lower()))

    imported_count = 0
    ct_unmatched = 0
    skipped_dup = 0
    for i, row in enumerate(rows_to_import):
        cell_type = row.get("cell_type", "").strip()
        species = row.get("species", "").strip()
        gene_symbol = row.get("gene_symbol", "").strip()
        # 按物种决定大小写（修复原 .upper() 误伤鼠源基因的 bug）
        gene_symbol_norm = normalize_gene_casing(gene_symbol, species)
        original_symbol = gene_symbol
        evidence_type = row["evidence_type"].strip()
        marker_polarity = row["marker_polarity"].strip()
        candidate_class = row["candidate_class"].strip()
        source_locator = row["source_locator"].strip()
        source_context = row.get("source_context", "").strip()
        review_status = row.get("review_status", "approved").strip()
        notes = row.get("notes", "").strip()

        # 幂等检查：paper_id + cell_type + gene_symbol 已存在则跳过
        subtype = row.get("subtype", "").strip()
        dedup_key = (cell_type.lower(), subtype.lower(), gene_symbol_norm.lower(), marker_polarity.lower())
        if dedup_key in existing_keys:
            skipped_dup += 1
            continue

        # 匹配 ct_id（D4：匹配失败不跳过，ct_id 留空待回填）
        ct_id = match_ct_id(ws_cell_types, paper_id, cell_type)
        if not ct_id:
            ct_unmatched += 1
            if notes:
                notes = notes + "; ct_id待回填"
            else:
                notes = "ct_id待回填"

        # 写入 markers sheet
        new_row_num = ws_markers.max_row + 1
        row_data = {
            "marker_id": next_marker_id,
            "ct_id": ct_id,                   # 可能为 None
            "subtype_id": subtype or None,
            "gene_symbol": gene_symbol_norm,  # 按物种大小写
            "original_symbol": original_symbol,
            "evidence_type": evidence_type,
            "marker_polarity": marker_polarity,
            "candidate_class": candidate_class,
            "source_locator": source_locator,
            "source_context": source_context,
            "review_status": review_status,
            "notes": notes,
            # 新增列（D4 解耦 ct_id）
            "paper_id": paper_id,
            "document_id": document_id,
            "document_role": row["document_role"].strip(),
            "cell_type": cell_type,
            "species": species or "NA",
        }
        for col_name, value in row_data.items():
            col_idx = marker_col_map.get(col_name)
            if col_idx:
                ws_markers.cell(row=new_row_num, column=col_idx, value=value)

        # 登记新写入的 key，防止同批次 CSV 内重复
        existing_keys.add(dedup_key)

        # 自增 ID
        next_marker_id = f"M{int(next_marker_id[1:]) + 1:05d}"
        imported_count += 1

    if skipped_dup:
        logger.info(f"  跳过 {skipped_dup} 条重复 marker（paper_id+cell_type+gene 已存在）")
    if ct_unmatched:
        logger.warning(f"  ⚠️ {ct_unmatched} 条 ct_id 未匹配（已写入 paper_id+cell_type，待回填）")

    # 更新 cell_types 的 mark_status
    update_mark_status(ws_cell_types, paper_id)

    wb.save(db_path)
    wb.close()

    logger.info(f"✅ 导入完成: {imported_count} 条 marker 写入 {db_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="复核数据入库")
    parser.add_argument("csv_path", help="复核 CSV 文件路径")
    parser.add_argument("--db", default=str(DB_PATH),
                        help=f"数据库路径 (默认: {DB_PATH})")
    args = parser.parse_args()

    csv_path = Path(args.csv_path)
    db_path = Path(args.db)

    if not csv_path.exists():
        logger.error(f"CSV 文件不存在: {csv_path}")
        return

    if not db_path.exists():
        logger.error(f"数据库文件不存在: {db_path}")
        return

    try:
        import_review_csv(csv_path, db_path)
    except ValueError as exc:
        logger.error("导入校验失败: %s", exc)


if __name__ == "__main__":
    main()
