"""
convert_gene_symbols.py — HGNC 批量基因名标准化（独立后处理步骤）

功能：
1. 读取 markers sheet 中所有 original_symbol（未被标准化的）
2. 尝试使用 mygene 批量查询 HGNC 官方符号
3. 更新 markers 表的 gene_symbol 列
4. 无法匹配的标记为 'unverified'

使用方式：
    python convert_gene_symbols.py [--db PATH] [--dry-run]

依赖：
    pip install mygene

注意：
    mygene 需要网络连接。如果网络不可用，可以跳过此步骤。
    本脚本可以多次运行——只处理 gene_symbol 为空或等于 original_symbol 的记录。
"""
import argparse
import logging
import re
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

DB_PATH = Path(r"D:\OneDrive\Desktop\组\db\pns-scrna.xlsx")


def collect_unverified_genes(ws_markers) -> list[tuple[int, str]]:
    """收集待标准化的基因

    返回: [(row_number, original_symbol), ...]
    """
    headers = [c.value for c in ws_markers[1]]
    col_map = {h: i + 1 for i, h in enumerate(headers)}

    genes: list[tuple[int, str]] = []
    for row in ws_markers.iter_rows(min_row=2, max_row=ws_markers.max_row,
                                    values_only=False):
        row_num = row[0].row
        original = row[col_map["original_symbol"] - 1].value
        current_gene = row[col_map["gene_symbol"] - 1].value

        if not original or str(original).strip() == "":
            continue
        original = str(original).strip()

        # 跳过已标准化的（gene_symbol 与 original_symbol 不同且不为空）
        if current_gene and str(current_gene).strip() not in ("", original, "unverified"):
            continue

        genes.append((row_num, original))

    return genes


def query_mygene(genes: list[tuple[int, str]]) -> dict[int, str]:
    """使用 mygene 查询 HGNC 标准符号

    返回: {row_number: hgnc_symbol, ...}
    """
    results: dict[int, str] = {}

    try:
        import mygene
    except ImportError:
        logger.error("需要安装 mygene: pip install mygene")
        return results

    mg = mygene.MyGeneInfo()

    # 去重
    unique_genes = sorted(set(g for _, g in genes))
    logger.info(f"  查询 {len(unique_genes)} 个唯一基因符号...")

    # 分批查询（mygene 每次最多 1000 个）
    batch_size = 200
    for i in range(0, len(unique_genes), batch_size):
        batch = unique_genes[i:i + batch_size]
        try:
            # 查询人类基因，返回 HGNC 符号
            query_results = mg.querymany(
                batch,
                scopes="symbol,alias,name",
                fields="symbol,name,taxid",
                species="human",
                returnall=True,
            )
            for item in query_results.get("out", []):
                query_term = item.get("query", "")
                symbol = item.get("symbol", "")
                if symbol and query_term:
                    # 找到匹配的 row_number
                    for row_num, orig in genes:
                        if orig == query_term:
                            results[row_num] = symbol
        except Exception as e:
            logger.warning(f"  批次 {i//batch_size + 1} 查询失败: {e}")

        if i + batch_size < len(unique_genes):
            import time
            time.sleep(0.5)

    logger.info(f"  成功匹配: {len(results)}/{len(genes)}")
    return results


def update_gene_symbols(ws_markers, updates: dict[int, str]) -> int:
    """更新 markers sheet 的 gene_symbol 列"""
    headers = [c.value for c in ws_markers[1]]
    col_map = {h: i + 1 for i, h in enumerate(headers)}
    gene_col = col_map.get("gene_symbol", 4)

    count = 0
    for row_num, hgnc_symbol in updates.items():
        ws_markers.cell(row=row_num, column=gene_col, value=hgnc_symbol)
        count += 1

    return count


def mark_unverified(ws_markers, genes: list[tuple[int, str]],
                    updates: dict[int, str]) -> int:
    """将未匹配的基因标记为 unverified"""
    headers = [c.value for c in ws_markers[1]]
    col_map = {h: i + 1 for i, h in enumerate(headers)}
    gene_col = col_map.get("gene_symbol", 4)

    updated_rows = set(updates.keys())
    count = 0
    for row_num, _ in genes:
        if row_num not in updated_rows:
            ws_markers.cell(row=row_num, column=gene_col, value="unverified")
            count += 1

    return count


def main() -> None:
    parser = argparse.ArgumentParser(description="HGNC 批量基因名标准化")
    parser.add_argument("--db", default=str(DB_PATH),
                        help=f"数据库路径 (默认: {DB_PATH})")
    parser.add_argument("--dry-run", action="store_true",
                        help="仅显示要转换的基因，不写入")
    args = parser.parse_args()

    db_path = Path(args.db)
    if not db_path.exists():
        logger.error(f"数据库文件不存在: {db_path}")
        return

    wb = load_workbook(db_path)
    ws_markers = wb["markers"]

    genes = collect_unverified_genes(ws_markers)
    if not genes:
        logger.info("没有待转换的基因")
        wb.close()
        return

    logger.info(f"待转换: {len(genes)} 条记录")
    if args.dry_run:
        for row_num, gene in genes:
            col = ws_markers.cell(row=row_num, column=1).value
            logger.info(f"  {col}: {gene}")
        wb.close()
        return

    updates = query_mygene(genes)
    n_updated = update_gene_symbols(ws_markers, updates)
    n_unverified = mark_unverified(ws_markers, genes, updates)

    wb.save(db_path)
    wb.close()

    logger.info(f"\n✅ 转换完成:")
    logger.info(f"  已标准化: {n_updated} 条")
    logger.info(f"  未匹配: {n_unverified} 条（标记为 unverified）")


if __name__ == "__main__":
    main()
