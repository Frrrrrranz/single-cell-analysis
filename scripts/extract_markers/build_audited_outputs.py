"""从 40 个终审 audit JSON 生成修正版总表、逐篇 review CSV 和汇总报告。

- 原始 our_markers.xlsx 只读，修正版另存 our_markers_audited.xlsx；
- 属于 40 篇的旧记录按终审结果替换，其余历史行保留并标记 not_in_40_article_audit；
- include 写入 markers sheet；context/exclude/unresolved 写入 audit_exclusions；
- 逐篇生成 <paper_id>_review.csv，另生成 audit_summary.csv 与 full-audit-report.md。
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from marker_schema import EVIDENCE_RANK
from run_full_audit import (
    ALL_EVIDENCE_TYPES,
    FORMAL_EVIDENCE_TYPES,
    NORMALIZATION_STATUSES,
    SPECIES,
    parse_scope_table,
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
LOGGER = logging.getLogger(__name__)

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_AUDIT_DIR = SCRIPT_DIR / "markers_audited"
DEFAULT_SCOPE_FILE = SCRIPT_DIR / "audits" / "task-scope-2026-08-14.md"
DEFAULT_SOURCE_XLSX = SCRIPT_DIR / "our_markers.xlsx"
DEFAULT_OUTPUT_XLSX = SCRIPT_DIR / "our_markers_audited.xlsx"

DECISIONS = {"include", "context_only", "exclude", "unresolved"}
PAPER_STATUSES = {"pass", "corrected", "no_formal_target_marker", "unresolved"}

REVIEW_CSV_HEADERS = [
    "paper_id",
    "task_no",
    "paper_status",
    "cell_type",
    "subtype",
    "species",
    "in_project_scope",
    "original_symbol",
    "normalized_symbol",
    "normalization_status",
    "evidence_type",
    "marker_polarity",
    "source_locator",
    "source_context",
    "citation_match_score",
    "citation_verified",
    "decision",
    "reason",
]

AUDIT_EXCLUSIONS_HEADERS = [
    "paper_id",
    "task_no",
    "cell_type",
    "subtype",
    "species",
    "original_symbol",
    "normalized_symbol",
    "normalization_status",
    "evidence_type",
    "marker_polarity",
    "decision",
    "reason",
    "source_locator",
    "source_context",
    "citation_match_score",
    "citation_verified",
    "audit_model",
    "superseded_marker_id",
]

AUDIT_SUMMARY_HEADERS = [
    "task_no",
    "paper_id",
    "paper_title",
    "task_species",
    "target_cell_scope",
    "tissue",
    "paper_status",
    "marker_records",
    "include",
    "context_only",
    "exclude",
    "unresolved",
    "issues",
    "audit_model",
]

NEW_MARKER_COLUMNS = [
    "audit_status",
    "normalization_status",
    "citation_verified",
    "audit_model",
    "audit_notes",
]

PNS_CELL_PATTERN = re.compile(
    r"schwann|satellite|neuron|gangli|nocicept|sensory|autonomic|sympathetic|"
    r"parasympathetic|enteric|glia|neuroblast",
    re.IGNORECASE,
)
ENDOCRINE_PATTERN = re.compile(r"neuroendocrine|enteroendocrine|pnec", re.IGNORECASE)


def is_pns_cell_value(cell_type: str, target_scope: str) -> str:
    """解剖学 PNS 判断，与 in_project_scope 分离：L4 内分泌层级在项目范围内但非解剖 PNS。"""
    if ENDOCRINE_PATTERN.search(cell_type or ""):
        return "false"
    if PNS_CELL_PATTERN.search(cell_type or ""):
        return "true"
    first_level = (target_scope or "—").split("/")[0].strip()
    if re.match(r"^L[123]:", first_level):
        return "true"
    return "false"


def normalize_gene_casing(gene: str, species: str) -> str:
    if not gene:
        return gene
    sp = (species or "").strip().lower()
    if sp == "human":
        return gene.upper()
    if sp in ("mouse", "rat"):
        return gene[:1].upper() + gene[1:].lower()
    return gene


def evidence_sort_key(marker: dict[str, Any]) -> int:
    return -EVIDENCE_RANK.get(marker.get("evidence_type", ""), 0)


def load_audit_jsons(audit_dir: Path) -> dict[str, dict[str, Any]]:
    audits: dict[str, dict[str, Any]] = {}
    for path in sorted(audit_dir.glob("*_audit.json")):
        data = json.loads(path.read_text(encoding="utf-8"))
        if data.get("paper_id") != path.stem.replace("_audit", ""):
            raise ValueError(f"{path.name} paper_id 与文件名不一致")
        audits[data["paper_id"]] = data
    return audits


def check_audit_invariants(audits: dict[str, dict[str, Any]]) -> None:
    for paper_id, data in audits.items():
        if data.get("paper_status") not in PAPER_STATUSES:
            raise ValueError(f"{paper_id}: 无效 paper_status {data.get('paper_status')!r}")
        for index, marker in enumerate(data.get("markers", [])):
            prefix = f"{paper_id} markers[{index}]"
            if marker.get("species") not in SPECIES:
                raise ValueError(f"{prefix}: 无效 species")
            if marker.get("evidence_type") not in ALL_EVIDENCE_TYPES:
                raise ValueError(f"{prefix}: 无效 evidence_type")
            if marker.get("normalization_status") not in NORMALIZATION_STATUSES:
                raise ValueError(f"{prefix}: 无效 normalization_status")
            if marker.get("decision") not in DECISIONS:
                raise ValueError(f"{prefix}: 无效 decision")
            if marker.get("decision") == "include":
                if marker.get("in_project_scope") is not True:
                    raise ValueError(f"{prefix}: include 但 in_project_scope != true")
                if marker.get("evidence_type") not in FORMAL_EVIDENCE_TYPES:
                    raise ValueError(f"{prefix}: include 但非正式证据")
                if marker.get("normalization_status") not in {"exact", "alias_resolved"}:
                    raise ValueError(f"{prefix}: include 但符号未唯一解析")
                if marker.get("citation_verified") is not True:
                    raise ValueError(f"{prefix}: include 但 citation 未验证")
                if marker.get("species") == "unknown":
                    raise ValueError(f"{prefix}: include 但 species=unknown")


def dedup_includes(audits: dict[str, dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """按计划 5.7 去重键合并 include；返回（主行, 附加证据行）。"""
    grouped: dict[tuple, list[dict[str, Any]]] = {}
    for paper_id, data in sorted(audits.items()):
        for marker in data.get("markers", []):
            if marker.get("decision") != "include":
                continue
            key = (
                paper_id,
                (marker.get("cell_type") or "").strip().lower(),
                (marker.get("subtype") or "").strip().lower(),
                marker.get("species"),
                (marker.get("normalized_symbol") or "").strip(),
                marker.get("marker_polarity"),
            )
            entry = dict(marker)
            entry["paper_id"] = paper_id
            entry["task"] = data.get("task", {})
            grouped.setdefault(key, []).append(entry)

    primary_rows: list[dict[str, Any]] = []
    extra_rows: list[dict[str, Any]] = []
    for key, entries in grouped.items():
        entries.sort(key=evidence_sort_key)
        primary = dict(entries[0])
        others = entries[1:]
        if others:
            extra_locators = "; ".join(
                f"{entry.get('source_locator', '')}" for entry in others
            )
            primary["audit_notes"] = f"同键附加证据 locator: {extra_locators}"
            for entry in others:
                entry["audit_notes"] = "同键去重：证据等级较低，保留在附加证据中"
                extra_rows.append(entry)
        primary_rows.append(primary)
    return primary_rows, extra_rows


def write_review_csvs(audits: dict[str, dict[str, Any]], audit_dir: Path) -> None:
    for paper_id, data in sorted(audits.items()):
        task = data.get("task", {})
        path = audit_dir / f"{paper_id}_review.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as stream:
            writer = csv.DictWriter(stream, fieldnames=REVIEW_CSV_HEADERS, extrasaction="ignore")
            writer.writeheader()
            for marker in data.get("markers", []):
                row = dict(marker)
                row["paper_id"] = paper_id
                row["task_no"] = task.get("task_no")
                row["paper_status"] = data.get("paper_status")
                row["subtype"] = marker.get("subtype") or ""
                row["in_project_scope"] = "true" if marker.get("in_project_scope") else "false"
                writer.writerow(row)
        LOGGER.info("生成 %s (%d 行)", path.name, len(data.get("markers", [])))


def write_audit_summary_csv(audits: dict[str, dict[str, Any]], audit_dir: Path) -> None:
    path = audit_dir / "audit_summary.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=AUDIT_SUMMARY_HEADERS)
        writer.writeheader()
        for paper_id, data in sorted(audits.items(), key=lambda kv: (kv[1].get("task", {}).get("task_no", 0), kv[0])):
            task = data.get("task", {})
            markers = data.get("markers", [])
            decisions = [marker.get("decision") for marker in markers]
            writer.writerow(
                {
                    "task_no": task.get("task_no"),
                    "paper_id": paper_id,
                    "paper_title": task.get("paper_title"),
                    "task_species": task.get("task_species"),
                    "target_cell_scope": task.get("target_cell_scope"),
                    "tissue": task.get("tissue"),
                    "paper_status": data.get("paper_status"),
                    "marker_records": len(markers),
                    "include": decisions.count("include"),
                    "context_only": decisions.count("context_only"),
                    "exclude": decisions.count("exclude"),
                    "unresolved": decisions.count("unresolved"),
                    "issues": len(data.get("issues", [])),
                    "audit_model": data.get("audit_model"),
                }
            )
    LOGGER.info("生成 %s", path.name)


def write_full_audit_report(
    audits: dict[str, dict[str, Any]],
    primary_rows: list[dict[str, Any]],
    removed_rows: list[dict[str, Any]],
    kept_historical: int,
    audit_dir: Path,
) -> None:
    status_counts: dict[str, int] = {}
    for data in audits.values():
        status_counts[data["paper_status"]] = status_counts.get(data["paper_status"], 0) + 1

    lines: list[str] = []
    lines.append("# 40 篇 Marker 全量审核汇总报告（2026-08-30）")
    lines.append("")
    lines.append("审核模型与规则：audit_markers_v1 提示词 + run_full_audit.py 自动降级规则（citation 词元覆盖率阈值 0.72）。")
    lines.append("")
    lines.append("## 文章状态统计")
    lines.append("")
    lines.append(f"- 审核论文数：{len(audits)}")
    for status in ("pass", "corrected", "no_formal_target_marker", "unresolved"):
        lines.append(f"- {status}: {status_counts.get(status, 0)}")
    lines.append(f"- 修正版正式 Marker（include，已去重）：{len(primary_rows)}")
    lines.append(f"- 旧总表移除记录（40 篇范围内未获终审 include）：{len(removed_rows)}")
    lines.append(f"- 旧总表保留历史行（不在 40 篇审计范围）：{kept_historical}")
    lines.append("")
    lines.append("## 逐篇结果")
    lines.append("")
    lines.append("| task | paper_id | 目标范围 | 状态 | include | context | exclude | unresolved |")
    lines.append("| --- | --- | --- | --- | --- | --- | --- | --- |")
    for paper_id, data in sorted(audits.items(), key=lambda kv: (kv[1].get("task", {}).get("task_no", 0), kv[0])):
        task = data.get("task", {})
        markers = data.get("markers", [])
        decisions = [marker.get("decision") for marker in markers]
        scope = task.get("target_cell_scope", "—")
        lines.append(
            f"| {task.get('task_no')} | {paper_id} | {scope} | {data['paper_status']} "
            f"| {decisions.count('include')} | {decisions.count('context_only')} "
            f"| {decisions.count('exclude')} | {decisions.count('unresolved')} |"
        )
    lines.append("")
    lines.append("## 主要问题（issue 汇总）")
    lines.append("")
    issue_types: dict[str, int] = {}
    for data in audits.values():
        for issue in data.get("issues", []):
            issue_types[issue.get("issue_type", "other")] = issue_types.get(issue.get("issue_type", "other"), 0) + 1
    if issue_types:
        lines.append("| issue_type | 数量 |")
        lines.append("| --- | --- |")
        for issue_type, count in sorted(issue_types.items(), key=lambda kv: -kv[1]):
            lines.append(f"| {issue_type} | {count} |")
    else:
        lines.append("无 issue 记录。")
    lines.append("")
    lines.append("## unresolved 条目及原因")
    lines.append("")
    unresolved_lines = 0
    for paper_id, data in sorted(audits.items()):
        for marker in data.get("markers", []):
            if marker.get("decision") == "unresolved":
                lines.append(
                    f"- {paper_id}: {marker.get('normalized_symbol') or marker.get('original_symbol')} "
                    f"({marker.get('cell_type')}) — {marker.get('reason', '')}"
                )
                unresolved_lines += 1
    if not unresolved_lines:
        lines.append("无 unresolved 条目。")
    lines.append("")
    lines.append("原始文件保留位置：`our_markers.xlsx`、`markers_output_v2/`、`review_md/` 均未修改。")
    path = audit_dir / "full-audit-report.md"
    path.write_text("\n".join(lines), encoding="utf-8")
    LOGGER.info("生成 %s", path.name)


def column_index(headers: list[str], name: str) -> int:
    try:
        return headers.index(name) + 1
    except ValueError as exc:
        raise ValueError(f"markers sheet 缺少列: {name}") from exc


def build_audited_workbook(
    audits: dict[str, dict[str, Any]],
    primary_rows: list[dict[str, Any]],
    source_xlsx: Path,
    output_xlsx: Path,
) -> tuple[dict[str, int], list[dict[str, Any]]]:
    wb = load_workbook(source_xlsx)
    ws = wb["markers"]
    headers = [cell.value for cell in ws[1]]
    pid_col = column_index(headers, "paper_id")
    mid_col = column_index(headers, "marker_id")

    audited_ids = set(audits)
    removed_rows: list[dict[str, Any]] = []
    rows_to_delete: list[int] = []
    for row_index in range(2, ws.max_row + 1):
        row_pid = ws.cell(row=row_index, column=pid_col).value
        if row_pid in audited_ids:
            removed_rows.append(
                {
                    "marker_id": ws.cell(row=row_index, column=mid_col).value,
                    "paper_id": row_pid,
                    "cell_type": ws.cell(row=row_index, column=column_index(headers, "cell_type")).value,
                    "gene_symbol": ws.cell(row=row_index, column=column_index(headers, "gene_symbol")).value,
                    "evidence_type": ws.cell(row=row_index, column=column_index(headers, "evidence_type")).value,
                    "review_status": ws.cell(row=row_index, column=column_index(headers, "review_status")).value,
                }
            )
            rows_to_delete.append(row_index)
    for row_index in reversed(rows_to_delete):
        ws.delete_rows(row_index)

    kept_row_ids: list[int] = [
        row_index
        for row_index in range(2, ws.max_row + 1)
        if ws.cell(row=row_index, column=pid_col).value not in audited_ids
    ]

    # 追加新列
    for offset, new_name in enumerate(NEW_MARKER_COLUMNS, start=1):
        ws.cell(row=1, column=len(headers) + offset, value=new_name)
    full_headers = headers + NEW_MARKER_COLUMNS
    ws.freeze_panes = "A2"
    for col_index, header in enumerate(full_headers, start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = min(
            max(len(str(header)) * 1.6, 12), 60
        )

    # 保留历史行打标；新行 marker_id 从最大编号后递增
    max_id = 0
    for row_index in range(2, ws.max_row + 1):
        current = ws.cell(row=row_index, column=mid_col).value
        match = re.match(r"M(\d+)", str(current or ""))
        if match:
            max_id = max(max_id, int(match.group(1)))

    imported_at = datetime(2026, 8, 30)
    for row_index in kept_row_ids:
        ws.cell(row=row_index, column=column_index(full_headers, "audit_status"), value="not_in_40_article_audit")
        ws.cell(row=row_index, column=column_index(full_headers, "audit_notes"), value="旧总表历史行，不在本轮 40 篇终审范围")

    for entry in sorted(primary_rows, key=lambda item: (item.get("task", {}).get("task_no", 0), item["paper_id"])):
        task = entry.get("task", {})
        if not task:
            task = audits[entry["paper_id"]].get("task", {})
        species = entry.get("species", "")
        gene_symbol = normalize_gene_casing(entry.get("normalized_symbol") or entry.get("original_symbol", ""), species)
        row_values = {
            "marker_id": f"M{max_id + 1:05d}",
            "task_no": task.get("task_no"),
            "dataset_id": task.get("dataset_id"),
            "paper_id": entry["paper_id"],
            "document_id": entry["paper_id"],
            "document_role": "primary",
            "ct_id": None,
            "subtype_id": None,
            "cell_type": entry.get("cell_type"),
            "subtype": entry.get("subtype"),
            "species": species,
            "is_pns_cell": is_pns_cell_value(entry.get("cell_type", ""), task.get("target_cell_scope", "")),
            "gene_symbol": gene_symbol,
            "original_symbol": entry.get("original_symbol"),
            "evidence_type": entry.get("evidence_type"),
            "marker_polarity": entry.get("marker_polarity"),
            "candidate_class": "formal_candidate",
            "source_locator": entry.get("source_locator"),
            "source_context": entry.get("source_context"),
            "review_status": "approved",
            "review_method": "full_audit_2026-08-30",
            "notes": entry.get("reason", ""),
            "source_file": f"scripts/extract_markers/markers_audited/{entry['paper_id']}_audit.json",
            "imported_at": imported_at,
            "audit_status": "audited_include",
            "normalization_status": entry.get("normalization_status"),
            "citation_verified": entry.get("citation_verified"),
            "audit_model": entry.get("audit_model") or audits[entry["paper_id"]].get("audit_model"),
            "audit_notes": entry.get("audit_notes", ""),
        }
        max_id += 1
        row_index = ws.max_row + 1
        for col_name in full_headers:
            if col_name in row_values:
                ws.cell(row=row_index, column=column_index(full_headers, col_name), value=row_values[col_name])

    # 说明与统计：追加修正版说明
    if "说明与统计" in wb.sheetnames:
        ws_intro = wb["说明与统计"]
        ws_intro.append([""])
        ws_intro.append(["修正版说明（full audit 2026-08-30）"])
        ws_intro.append([f"40 篇全部终审；修正版正式 marker {len(primary_rows)} 条；旧记录移除 {len(removed_rows)} 条；历史保留 {len(kept_row_ids)} 条。原始 our_markers.xlsx 未修改。"])

    # audit_exclusions sheet
    if "audit_exclusions" in wb.sheetnames:
        del wb["audit_exclusions"]
    ws_ex = wb.create_sheet("audit_exclusions")
    ws_ex.append(AUDIT_EXCLUSIONS_HEADERS)
    superseded_map: dict[str, list[str]] = {}
    for removed in removed_rows:
        superseded_map.setdefault(removed["paper_id"], []).append(f"{removed['marker_id']}:{removed['gene_symbol']}")
    for paper_id, data in sorted(audits.items()):
        task = data.get("task", {})
        for marker in data.get("markers", []):
            if marker.get("decision") == "include":
                continue
            ws_ex.append(
                [
                    paper_id,
                    task.get("task_no"),
                    marker.get("cell_type"),
                    marker.get("subtype"),
                    marker.get("species"),
                    marker.get("original_symbol"),
                    marker.get("normalized_symbol"),
                    marker.get("normalization_status"),
                    marker.get("evidence_type"),
                    marker.get("marker_polarity"),
                    marker.get("decision"),
                    marker.get("reason"),
                    marker.get("source_locator"),
                    marker.get("source_context"),
                    marker.get("citation_match_score"),
                    marker.get("citation_verified"),
                    data.get("audit_model"),
                    None,
                ]
            )
    for removed in removed_rows:
        ws_ex.append(
            [
                removed["paper_id"],
                audits.get(removed["paper_id"], {}).get("task", {}).get("task_no"),
                removed["cell_type"],
                None,
                None,
                removed["gene_symbol"],
                removed["gene_symbol"],
                None,
                removed["evidence_type"],
                None,
                "removed_by_audit",
                "旧总表记录未获 40 篇终审 include，已从修正版移除",
                None,
                None,
                None,
                None,
                None,
                removed["marker_id"],
            ]
        )

    # audit_summary sheet
    if "audit_summary" in wb.sheetnames:
        del wb["audit_summary"]
    ws_sum = wb.create_sheet("audit_summary")
    ws_sum.append(AUDIT_SUMMARY_HEADERS)
    for paper_id, data in sorted(audits.items(), key=lambda kv: (kv[1].get("task", {}).get("task_no", 0), kv[0])):
        task = data.get("task", {})
        markers = data.get("markers", [])
        decisions = [marker.get("decision") for marker in markers]
        ws_sum.append(
            [
                task.get("task_no"),
                paper_id,
                task.get("paper_title"),
                task.get("task_species"),
                task.get("target_cell_scope"),
                task.get("tissue"),
                data.get("paper_status"),
                len(markers),
                decisions.count("include"),
                decisions.count("context_only"),
                decisions.count("exclude"),
                decisions.count("unresolved"),
                len(data.get("issues", [])),
                data.get("audit_model"),
            ]
        )
    for sheet in (ws_ex, ws_sum):
        sheet.freeze_panes = "A2"
        for col_index in range(1, len(sheet[1]) + 1):
            sheet.column_dimensions[get_column_letter(col_index)].width = min(
                max(len(str(sheet.cell(row=1, column=col_index).value)) * 1.6, 12), 45
            )

    # import_log：追加本次导入记录
    if "import_log" in wb.sheetnames:
        ws_log = wb["import_log"]
        ws_log.append(
            [
                "B20260830-AUDIT",
                None,
                "40 papers full audit",
                None,
                None,
                "full_audit",
                len(primary_rows) + len(removed_rows),
                len(primary_rows),
                len(primary_rows),
                len(removed_rows),
                None,
                len(primary_rows),
                "scripts/extract_markers/markers_audited/full-audit-report.md",
                imported_at,
                "40 篇终审：旧记录替换为 include 终审结果；非 40 篇历史行标记 not_in_40_article_audit",
            ]
        )

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    wb.close()
    LOGGER.info("生成 %s（include=%d, removed=%d, kept_historical=%d）", output_xlsx, len(primary_rows), len(removed_rows), len(kept_row_ids))
    return {"include": len(primary_rows), "removed": len(removed_rows), "kept": len(kept_row_ids)}, removed_rows


def main() -> None:
    parser = argparse.ArgumentParser(description="生成 40 篇终审修正版产物")
    parser.add_argument("--audit-dir", type=Path, default=DEFAULT_AUDIT_DIR)
    parser.add_argument("--scope-file", type=Path, default=DEFAULT_SCOPE_FILE)
    parser.add_argument("--source-xlsx", type=Path, default=DEFAULT_SOURCE_XLSX)
    parser.add_argument("--output-xlsx", type=Path, default=DEFAULT_OUTPUT_XLSX)
    parser.add_argument("--skip-csv", action="store_true")
    args = parser.parse_args()

    audits = load_audit_jsons(args.audit_dir)
    if len(audits) != 40:
        raise SystemExit(f"audit JSON 数量不是 40: {len(audits)}")

    task_map = parse_scope_table(args.scope_file)
    missing_scope = [paper_id for paper_id in audits if paper_id not in task_map]
    if missing_scope:
        raise SystemExit(f"缺少任务范围: {missing_scope}")

    check_audit_invariants(audits)
    primary_rows, extra_rows = dedup_includes(audits)
    LOGGER.info("去重后 include %d 条（附加证据 %d 条）", len(primary_rows), len(extra_rows))

    if not args.skip_csv:
        write_review_csvs(audits, args.audit_dir)
        write_audit_summary_csv(audits, args.audit_dir)

    stats, removed = build_audited_workbook(audits, primary_rows, args.source_xlsx, args.output_xlsx)
    write_full_audit_report(
        audits,
        primary_rows,
        removed_rows=removed,
        kept_historical=stats["kept"],
        audit_dir=args.audit_dir,
    )
    LOGGER.info("完成：include=%d removed=%d kept=%d", stats["include"], stats["removed"], stats["kept"])


if __name__ == "__main__":
    main()
