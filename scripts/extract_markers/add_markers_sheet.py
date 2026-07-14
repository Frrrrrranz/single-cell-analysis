"""Task 1: 在 pns-scrna.xlsx 中新增 markers sheet 和 cell_types.mark_status 列"""
import logging
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

DB_PATH = Path(r"D:\OneDrive\Desktop\组\db\pns-scrna.xlsx")

MARKERS_HEADERS = [
    ("marker_id", "自增主键 (M00001+)"),
    ("ct_id", "关联 cell_types.ct_id"),
    ("subtype_id", "可选，关联 cell_subtypes.subtype_id"),
    ("gene_symbol", "HGNC 标准化后的符号"),
    ("original_symbol", "论文原文的基因写法"),
    ("evidence_level", "explicit / implied / inferred / imported"),
    ("source_section", "在论文中的位置 (Table S2 / Fig.1A)"),
    ("source_context", "原文上下文片段"),
    ("review_status", "pending / approved / modified / rejected"),
    ("notes", "复核备注"),
]

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


def add_markers_sheet() -> None:
    wb = load_workbook(DB_PATH)

    # --- Step 1: Add mark_status to cell_types (after markers column) ---
    if "cell_types" in wb.sheetnames:
        ws = wb["cell_types"]
        headers = [c.value for c in ws[1]]
        if "mark_status" not in headers:
            # Insert mark_status column right after 'markers'
            markers_col = headers.index("markers") + 1  # 1-based
            ws.insert_cols(markers_col + 1)
            ws.cell(row=1, column=markers_col + 1, value="mark_status")
            # Style the new header
            cell = ws.cell(row=1, column=markers_col + 1)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            # Fill existing rows with 'old'
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=markers_col + 1, value="old")
            logger.info("Added 'mark_status' column to cell_types sheet")
        else:
            logger.info("'mark_status' column already exists in cell_types")

    # --- Step 2: Create markers sheet ---
    if "markers" not in wb.sheetnames:
        ws_markers = wb.create_sheet("markers")
        col_letters = []
        for col_idx, (header, comment) in enumerate(MARKERS_HEADERS, 1):
            cell = ws_markers.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            col_letters.append(cell.column_letter)
        # Set column widths
        widths = [14, 14, 14, 18, 18, 16, 22, 40, 16, 30]
        for letter, w in zip(col_letters, widths):
            ws_markers.column_dimensions[letter].width = w
        # Freeze header row
        ws_markers.freeze_panes = "A2"
        logger.info("Created 'markers' sheet with headers")
    else:
        logger.info("'markers' sheet already exists")

    wb.save(DB_PATH)
    logger.info(f"Database updated: {DB_PATH}")


if __name__ == "__main__":
    add_markers_sheet()
