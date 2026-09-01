"""门 3：把恢复复核结果合并进总表（B-lite 补充轮，2026-09-01）。

规则：
- 现有 markers sheet 97 行冻结不动（追加前快照、保存后逐行比对）；
- 恢复 include（池 A/B verifications + new_findings）按唯一键去重后追加，
  新增 four_layer_category、recovery_source 两列（现有行留空）；
- audit_exclusions 旧行不动，新增 recovery_outcome 列标注重查结论；
  池 B 与 new_findings 中非 include 记录追加为新的排除行；
- import_log、说明与统计追加本轮记录；
- 产出 recovery_diff.csv 与 recovery-report.md（分项计数）。
"""

from __future__ import annotations

import csv
import json
import logging
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from build_audited_outputs import (
    AUDIT_EXCLUSIONS_HEADERS,
    catalog_cell_layers,
    is_pns_cell_value,
    normalize_gene_casing,
)
from marker_schema import EVIDENCE_RANK
from run_full_audit import FORMAL_EVIDENCE_TYPES

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
LOGGER = logging.getLogger(__name__)

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent
AUDIT_DIR = SCRIPT_DIR / "audited-extraction" / "markers"
RECOVERY_DIR = SCRIPT_DIR / "audited-extraction" / "recovery"
MASTER_XLSX = PROJECT_ROOT / "db" / "cellxgene" / "our_markers.xlsx"

IMPORTED_AT = datetime(2026, 9, 1)
REVIEW_METHOD = "recovery_verify_2026-09-01"

# 与 build_recovery_pool.py 的过滤保持一致
SUBSTANTIVE_PATTERNS = re.compile(
    r"非基因|基因实体|对应错误|不存在|重复|污染|OCR|无法唯一|不可读|不可辨认|"
    r"并非\s*[Mm]arker|不是\s*[Mm]arker|非\s*[Mm]arker|不是正式|无法解析|拼写错误|幻觉"
)

DIFF_HEADERS = [
    "paper_id",
    "source",
    "cell_type",
    "original_symbol",
    "old_decision",
    "gate1_status",
    "new_decision",
    "action",
    "marker_id",
    "reason",
]


def dedup_key(
    paper_id: str,
    cell_type: str,
    subtype: str,
    species: str,
    gene_symbol: str,
    polarity: str,
) -> tuple:
    return (
        paper_id,
        (cell_type or "").strip().lower(),
        (subtype or "").strip().lower(),
        species,
        (gene_symbol or "").strip(),
        polarity or "unknown",
    )


def load_verify_jsons() -> dict[str, dict[str, Any]]:
    verifies: dict[str, dict[str, Any]] = {}
    for path in sorted(RECOVERY_DIR.glob("*_verify.json")):
        data = json.loads(path.read_text(encoding="utf-8"))
        if data.get("paper_id") != path.name.replace("_verify.json", ""):
            raise ValueError(f"{path.name} paper_id 不一致")
        verifies[data["paper_id"]] = data
    if len(verifies) != 40:
        raise SystemExit(f"预期 40 个 verify JSON，实际 {len(verifies)}")
    return verifies


def load_pool(paper_id: str) -> list[dict[str, Any]]:
    """全量池记录（含 duplicate，供排除行映射）；无候选论文返回空列表。"""
    path = RECOVERY_DIR / f"{paper_id}_pool.json"
    if not path.exists():
        return []
    return json.loads(path.read_text(encoding="utf-8"))


def annotate_candidate_indices(pools: dict[str, list[dict[str, Any]]]) -> None:
    """按 run_recovery_verify.load_candidates 的规则就地编号：池顺序、跳过 duplicate。

    池 JSON 落盘时不包含 candidate_index（verify 时才计算），合并前必须用同一规则补齐。
    """
    for records in pools.values():
        index = 0
        for record in records:
            if record.get("gate1_status") in ("duplicate_pool", "duplicate_existing"):
                continue
            record["candidate_index"] = index
            index += 1


def map_exclusion_rows_to_pool(
    audits: dict[str, dict[str, Any]],
    pools: dict[str, list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    """按 build_audited_outputs 写 audit_exclusions 的顺序重放，映射旧行 → 池记录。"""
    mapping: list[dict[str, Any]] = []
    for paper_id, data in sorted(audits.items()):
        pool = pools[paper_id]
        pool_index = 0
        for marker in data.get("markers", []):
            if marker.get("decision") == "include":
                continue
            # 复刻建池过滤
            in_pool = False
            if marker.get("decision") == "exclude":
                in_pool = not SUBSTANTIVE_PATTERNS.search(marker.get("reason") or "")
            else:  # context_only / unresolved
                in_pool = marker.get("evidence_type") in FORMAL_EVIDENCE_TYPES
            record = pool[pool_index] if in_pool else None
            if in_pool:
                pool_index += 1
            mapping.append({"paper_id": paper_id, "marker": marker, "pool_record": record})
        if pool_index != sum(1 for r in pool if r.get("pool") != "B_unaudited"):
            raise ValueError(
                f"{paper_id}: 排除行映射数 {pool_index} != A池记录数 "
                f"{sum(1 for r in pool if r.get('pool') != 'B_unaudited')}"
            )
    return mapping


def collect_includes(
    verifies: dict[str, dict[str, Any]],
    pools: dict[str, list[dict[str, Any]]],
):
    """收集全部 include 候选：verifications + new_findings。

    pools 必须先经 annotate_candidate_indices 编号（与门 2 送检口径一致）。
    """
    entries: list[dict[str, Any]] = []
    for paper_id, data in verifies.items():
        verifications = {
            v.get("candidate_index"): v for v in data.get("verifications", [])
        }
        for record in pools[paper_id]:
            if record.get("gate1_status") in ("duplicate_pool", "duplicate_existing"):
                continue
            v = verifications.get(record.get("candidate_index"))
            if v is None:
                raise ValueError(f"{paper_id}: 候选 {record.get('candidate_index')} 无对应复核")
            if v.get("decision") != "include":
                continue
            entry = dict(v)
            entry["paper_id"] = paper_id
            entry["recovery_source"] = record.get("pool")
            entry["gate1_status"] = record.get("gate1_status")
            entry["old_decision"] = record.get("old_decision")
            entries.append(entry)
        for nf in data.get("new_findings", []):
            if nf.get("decision") != "include":
                continue
            entry = dict(nf)
            entry["paper_id"] = paper_id
            entry["recovery_source"] = "new_finding"
            entry["gate1_status"] = "pass"
            entry["old_decision"] = "never_reported"
            entries.append(entry)
    return entries


def check_include_invariants(entry: dict[str, Any], paper_id: str) -> None:
    if entry.get("evidence_type") not in FORMAL_EVIDENCE_TYPES:
        raise ValueError(f"{paper_id} {entry.get('original_symbol')}: include 但非正式证据")
    if entry.get("normalization_status") not in ("exact", "alias_resolved"):
        raise ValueError(f"{paper_id} {entry.get('original_symbol')}: include 但符号未唯一解析")
    if entry.get("species") == "unknown":
        raise ValueError(f"{paper_id} {entry.get('original_symbol')}: include 但物种未知")
    if entry.get("citation_verified") is not True:
        raise ValueError(f"{paper_id} {entry.get('original_symbol')}: include 但引用未核验")


def snapshot_rows(ws) -> list[tuple]:
    return [tuple(cell.value for cell in row) for row in ws.iter_rows()]


def main() -> None:
    audits = {
        path.name.replace("_audit.json", ""): json.loads(path.read_text(encoding="utf-8"))
        for path in sorted(AUDIT_DIR.glob("*_audit.json"))
    }
    pools = {pid: load_pool(pid) for pid in audits}
    annotate_candidate_indices(pools)
    verifies = load_verify_jsons()

    # ---- 收集与去重 ----
    entries = collect_includes(verifies, pools)
    for entry in entries:
        check_include_invariants(entry, entry.get("paper_id", ""))

    wb = load_workbook(MASTER_XLSX)
    ws = wb["markers"]
    headers = [cell.value for cell in ws[1]]
    col = {name: i for i, name in enumerate(headers)}
    existing_rows = snapshot_rows(ws)
    frozen_count = len(existing_rows) - 1

    existing_keys = set()
    for row in existing_rows[1:]:
        existing_keys.add(
            dedup_key(
                row[col["paper_id"]],
                row[col["cell_type"]],
                row[col.get("subtype", -1)] or "",
                row[col["species"]],
                row[col["gene_symbol"]],
                row[col["marker_polarity"]],
            )
        )

    grouped: dict[tuple, list[dict[str, Any]]] = {}
    for entry in entries:
        gene_symbol = normalize_gene_casing(
            entry.get("normalized_symbol") or entry.get("original_symbol", ""),
            entry.get("species", ""),
        )
        key = dedup_key(
            entry.get("paper_id", ""),
            entry.get("cell_type", ""),
            entry.get("subtype") or "",
            entry.get("species", ""),
            gene_symbol,
            entry.get("marker_polarity"),
        )
        entry["_key"] = key
        entry["_gene_symbol"] = gene_symbol
        grouped.setdefault(key, []).append(entry)

    diff_rows: list[dict[str, str]] = []
    append_entries: list[dict[str, Any]] = []
    action_counts: Counter = Counter()
    for key, group in sorted(grouped.items()):
        group.sort(key=lambda e: -EVIDENCE_RANK.get(e.get("evidence_type", ""), 0))
        primary = group[0]
        if key in existing_keys:
            action_counts["duplicate_of_existing"] += len(group)
            for entry in group:
                diff_rows.append(
                    {
                        "paper_id": entry.get("paper_id", ""),
                        "source": entry.get("recovery_source", ""),
                        "cell_type": entry.get("cell_type", ""),
                        "original_symbol": entry.get("original_symbol", ""),
                        "old_decision": entry.get("old_decision", ""),
                        "gate1_status": entry.get("gate1_status", ""),
                        "new_decision": "include",
                        "action": "duplicate_of_existing（已入表，跳过）",
                        "marker_id": "",
                        "reason": str(entry.get("reason", ""))[:200],
                    }
                )
            continue
        merged_note = ""
        if len(group) > 1:
            locators = "; ".join(
                f"{e.get('recovery_source')}:{e.get('source_locator', '')}" for e in group[1:]
            )
            merged_note = f"同键合并，附加证据: {locators}"
            action_counts["merged_within_pool"] += len(group) - 1
            for entry in group[1:]:
                diff_rows.append(
                    {
                        "paper_id": entry.get("paper_id", ""),
                        "source": entry.get("recovery_source", ""),
                        "cell_type": entry.get("cell_type", ""),
                        "original_symbol": entry.get("original_symbol", ""),
                        "old_decision": entry.get("old_decision", ""),
                        "gate1_status": entry.get("gate1_status", ""),
                        "new_decision": "include",
                        "action": "merged_within_pool（并入主行）",
                        "marker_id": "",
                        "reason": merged_note[:200],
                    }
                )
        primary["_merged_note"] = merged_note
        append_entries.append(primary)
        action_counts["appended"] += 1

    # ---- 追加新列与行 ----
    new_columns = ["four_layer_category", "recovery_source"]
    for offset, name in enumerate(new_columns, start=1):
        ws.cell(row=1, column=len(headers) + offset, value=name)
    full_headers = headers + new_columns
    for col_index, header in enumerate(full_headers, start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = min(
            max(len(str(header)) * 1.6, 12), 60
        )

    max_id = 0
    id_col = col["marker_id"] + 1
    for row_index in range(2, ws.max_row + 1):
        match = re.match(r"M(\d+)", str(ws.cell(row=row_index, column=id_col).value or ""))
        if match:
            max_id = max(max_id, int(match.group(1)))

    def full_col(name: str) -> int:
        return full_headers.index(name) + 1

    for entry in sorted(
        append_entries,
        key=lambda e: ((verifies.get(e.get("paper_id", ""), {}).get("task") or {}).get("task_no", 0), e.get("paper_id", "")),
    ):
        paper_id = entry.get("paper_id", "")
        task = verifies.get(paper_id, {}).get("task") or {}
        max_id += 1
        marker_id = f"M{max_id:05d}"
        notes = entry.get("reason") or ""
        if entry.get("_merged_note"):
            notes = f"{notes}; {entry['_merged_note']}"
        row_values = {
            "marker_id": marker_id,
            "task_no": task.get("task_no"),
            "dataset_id": task.get("dataset_id"),
            "paper_id": paper_id,
            "document_id": paper_id,
            "document_role": "primary",
            "ct_id": None,
            "subtype_id": None,
            "cell_type": entry.get("cell_type"),
            "subtype": entry.get("subtype"),
            "species": entry.get("species"),
            "is_pns_cell": is_pns_cell_value(entry.get("cell_type", ""), catalog_cell_layers(task)),
            "gene_symbol": entry.get("_gene_symbol"),
            "original_symbol": entry.get("original_symbol"),
            "evidence_type": entry.get("evidence_type"),
            "marker_polarity": entry.get("marker_polarity"),
            "candidate_class": "formal_candidate",
            "source_locator": entry.get("source_locator"),
            "source_context": entry.get("source_context"),
            "review_status": "approved",
            "review_method": REVIEW_METHOD,
            "notes": notes,
            "source_file": f"scripts/extract_markers/audited-extraction/recovery/{paper_id}_verify.json",
            "imported_at": IMPORTED_AT,
            "audit_status": "recovery_include",
            "normalization_status": entry.get("normalization_status"),
            "citation_verified": entry.get("citation_verified"),
            "audit_model": verifies.get(paper_id, {}).get("verify_model"),
            "audit_notes": f"恢复轮 2026-09-01；来源 {entry.get('recovery_source')}",
            "four_layer_category": entry.get("four_layer_category"),
            "recovery_source": entry.get("recovery_source"),
        }
        row_index = ws.max_row + 1
        for name, value in row_values.items():
            ws.cell(row=row_index, column=full_col(name), value=value)
        diff_rows.append(
            {
                "paper_id": paper_id,
                "source": entry.get("recovery_source", ""),
                "cell_type": entry.get("cell_type", ""),
                "original_symbol": entry.get("original_symbol", ""),
                "old_decision": entry.get("old_decision", ""),
                "gate1_status": entry.get("gate1_status", ""),
                "new_decision": "include",
                "action": "appended",
                "marker_id": marker_id,
                "reason": str(entry.get("reason", ""))[:200],
            }
        )

    # ---- audit_exclusions：加 recovery_outcome 列 + 追加新排除行 ----
    ws_ex = wb["audit_exclusions"]
    ex_headers = [cell.value for cell in ws_ex[1]]
    if "recovery_outcome" in ex_headers:
        raise SystemExit("audit_exclusions 已有 recovery_outcome 列，疑似重复执行")
    outcome_col = len(ex_headers) + 1
    ws_ex.cell(row=1, column=outcome_col, value="recovery_outcome")
    ex_full_headers = ex_headers + ["recovery_outcome"]

    mapping = map_exclusion_rows_to_pool(audits, pools)
    verifications_by_paper = {
        pid: {v.get("candidate_index"): v for v in data.get("verifications", [])}
        for pid, data in verifies.items()
    }

    # 排除行前段为旧审计非 include 记录（与 mapping 顺序一致），
    # 末尾可能还有 removed_by_audit 的旧总表移除行，不在重查范围
    ex_row_index = 2
    for item in mapping:
        record = item["pool_record"]
        if record is None:
            marker = item["marker"]
            if marker.get("decision") == "exclude":
                outcome = "substantive_exclusion_retained"
            else:
                outcome = "not_rechecked_non_formal_evidence"
        else:
            v = verifications_by_paper.get(item["paper_id"], {}).get(
                record.get("candidate_index")
            )
            if v is None:
                outcome = f"not_sent_{record.get('gate1_status')}"
            else:
                outcome = f"recheck_{v.get('decision')}"
        ws_ex.cell(row=ex_row_index, column=outcome_col, value=outcome)
        ex_row_index += 1
    while ex_row_index <= ws_ex.max_row:
        ws_ex.cell(row=ex_row_index, column=outcome_col, value="not_rechecked_master_history")
        ex_row_index += 1

    # 池 B 与 new_findings 中非 include 记录 → 新排除行
    def ex_append(values: dict[str, Any]) -> None:
        row = [
            values.get("paper_id"),
            values.get("task_no"),
            values.get("cell_type"),
            values.get("subtype"),
            values.get("species"),
            values.get("original_symbol"),
            values.get("normalized_symbol"),
            values.get("normalization_status"),
            values.get("evidence_type"),
            values.get("marker_polarity"),
            values.get("decision"),
            values.get("reason"),
            values.get("source_locator"),
            values.get("source_context"),
            values.get("citation_match_score"),
            values.get("citation_verified"),
            values.get("verify_model"),
            None,
            values.get("recovery_outcome"),
        ]
        ws_ex.append(row)

    for paper_id, data in sorted(verifies.items()):
        task = data.get("task") or {}
        verifications = verifications_by_paper[paper_id]
        for record in pools[paper_id]:
            if record.get("gate1_status") in ("duplicate_pool", "duplicate_existing"):
                continue
            if record.get("pool") != "B_unaudited":
                continue  # A 池旧行已在表中并标 outcome，不重复追加
            v = verifications.get(record.get("candidate_index"))
            if v is None or v.get("decision") == "include":
                continue
            ex_append(
                {
                    **v,
                    "paper_id": paper_id,
                    "task_no": task.get("task_no"),
                    "verify_model": data.get("verify_model"),
                    "recovery_outcome": f"recheck_{v.get('decision')}",
                }
            )
        for nf in data.get("new_findings", []):
            if nf.get("decision") == "include":
                continue
            ex_append(
                {
                    **nf,
                    "paper_id": paper_id,
                    "task_no": task.get("task_no"),
                    "verify_model": data.get("verify_model"),
                    "recovery_outcome": "new_finding_recheck_" + str(nf.get("decision")),
                }
            )

    for col_index in range(1, len(ex_full_headers) + 1):
        ws_ex.column_dimensions[get_column_letter(col_index)].width = min(
            max(len(str(ex_full_headers[col_index - 1])) * 1.6, 12), 45
        )

    # ---- import_log / 说明与统计 ----
    if "import_log" in wb.sheetnames:
        wb["import_log"].append(
            [
                "B20260901-RECOVERY",
                None,
                "40 papers recovery supplement",
                None,
                None,
                "recovery_verify",
                len(append_entries),
                len(append_entries),
                len(append_entries),
                action_counts["duplicate_of_existing"] + action_counts["merged_within_pool"],
                None,
                len(append_entries),
                "scripts/extract_markers/audited-extraction/recovery/recovery-report.md",
                IMPORTED_AT,
                "B-lite 恢复轮：旧范围性排除经三道门复核后追加；现有 97 行冻结未动",
            ]
        )
    if "说明与统计" in wb.sheetnames:
        wb["说明与统计"].append([""])
        wb["说明与统计"].append(["恢复轮补充说明（recovery 2026-09-01）"])
        wb["说明与统计"].append(
            [
                f"B-lite 恢复轮：现有 97 行冻结未动；追加恢复/新发现 include "
                f"{len(append_entries)} 条（池 A/B 恢复与全文漏提扫描）；"
                f"去重跳过 {action_counts['duplicate_of_existing']} 条、同键合并 "
                f"{action_counts['merged_within_pool']} 条。规则见 MARKER_POLICY.md 与 "
                ".agents/plan/marker-full-supplement-blite-2026-09-01.md。"
            ]
        )

    MASTER_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(MASTER_XLSX)
    LOGGER.info(
        "总表更新：追加 %d 行（duplicate_of_existing=%d, merged=%d）",
        len(append_entries),
        action_counts["duplicate_of_existing"],
        action_counts["merged_within_pool"],
    )

    # ---- 保存后校验：现有行零变动（只比较原始列宽，新增列不计入） ----
    wb_check = load_workbook(MASTER_XLSX, read_only=True)
    ws_check = wb_check["markers"]
    after_rows = [tuple(row) for row in ws_check.iter_rows(values_only=True)]
    wb_check.close()
    orig_width = len(headers)
    # after_rows 含表头行，取 [1, frozen_count+1) 才是与 existing_rows[1:] 对应的冻结数据行
    if [r[:orig_width] for r in after_rows[1 : frozen_count + 1]] != [
        r[:orig_width] for r in existing_rows[1:]
    ]:
        raise SystemExit("校验失败：现有 97 行内容发生变动")
    LOGGER.info("冻结校验通过：现有 %d 行内容与追加前逐行一致", frozen_count)

    baseline_path = RECOVERY_DIR / "frozen_baseline.json"
    baseline_path.write_text(
        json.dumps(
            {
                "frozen_count": frozen_count,
                "headers": headers,
                "rows": [
                    [cell if cell is None or isinstance(cell, (str, int, float, bool, datetime)) else str(cell) for cell in row]
                    for row in existing_rows[1:]
                ],
            },
            ensure_ascii=False,
            indent=2,
            default=str,
        ),
        encoding="utf-8",
    )

    # ---- diff CSV ----
    diff_path = RECOVERY_DIR / "recovery_diff.csv"
    with diff_path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=DIFF_HEADERS)
        writer.writeheader()
        writer.writerows(diff_rows)
    LOGGER.info("生成 %s（%d 行）", diff_path.name, len(diff_rows))

    # ---- 报告 ----
    source_counts = Counter(e.get("recovery_source") for e in append_entries)
    verify_decision_counts: Counter = Counter()
    for data in verifies.values():
        for v in data.get("verifications", []):
            verify_decision_counts[v.get("decision")] += 1
    new_finding_counts: Counter = Counter()
    for data in verifies.values():
        for nf in data.get("new_findings", []):
            new_finding_counts[nf.get("decision")] += 1

    lines = [
        "# B-lite 恢复轮汇总报告（2026-09-01）",
        "",
        "口径：MARKER_POLICY.md（全部有可靠证据的 Marker 保留）+ "
        "同学规范双重门槛；现有 97 行冻结未动。",
        "",
        "## 分项计数",
        "",
        f"- 现有冻结行：{frozen_count}（audited_include 87 + 历史 10）",
        f"- 本轮追加 include：{len(append_entries)}",
        f"  - 池 A（旧范围性排除恢复）：{source_counts.get('A_exclude', 0)}",
        f"  - 池 A（旧降级记录恢复）：{source_counts.get('A_downgraded', 0)}",
        f"  - 池 B（从未送审候选恢复）：{source_counts.get('B_unaudited', 0)}",
        f"  - 全文漏提扫描新发现：{source_counts.get('new_finding', 0)}",
        f"- 去重跳过（已入表）：{action_counts['duplicate_of_existing']}",
        f"- 同键合并：{action_counts['merged_within_pool']}",
        f"- 池候选复核判定分布：{dict(verify_decision_counts)}",
        f"- 新发现判定分布：{dict(new_finding_counts)}",
        "",
        "## 逐篇新旧数量对照",
        "",
        "| paper_id | 旧 include | 追加恢复 | 追加新发现 | 新合计 |",
        "| --- | ---: | ---: | ---: | ---: |",
    ]
    per_paper_appended: Counter = Counter()
    per_paper_new: Counter = Counter()
    for entry in append_entries:
        if entry.get("recovery_source") == "new_finding":
            per_paper_new[entry.get("paper_id", "")] += 1
        else:
            per_paper_appended[entry.get("paper_id", "")] += 1
    for paper_id in sorted(audits):
        old_include = sum(
            1 for m in audits[paper_id].get("markers", []) if m.get("decision") == "include"
        )
        lines.append(
            f"| {paper_id} | {old_include} | {per_paper_appended.get(paper_id, 0)} "
            f"| {per_paper_new.get(paper_id, 0)} "
            f"| {old_include + per_paper_appended.get(paper_id, 0) + per_paper_new.get(paper_id, 0)} |"
        )
    lines += [
        "",
        "明细见 recovery_diff.csv；逐篇复核 JSON 在本目录；总表已更新 db/cellxgene/our_markers.xlsx。",
    ]
    report_path = RECOVERY_DIR / "recovery-report.md"
    report_path.write_text("\n".join(lines), encoding="utf-8")
    LOGGER.info("生成 %s", report_path.name)


if __name__ == "__main__":
    main()
