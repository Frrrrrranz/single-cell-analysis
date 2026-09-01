"""生成当前修正版离线汇总页 audited-extraction/marker-summary.html。

数据源：
- 终审 audit JSON：文章状态、非 include 候选（叠加恢复轮复核后的新判定）；
- marker提取/表单/our_markers.xlsx：正式 Marker 总表（含恢复追加行与来源列）；
- 恢复轮 verify/pool JSON：B 池与新发现的非 include 记录、来源统计。

自包含单文件（内嵌 JSON + 原生 JS 筛选）。
"""

from __future__ import annotations

import argparse
import json
import logging
from collections import Counter
from pathlib import Path
from typing import Any

from marker_schema import EVIDENCE_RANK

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
LOGGER = logging.getLogger(__name__)

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
MARKER_DIR = PROJECT_ROOT / "marker提取"
DEFAULT_AUDIT_DIR = MARKER_DIR / "audited-extraction" / "markers"
DEFAULT_RECOVERY_DIR = MARKER_DIR / "audited-extraction" / "recovery"
DEFAULT_MASTER_XLSX = MARKER_DIR / "表单" / "our_markers.xlsx"
DEFAULT_OUTPUT = MARKER_DIR / "audited-extraction" / "marker-summary.html"

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Marker 全量终审汇总</title>
<style>
:root {
  --bg: #f5f7fa; --card: #ffffff; --ink: #2c3440; --muted: #6b7686;
  --line: #e3e8ef; --accent: #4a7ab5; --accent-soft: #eaf1f8;
  --ok: #3d8f6e; --warn: #b3813a; --bad: #b0563f;
  --chip: #f0f3f7; --chip-active: #4a7ab5;
}
* { box-sizing: border-box; }
body { margin: 0; background: var(--bg); color: var(--ink);
  font: 14px/1.6 "Segoe UI", "Microsoft YaHei", system-ui, sans-serif; }
.wrap { max-width: 1240px; margin: 0 auto; padding: 28px 20px 60px; }
h1 { font-size: 22px; margin: 0 0 4px; }
.sub { color: var(--muted); font-size: 13px; margin-bottom: 22px; }
.grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 12px; margin-bottom: 14px; }
.stat { background: var(--card); border: 1px solid var(--line); border-radius: 10px; padding: 14px 16px; }
.stat .num { font-size: 26px; font-weight: 600; color: var(--accent); }
.stat .label { font-size: 12px; color: var(--muted); margin-top: 2px; }
.panel { background: var(--card); border: 1px solid var(--line); border-radius: 12px; padding: 18px 20px; margin-top: 16px; }
.panel h2 { font-size: 16px; margin: 0 0 10px; }
.bar-row { display: flex; align-items: center; gap: 10px; margin: 6px 0; font-size: 13px; }
.bar-label { width: 220px; color: var(--muted); text-align: right; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.bar { height: 12px; border-radius: 6px; background: linear-gradient(90deg, #7ba3cc, var(--accent)); min-width: 2px; }
.bar-val { color: var(--muted); font-size: 12px; }
#dist-bars { max-height: 430px; overflow: auto; }
.filters { display: flex; flex-wrap: wrap; gap: 8px; align-items: center; margin-bottom: 12px; }
select, input { background: var(--chip); border: 1px solid var(--line); border-radius: 8px;
  padding: 6px 10px; color: var(--ink); font-size: 13px; }
input { min-width: 220px; }
table { width: 100%; border-collapse: collapse; font-size: 12.5px; }
th { position: sticky; top: 0; background: var(--card); text-align: left; padding: 8px 10px;
  border-bottom: 2px solid var(--line); color: var(--muted); font-weight: 600; white-space: nowrap; }
td { padding: 7px 10px; border-bottom: 1px solid var(--line); vertical-align: top; }
tr:hover td { background: var(--accent-soft); }
.tag { display: inline-block; border-radius: 999px; padding: 1px 9px; font-size: 11px; white-space: nowrap; }
.tag.include { background: #e4f1eb; color: var(--ok); }
.tag.context_only { background: #f4eedd; color: var(--warn); }
.tag.exclude { background: #f3e6e1; color: var(--bad); }
.tag.unresolved { background: #eceded; color: #61666d; }
.src { display: inline-block; border-radius: 999px; padding: 1px 8px; font-size: 11px; background: var(--accent-soft); color: var(--accent); white-space: nowrap; }
.src.old { background: #eef0f3; color: var(--muted); }
.mono { font-family: Consolas, "Courier New", monospace; }
.ctx { max-width: 460px; color: var(--muted); font-size: 12px; }
.count { color: var(--muted); font-size: 12px; margin: 4px 0 10px; }
.tabs { display: flex; gap: 8px; margin-bottom: 12px; }
.tab { border: 1px solid var(--line); background: var(--chip); border-radius: 8px;
  padding: 5px 14px; cursor: pointer; font-size: 13px; color: var(--ink); }
.tab.active { background: var(--chip-active); color: #fff; border-color: var(--chip-active); }
.note { color: var(--muted); font-size: 12px; margin-top: 14px; }
.papers { max-height: 560px; overflow: auto; }
</style>
</head>
<body>
<div class="wrap">
  <h1>Marker 全量终审汇总</h1>
  <div class="sub">当前全部论文 · 正式 Marker 来自 marker提取/表单/our_markers.xlsx · 非 include 视图为恢复轮复核后判定</div>
  <div id="stats" class="grid"></div>

  <div class="panel">
    <h2>文章状态</h2>
    <div id="status-bars"></div>
  </div>
  <div class="panel">
    <h2>正式 Marker 分布（现总表全部行）</h2>
    <div id="dist-bars"></div>
  </div>

  <div class="panel">
    <h2>逐篇结果</h2>
    <div class="papers"><table id="papers-table"></table></div>
  </div>

  <div class="panel">
    <h2>Marker 明细</h2>
    <div class="tabs" id="decision-tabs"></div>
    <div class="filters">
      <select id="f-species"><option value="">全部物种</option></select>
      <select id="f-evidence"><option value="">全部证据</option></select>
      <select id="f-paper"><option value="">全部论文</option></select>
      <select id="f-source"><option value="">全部来源</option></select>
      <input id="f-search" placeholder="搜索基因 / 细胞类型 / 基因符号…">
    </div>
    <div class="count" id="row-count"></div>
    <div class="papers"><table id="markers-table"></table></div>
    <div class="note">include 视图为现总表（终审 97 行冻结 + 恢复轮追加）；unresolved/exclude/context_only 视图含旧终审候选（叠加复核新判定）、B 池与新发现中未收录的记录，均不进入正式总表。</div>
  </div>
</div>
<script id="data" type="application/json">__DATA__</script>
<script>
const DATA = JSON.parse(document.getElementById('data').textContent);
const ALL_DECISIONS = ['include', 'context_only', 'exclude', 'unresolved'];
let activeDecision = 'include';

function el(tag, cls, text) {
  const node = document.createElement(tag);
  if (cls) node.className = cls;
  if (text !== undefined) node.textContent = text;
  return node;
}

function fmt(n) { return n.toLocaleString('zh-CN'); }

(function renderStats() {
  const stats = DATA.stats;
  const items = [
    ['审核论文', stats.papers],
    ['正式 Marker（现总表）', stats.include],
    ['其中恢复轮追加', stats.recovery],
    ['涉及论文', stats.papers_with_include],
    ['corrected', stats.status.corrected],
    ['pass', stats.status.pass],
    ['no_formal_marker', (stats.status.no_formal_marker || 0) + (stats.status.no_formal_target_marker || 0)],
    ['unresolved 文章', stats.status.unresolved],
    ['未收录候选', stats.excluded_records],
  ];
  const root = document.getElementById('stats');
  items.forEach(([label, value]) => {
    const card = el('div', 'stat');
    card.append(el('div', 'num', fmt(value)), el('div', 'label', label));
    root.append(card);
  });
})();

function renderBars(containerId, rows, total) {
  const root = document.getElementById(containerId);
  root.innerHTML = '';
  const maxVal = Math.max(...rows.map(r => r[1]), 1);
  rows.forEach(([label, value]) => {
    const row = el('div', 'bar-row');
    const bar = el('div', 'bar');
    bar.style.width = Math.max((value / maxVal) * 640, 2) + 'px';
    row.append(el('div', 'bar-label', label + '（' + value + '）'), bar, el('div', 'bar-val', (total ? (value / total * 100).toFixed(1) + '%' : '')));
    root.append(row);
  });
}

renderBars('status-bars', Object.entries(DATA.stats.status));
renderBars('dist-bars', [
  ...Object.entries(DATA.stats.by_source).map(([k, v]) => ['来源: ' + k, v]),
  ...Object.entries(DATA.stats.by_cell_type).sort((a, b) => b[1] - a[1]).map(([k, v]) => ['细胞: ' + k, v]),
  ...Object.entries(DATA.stats.by_species).map(([k, v]) => ['物种: ' + k, v]),
  ...Object.entries(DATA.stats.by_evidence).map(([k, v]) => ['证据: ' + k, v]),
], DATA.stats.include);

(function renderPapers() {
  const table = document.getElementById('papers-table');
  table.innerHTML = '';
  const head = el('tr');
  ['task', 'paper_id', '目标范围', '物种', '状态', 'include(旧)', '+恢复', '+新发现', '新合计', 'context', 'exclude', 'unresolved']
    .forEach(h => head.append(el('th', null, h)));
  const thead = el('thead'); thead.append(head); table.append(thead);
  const tbody = el('tbody');
  DATA.papers.forEach(p => {
    const row = el('tr');
    [p.task_no, p.paper_id, p.scope, p.species, p.status, p.include_old, p.recovery, p.new_finding, p.total, p.context_only, p.exclude, p.unresolved]
      .forEach(v => row.append(el('td', 'mono', v)));
    tbody.append(row);
  });
  table.append(tbody);
})();

(function renderTabs() {
  const root = document.getElementById('decision-tabs');
  ALL_DECISIONS.forEach(decision => {
    const count = DATA.markers.filter(m => m.decision === decision).length;
    const tab = el('div', 'tab' + (decision === activeDecision ? ' active' : ''), decision + '（' + count + '）');
    tab.onclick = () => {
      activeDecision = decision;
      root.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
      tab.classList.add('active');
      renderMarkers();
    };
    root.append(tab);
  });
})();

(function fillFilters() {
  const unique = (values) => [...new Set(values.filter(Boolean))].sort();
  [['f-species', unique(DATA.markers.map(m => m.species))],
   ['f-evidence', unique(DATA.markers.map(m => m.evidence_type))],
   ['f-paper', unique(DATA.markers.map(m => m.paper_id))],
   ['f-source', unique(DATA.markers.map(m => m.source))]].forEach(([id, values]) => {
    const select = document.getElementById(id);
    values.forEach(v => {
      const option = el('option', null, v);
      option.value = v;
      select.append(option);
    });
    select.onchange = renderMarkers;
  });
  document.getElementById('f-search').oninput = renderMarkers;
})();

function renderMarkers() {
  const species = document.getElementById('f-species').value;
  const evidence = document.getElementById('f-evidence').value;
  const paper = document.getElementById('f-paper').value;
  const source = document.getElementById('f-source').value;
  const query = document.getElementById('f-search').value.trim().toLowerCase();
  const rows = DATA.markers.filter(m => {
    if (m.decision !== activeDecision) return false;
    if (species && m.species !== species) return false;
    if (evidence && m.evidence_type !== evidence) return false;
    if (paper && m.paper_id !== paper) return false;
    if (source && m.source !== source) return false;
    if (query && !(`${m.symbol} ${m.cell_type} ${m.paper_id}`.toLowerCase().includes(query))) return false;
    return true;
  });
  document.getElementById('row-count').textContent = '共 ' + rows.length + ' 条';
  const table = document.getElementById('markers-table');
  table.innerHTML = '';
  const head = el('tr');
  ['paper_id', '细胞类型', '亚型', '物种', '基因', '证据类型', '极性', 'decision', '来源', 'citation', '定位', '原文证据'].forEach(h => head.append(el('th', null, h)));
  const thead = el('thead'); thead.append(head); table.append(thead);
  const tbody = el('tbody');
  rows.forEach(m => {
    const row = el('tr');
    row.append(el('td', 'mono', m.paper_id));
    row.append(el('td', null, m.cell_type));
    row.append(el('td', null, m.subtype || '—'));
    row.append(el('td', null, m.species));
    row.append(el('td', 'mono', m.symbol));
    row.append(el('td', null, m.evidence_type));
    row.append(el('td', null, m.polarity));
    const decision = el('td');
    decision.append(el('span', 'tag ' + m.decision, m.decision));
    row.append(decision);
    const src = el('td');
    src.append(el('span', 'src' + (m.decision === 'include' && !m.recovered ? ' old' : ''), m.source || '—'));
    row.append(src);
    row.append(el('td', null, m.citation === true ? '✓ ' + (m.score ?? '') : (m.score != null ? '✗ ' + m.score : '')));
    row.append(el('td', 'mono', m.locator || ''));
    row.append(el('td', 'ctx', m.context || ''));
    tbody.append(row);
  });
  table.append(tbody);
}
renderMarkers();
</script>
</body>
</html>
"""

SOURCE_LABELS = {
    "audited_include": "终审(2026-08-30)",
    "recovery_include": "恢复轮",
    "not_in_40_article_audit": "历史行",
}
RECOVERY_SOURCE_LABELS = {
    "A_exclude": "恢复·旧范围排除",
    "A_downgraded": "恢复·旧降级",
    "B_unaudited": "恢复·B池",
    "new_finding": "新发现",
}


def marker_row_from_master(row: dict[str, Any]) -> dict[str, Any]:
    source = row.get("recovery_source")
    label = RECOVERY_SOURCE_LABELS.get(source or "", SOURCE_LABELS.get(row.get("audit_status"), "终审"))
    return {
        "paper_id": row.get("paper_id"),
        "cell_type": row.get("cell_type"),
        "subtype": row.get("subtype"),
        "species": row.get("species"),
        "symbol": row.get("gene_symbol"),
        "original_symbol": row.get("original_symbol"),
        "evidence_type": row.get("evidence_type"),
        "polarity": row.get("marker_polarity"),
        "decision": "include",
        "citation": row.get("citation_verified"),
        "score": None,
        "locator": row.get("source_locator"),
        "context": row.get("source_context"),
        "source": label,
        "recovered": source is not None,
    }


def marker_row_from_record(
    record: dict[str, Any],
    paper_id: str,
    source_label: str,
) -> dict[str, Any]:
    return {
        "paper_id": paper_id,
        "cell_type": record.get("cell_type"),
        "subtype": record.get("subtype"),
        "species": record.get("species"),
        "symbol": record.get("normalized_symbol") or record.get("original_symbol"),
        "original_symbol": record.get("original_symbol"),
        "evidence_type": record.get("evidence_type"),
        "polarity": record.get("marker_polarity"),
        "decision": record.get("decision"),
        "citation": record.get("citation_verified"),
        "score": record.get("citation_match_score"),
        "locator": record.get("source_locator"),
        "context": record.get("source_context"),
        "source": source_label,
        "recovered": False,
    }


def load_master_rows(master_xlsx: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(master_xlsx, read_only=True)
    ws = wb["markers"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = list(rows[0])
    col = {name: i for i, name in enumerate(headers)}
    data = []
    for values in rows[1:]:
        if not values[col["marker_id"]]:
            continue
        data.append({name: values[i] for name, i in col.items()})
    return data


def load_recovery_overlay(
    audits: dict[str, dict[str, Any]],
    recovery_dir: Path,
):
    """返回 (paper→非include旧判定覆盖列表, B池与新发现的非include记录, verifies, pools)。"""
    from apply_recovery import (
        annotate_candidate_indices,
        load_pool,
        map_exclusion_rows_to_pool,
    )

    pools = {pid: load_pool(pid) for pid in audits}
    annotate_candidate_indices(pools)
    verifies = {}
    for path in sorted(recovery_dir.glob("*_verify.json")):
        data = json.loads(path.read_text(encoding="utf-8"))
        verifies[data["paper_id"]] = data
    if len(verifies) != len(audits):
        LOGGER.warning("verify JSON %d 篇 != 论文 %d 篇，恢复视图可能不完整", len(verifies), len(audits))

    verifications_by_paper = {
        pid: {v.get("candidate_index"): v for v in data.get("verifications", [])}
        for pid, data in verifies.items()
    }
    mapping = map_exclusion_rows_to_pool(audits, pools)

    paper_outcomes: dict[str, list[str | None]] = {}
    for item in mapping:
        paper_id = item["paper_id"]
        record = item["pool_record"]
        if record is None:
            outcome = None
        else:
            v = verifications_by_paper.get(paper_id, {}).get(record.get("candidate_index"))
            outcome = v.get("decision") if v else None
        paper_outcomes.setdefault(paper_id, []).append(outcome)

    extra_rows: list[tuple[str, dict[str, Any], str]] = []
    for paper_id in sorted(verifies):
        verifications = verifications_by_paper[paper_id]
        for record in pools[paper_id]:
            if record.get("gate1_status") in ("duplicate_pool", "duplicate_existing"):
                continue
            if record.get("pool") != "B_unaudited":
                continue
            v = verifications.get(record.get("candidate_index"))
            if v is None or v.get("decision") == "include":
                continue
            extra_rows.append((paper_id, v, "B池·未收录"))
        for nf in verifies[paper_id].get("new_findings", []):
            if nf.get("decision") == "include":
                continue
            extra_rows.append((paper_id, nf, "新发现·未收录"))
    return paper_outcomes, extra_rows


def build_payload(
    audit_dir: Path,
    master_xlsx: Path,
    recovery_dir: Path,
) -> dict[str, Any]:
    audits: dict[str, dict[str, Any]] = {}
    for path in sorted(audit_dir.glob("*_audit.json")):
        data = json.loads(path.read_text(encoding="utf-8"))
        audits[data["paper_id"]] = data

    master_rows = load_master_rows(master_xlsx)
    include_rows = [marker_row_from_master(r) for r in master_rows]
    LOGGER.info("总表 include 行 %d（恢复追加 %d）", len(include_rows), sum(1 for r in include_rows if r["recovered"]))

    paper_outcomes, extra_rows = load_recovery_overlay(audits, recovery_dir)

    status_counts: dict[str, int] = {"pass": 0, "corrected": 0, "no_formal_marker": 0, "no_formal_target_marker": 0, "unresolved": 0}
    by_species: dict[str, int] = {}
    by_evidence: dict[str, int] = {}
    by_cell_type: dict[str, int] = {}
    by_source: dict[str, int] = {}
    papers: list[dict[str, Any]] = []

    master_by_paper: dict[str, dict[str, int]] = {}
    for row in master_rows:
        entry = master_by_paper.setdefault(row["paper_id"], {"old": 0, "recovery": 0, "new_finding": 0})
        if row.get("audit_status") == "recovery_include":
            if row.get("recovery_source") == "new_finding":
                entry["new_finding"] += 1
            else:
                entry["recovery"] += 1
        else:
            entry["old"] += 1

    excluded_markers: list[dict[str, Any]] = []
    paper_excluded_counts: dict[str, dict[str, int]] = {}

    def count_excluded(paper_id: str, decision: str) -> None:
        entry = paper_excluded_counts.setdefault(paper_id, {"context_only": 0, "exclude": 0, "unresolved": 0})
        entry[decision] = entry.get(decision, 0) + 1

    for paper_id, data in sorted(audits.items(), key=lambda kv: (kv[1].get("task", {}).get("task_no", 0), kv[0])):
        task = data.get("task", {})
        status_counts[data["paper_status"]] = status_counts.get(data["paper_status"], 0) + 1
        outcomes = paper_outcomes.get(paper_id, [])
        outcome_pos = 0
        for marker in data.get("markers", []):
            if marker.get("decision") == "include":
                continue
            new_decision = outcomes[outcome_pos] if outcome_pos < len(outcomes) else None
            outcome_pos += 1
            if new_decision == "include":
                continue  # 已在 include 视图（总表恢复行）
            decision = new_decision or marker.get("decision")
            row = marker_row_from_record(marker, paper_id, "旧候选·复核后维持" if new_decision else "旧候选")
            row["decision"] = decision
            excluded_markers.append(row)
            count_excluded(paper_id, decision)

        entry = master_by_paper.get(paper_id, {"old": 0, "recovery": 0, "new_finding": 0})
        excluded = paper_excluded_counts.get(paper_id, {"context_only": 0, "exclude": 0, "unresolved": 0})
        papers.append(
            {
                "task_no": task.get("task_no"),
                "paper_id": paper_id,
                "scope": task.get("catalog_cell_layers") or task.get("target_cell_scope") or "—",
                "species": task.get("task_species"),
                "status": data["paper_status"],
                "include_old": entry["old"],
                "recovery": entry["recovery"],
                "new_finding": entry["new_finding"],
                "total": entry["old"] + entry["recovery"] + entry["new_finding"],
                "context_only": excluded["context_only"],
                "exclude": excluded["exclude"],
                "unresolved": excluded["unresolved"],
            }
        )

    for paper_id, record, label in extra_rows:
        row = marker_row_from_record(record, paper_id, label)
        excluded_markers.append(row)
        count_excluded(paper_id, row["decision"])

    for row in include_rows:
        by_species[row["species"] or "unknown"] = by_species.get(row["species"] or "unknown", 0) + 1
        by_evidence[row["evidence_type"] or "unknown"] = by_evidence.get(row["evidence_type"] or "unknown", 0) + 1
        by_cell_type[row["cell_type"] or "未知"] = by_cell_type.get(row["cell_type"] or "未知", 0) + 1
        by_source[row["source"]] = by_source.get(row["source"], 0) + 1

    papers_with_include = sum(1 for p in papers if p["total"] > 0)
    return {
        "stats": {
            "papers": len(audits),
            "include": len(include_rows),
            "recovery": sum(1 for r in include_rows if r["recovered"]),
            "papers_with_include": papers_with_include,
            "excluded_records": len(excluded_markers),
            "status": status_counts,
            "by_species": by_species,
            "by_evidence": by_evidence,
            "by_cell_type": by_cell_type,
            "by_source": by_source,
        },
        "papers": papers,
        "markers": include_rows + excluded_markers,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="生成修正版 HTML 汇总页")
    parser.add_argument("--audit-dir", type=Path, default=DEFAULT_AUDIT_DIR)
    parser.add_argument("--master-xlsx", type=Path, default=DEFAULT_MASTER_XLSX)
    parser.add_argument("--recovery-dir", type=Path, default=DEFAULT_RECOVERY_DIR)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    payload = build_payload(args.audit_dir, args.master_xlsx, args.recovery_dir)
    if payload["stats"]["papers"] == 0:
        raise SystemExit("未找到 audit JSON")
    html = HTML_TEMPLATE.replace("__DATA__", json.dumps(payload, ensure_ascii=False))
    args.output.write_text(html, encoding="utf-8")
    LOGGER.info(
        "生成 %s（include=%d，其中恢复追加=%d，论文=%d）",
        args.output,
        payload["stats"]["include"],
        payload["stats"]["recovery"],
        payload["stats"]["papers"],
    )


if __name__ == "__main__":
    main()
