"""构建 40 篇论文的完整复核计划（2026-09-02 复核轮）。

四道机械检查门 + 人工清单：
- Gate A 漏提候选：review_md 中含作者 marker 措辞的句子里提取类基因 token，
  与该篇总表行比对，未收录者列为候选（词表命中优先）。
- Gate B 证据类型候选：annotation_marker/figure_labeled 行的 source_context
  已含明确 marker 措辞 → 升级 author_declared 候选；context_only 行含 marker
  措辞 → 身份恢复候选。
- Gate C 语义重复：同篇同基因 + cell_type 规范化后相近 → 合并候选。
- Gate D 人工清单：聚类对账、归属、物种、基因写法、跨篇一致性。

输出：marker提取/audits/recheck-2026-09-02/recheck_plan.xlsx
"""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(r"d:\OneDrive\Desktop\组\marker提取")
XLSX = ROOT / "表单" / "our_markers.xlsx"
MD_DIR = ROOT / "review_md"
OUT_DIR = ROOT / "audits" / "recheck-2026-09-02"

MARKER_LANGUAGE = re.compile(
    r"\bmarkers?\b|\bmarked\s+by\b|\bmarks\b|markers?\s+highlighting|"
    r"characteri[sz]ed\s+by|defined\s+by|\bspecified\s+by\b|\bsignature\b",
    re.IGNORECASE,
)
TOKEN_RE = re.compile(r"\b([A-Z][A-Za-z0-9]{1,11})\b")

DOMAIN_STOPWORDS = {
    "FIG", "FIGS", "FIGURE", "FIGURES", "TABLE", "TABLES", "CLUSTER", "CLUSTERS",
    "EXTENDED", "DATA", "SUPP", "SUPPLEMENTARY", "SUPPLEMENT", "METHODS", "RESULTS",
    "GENE", "GENES", "CELL", "CELLS", "MOUSE", "HUMAN", "RAT", "MONKEY", "GFP", "RFP",
    "RNA", "DNAS", "DNA", "UMAP", "TSNE", "FACS", "GWAS", "IBD", "SNP", "SNPS", "QC",
    "GO", "PCR", "FISH", "IF", "IHC", "PMCID", "PMID", "DOI", "USA", "UK", "EU", "NIH",
    "GEO", "GSE", "CRE", "DAPI", "PI", "EDTA", "BSA", "PBS", "UMI", "UMIS", "READS",
    "DPI", "HIV", "SARS", "COVID", "UCSC", "ENSEMBL", "HGNC", "MGI", "FDR", "AUC",
    "ROC", "PCA", "ICA", "LR", "TDT", "IHCIF", "ANOVA", "STAR", "SEURAT", "SCANPY",
    "MONOCLE", "SLINGSHOT", "SCANORAMA", "HARMONY", "LIGER", "SCVI", "TOTALSEQ",
    "HASHTAG", "CITESEQ", "MULTIOME", "ATAC", "CHIP", "H3K", "GG", "TT", "CC", "AA",
    "WT", "KO", "KI", "OE", "NC", "PB", "DT", "MEN", "II", "III", "IV", "VI", "VII",
    "VIII", "IX", "XI", "XII", "XIII", "XIV", "XV", "TS", "CS", "PBS", "MLN", "GVL",
    "ALPHA", "BETA", "GAMMA", "DELTA", "OMEGA", "SIGMA", "LAMDA", "CTRL", "TREATED",
    "WEEK", "WEEKS", "MONTH", "MONTHS", "YEAR", "YEARS", "DAY", "DAYS", "P0", "P1",
    "P3", "P5", "P7", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "S10",
    "A1", "A2", "B1", "B2", "C1", "C2", "D1", "D2", "E1", "E2", "F1", "F2", "G1", "G2",
    "T1", "T2", "NK", "ILC", "DC", "UC", "CD", "CRC", "HCC", "NSCLC", "SCLC", "AML",
    "CML", "ALL", "T1D", "T2D", "MFI", "RIA", "ELISA", "WB", "GEL", "KDA", "MRNA",
    "SMRNA", "SNRNA", "SCRNA", "CMB", "DMEM", "FBS", "RT", "QPCR", "RTPCR", "CPM",
    "TPM", "RPKM", "FPKM", "PCE", "ACE2", "APCs", "MHC", "BD", "CV", "CVS", "SAN",
    "RV", "LV", "LA", "RA", "AV", "SA", "ECG", "EEG", "MEG", "MRI", "CT", "PET",
}
# “特征基因”等中文界面不涉及；下列为常见英文普通词（首字母大写无数字型 token 的拦截靠词表）。
WEAK_PATTERNS_NEED_VOCAB = True

# --- 噪声修正 1：常用蛋白别名 -> 基因符号（仅收录跨篇词表比对用） ---
ALIAS_TO_GENE = {
    "cd45": "ptprc",
    "glut1": "slc2a1",
    "cd31": "pecam1",
    "pd1": "pdcd1",
    "pd-l1": "cd274",
    "pdl1": "cd274",
    "cd56": "ncam1",
    "cd90": "thy1",
    "cd11b": "itgam",
    "cd11c": "itgax",
}

# --- 噪声修正 2：小鼠构建体名（GENE-Cre / GENE-Sun1 / GENE-GFP 等）整段剔除 ---
CONSTRUCT_RE = re.compile(
    r"\b[A-Z][A-Za-z0-9]*(?:-(?:Cre|CRE|Sun1|sfGFP|GFP|RFP|tdT|rtTA|DTR|ERT2|P2A|IRES|Luc))+\b"
)

# --- 噪声修正 3：Gate C 状态词（状态不同 = 不同细胞，不报重复） ---
STATE_WORDS = {
    "cycling", "mature", "immature", "proliferating", "activated", "memory",
    "naive", "effector", "resident", "inflammatory", "transitional",
    "intermediate", "putative",
}

# --- 噪声修正 4：引文上标粘连（Cd45^28 -> "Cd4528"）尾部数字剥离 ---
# 基因符号本身常以数字结尾（Cd45/Slc2a1），必须逐档尝试剥离并与已知基因比对，
# 不能假设词干以字母结尾。


def _superscript_hits(base: str, paper_genes: set[str], vocab: dict[str, list[str]]):
    """返回 (skip_stem, vocab_stem)：剥离 1-4 位尾部数字后的命中。

    skip_stem：词干命中该篇已收录基因（或其别名）→ 引文粘连，跳过；
    vocab_stem：词干命中跨篇词表 → 以词干为候选名。
    """
    skip_stem = None
    vocab_stem = None
    for n in range(1, 5):
        stem = base[:-n]
        if len(stem) < 3 or not re.search(r"[A-Za-z]", stem):
            break
        sk = stem.casefold()
        if sk in paper_genes or ALIAS_TO_GENE.get(sk) in paper_genes:
            skip_stem = stem
            break
        if vocab_stem is None and (sk in vocab or ALIAS_TO_GENE.get(sk) in vocab):
            vocab_stem = stem
    return skip_stem, vocab_stem

HEADER_RE = re.compile(r"^#{1,4}\s+(.*)$")
CAPTION_RE = re.compile(r"(Figure|Fig\.?|Table|Extended Data|Supp)\s*\.?\s*S?\d", re.IGNORECASE)


def load_table():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["markers"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    idx = {h: i for i, h in enumerate(header)}
    markers = []
    for r in rows[1:]:
        if not r or not r[idx["marker_id"]]:
            continue
        markers.append({h: r[i] for h, i in idx.items()})
    ex = wb["audit_exclusions"]
    exrows = list(ex.iter_rows(values_only=True))
    exheader = list(exrows[0])
    exidx = {h: i for i, h in enumerate(exheader)}
    exclusions = []
    for r in exrows[1:]:
        if not r or not r[exidx["paper_id"]]:
            continue
        exclusions.append({h: r[i] for h, i in exidx.items()})
    log = wb["import_log"]
    logrows = list(log.iter_rows(values_only=True))
    logheader = list(logrows[0])
    logidx = {h: i for i, h in enumerate(logheader)}
    titles = {}
    for r in logrows[1:]:
        if not r:
            continue
        pid = r[logidx["paper_id"]]
        title = r[logidx["paper_title"]]
        if pid and title and pid not in titles:
            titles[pid] = str(title)
    wb.close()
    return markers, exclusions, titles


def split_sentences(text: str):
    """按行切分（保留图注/标题上下文），再按句号细分。"""
    units = []
    header_stack = []
    for raw_line in text.splitlines():
        h = HEADER_RE.match(raw_line.strip())
        if h:
            header_stack.append(h.group(1)[:80])
            continue
        line = raw_line.strip()
        if not line or line.startswith("|") or line.startswith("!["):
            continue
        for sent in re.split(r"(?<=[.;])\s+", line):
            sent = sent.strip()
            if 25 <= len(sent) <= 800:
                units.append((sent, header_stack[-1] if header_stack else ""))
    return units


def gene_candidates(sentence: str, vocab: dict[str, list[str]], paper_genes: set[str]):
    """提取候选基因 token。返回 {token: (候选名, strong, in_vocab)}。

    已收录判定（返回前过滤）：token / 别名 / 上标剥离后任一命中该篇总表基因 → 跳过。
    候选命名：上标剥离命中的用剥离后词干，否则用原 token。
    """
    # 先剔除构建体名（Mpz-Sun1 / ActB-Cre / ChAT-Cre-tdT 等）
    sentence = CONSTRUCT_RE.sub(" ", sentence)
    found = {}
    for m in TOKEN_RE.finditer(sentence):
        tok = m.group(1)
        base = re.sub(r"[\+\-]$", "", tok)
        if base in DOMAIN_STOPWORDS or tok in DOMAIN_STOPWORDS:
            continue
        strong = False
        if re.fullmatch(r"[A-Z][A-Z0-9]{2,11}", base) and re.search(r"[A-Z]{2}", base):
            strong = True
        elif re.search(r"\d", base) and re.fullmatch(r"[A-Z][A-Za-z0-9]{2,11}", base):
            strong = True
        key = base.casefold()
        # 1) 直接命中该篇已收录基因 → 跳过
        if key in paper_genes:
            continue
        # 2) 别名命中该篇已收录基因 → 跳过
        alias_canon = ALIAS_TO_GENE.get(key)
        if alias_canon and alias_canon in paper_genes:
            continue
        # 3) 引文上标粘连：尾部数字剥离后命中该篇基因 → 跳过
        skip_stem = vocab_stem = None
        if base[-1:].isdigit():
            skip_stem, vocab_stem = _superscript_hits(base, paper_genes, vocab)
            if skip_stem:
                continue
        # --- 未收录，判定候选命名与置信 ---
        if vocab_stem:
            # 上标剥离后命中跨篇词表：候选名用词干
            found[tok] = (vocab_stem, strong, True)
        elif key in vocab or (alias_canon and alias_canon in vocab):
            canon_name = base if key in vocab else ALIAS_TO_GENE.get(key, "")
            found[tok] = (canon_name or base, strong, True)
        elif strong:
            found[tok] = (base, True, False)
    return found


def normalize_cell_type(ct: str) -> str:
    s = str(ct or "").lower()
    s = re.sub(r"\([^)]*\)", " ", s)
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    markers, exclusions, titles = load_table()

    by_paper = defaultdict(list)
    for m in markers:
        by_paper[m["paper_id"]].append(m)
    excl_by_paper = defaultdict(list)
    for e in exclusions:
        excl_by_paper[e["paper_id"]].append(e)

    # 跨篇基因词表：基因 -> 出现的 paper_id 列表
    vocab: dict[str, list[str]] = defaultdict(list)
    for m in markers:
        for key in (m["gene_symbol"], m["original_symbol"]):
            if key:
                k = str(key).casefold()
                if m["paper_id"] not in vocab[k]:
                    vocab[k].append(m["paper_id"])

    gate_a_rows = []
    gate_b_rows = []
    gate_c_rows = []
    overview = []

    for paper_id in sorted(by_paper, key=lambda p: (len(by_paper[p]), p)):
        rows = by_paper[paper_id]
        paper_genes = set()
        for m in rows:
            for key in (m["gene_symbol"], m["original_symbol"]):
                if key:
                    paper_genes.add(str(key).casefold())
        excl_genes = {}
        for e in excl_by_paper[paper_id]:
            for key in (e["normalized_symbol"], e["original_symbol"]):
                if key:
                    excl_genes.setdefault(str(key).casefold(), str(e.get("reason") or ""))

        md_path = MD_DIR / f"{paper_id}.md"
        a_vocab = a_pattern = 0
        if not md_path.exists():
            overview.append(
                [paper_id, titles.get(paper_id, ""), len(rows), "无 review_md（需先转换PDF）", "", "", "", "中", "待复核", "", ""]
            )
            continue

        text = md_path.read_text(encoding="utf-8", errors="ignore")
        sentences = split_sentences(text)
        seen_tokens = {}
        for sent, section in sentences:
            if not MARKER_LANGUAGE.search(sent):
                continue
            # 纯图注/方法学语句降噪：跳过只谈统计工具的句子
            if re.search(r"Seurat|Scanpy|harmony\s+correction|batch\s+correct|doublet", sent, re.I):
                continue
            cands = gene_candidates(sent, vocab, paper_genes)
            if not cands:
                continue
            for tok, (cand_name, strong, in_vocab) in cands.items():
                key = cand_name.casefold()
                if key in paper_genes:
                    continue
                if key in seen_tokens:
                    seen_tokens[key][5].append((sent, section))
                    continue
                if key in excl_genes:
                    status = f"此前已排除：{excl_genes[key][:60]}"
                    bucket = "excluded"
                else:
                    status = "未收录"
                    bucket = "vocab" if in_vocab else "pattern"
                    if in_vocab:
                        a_vocab += 1
                    elif strong:
                        a_pattern += 1
                vocab_papers = vocab.get(key, []) or vocab.get(ALIAS_TO_GENE.get(key, ""), [])
                seen_tokens[key] = [
                    paper_id,
                    titles.get(paper_id, "")[:40],
                    cand_name,
                    "是" if in_vocab else "否",
                    "、".join(vocab_papers[:4]) if in_vocab else "",
                    [(sent, section)],
                    status,
                    bucket,
                ]

        # Gate A 输出：词表命中优先，其次强模式
        ordered = sorted(
            seen_tokens.values(),
            key=lambda v: ({"vocab": 0, "pattern": 1, "excluded": 2}[v[7]], -len(v[5])),
        )
        for v in ordered[:40]:
            sent, section = v[5][0]
            extra = f"（另有 {len(v[5]) - 1} 处提及）" if len(v[5]) > 1 else ""
            gate_a_rows.append(
                [
                    v[0], v[1], v[2], v[3], v[4],
                    (f"[{section}] " if section else "") + sent[:280] + extra,
                    v[6],
                    "人工核对原文上下文：满足双重门槛则补录，否则注明不录原因",
                    "待处理",
                ]
            )

        # Gate B：证据类型候选
        b_up = b_recover = 0
        for m in rows:
            ctx = str(m.get("source_context") or "")
            loc = str(m.get("source_locator") or "")
            if not MARKER_LANGUAGE.search(ctx):
                continue
            ev = m["evidence_type"]
            if ev in ("annotation_marker", "figure_labeled") and not re.search(r"^Fig|^Extended|^Table|^S\d", loc.strip(), re.I):
                gate_b_rows.append(
                    [m["marker_id"], paper_id, m["cell_type"], m["gene_symbol"], ev,
                     "升级 author_declared 候选", loc[:50], ctx[:200], "待处理"]
                )
                b_up += 1
            elif m.get("candidate_class") == "context_only":
                gate_b_rows.append(
                    [m["marker_id"], paper_id, m["cell_type"], m["gene_symbol"], ev,
                     "身份恢复候选（context_only 但含 marker 措辞）", loc[:50], ctx[:200], "待处理"]
                )
                b_recover += 1

        # Gate C：语义重复
        groups = defaultdict(list)
        for m in rows:
            groups[(m["gene_symbol"] or "").casefold()].append(m)
        c_groups = 0
        for gene, grp in groups.items():
            if len(grp) < 2:
                continue
            normed = [(normalize_cell_type(g["cell_type"]), g) for g in grp]
            for i in range(len(grp)):
                for j in range(i + 1, len(grp)):
                    a_norm, a_m = normed[i]
                    b_norm, b_m = normed[j]
                    if not a_norm or not b_norm or a_norm == b_norm:
                        continue
                    a_set, b_set = set(a_norm.split()), set(b_norm.split())
                    if not a_set or not b_set or not (a_set <= b_set or b_set <= a_set):
                        continue
                    # 状态词守卫：状态不同（如 B cells vs Cycling B cells）= 不同细胞状态，
                    # 共享谱系 marker 属正常，不报重复；plasma 等分化阶段词除外（可能是误归属）
                    a_states, b_states = a_set & STATE_WORDS, b_set & STATE_WORDS
                    if a_states != b_states:
                        continue
                    reason = "cell_type 规范化后互为包含"
                    if a_norm == b_norm:
                        reason = "cell_type 规范化后完全相同（括号差异）"
                    gate_c_rows.append(
                        [paper_id, gene,
                         f"{a_m['marker_id']} {a_m['cell_type']}/{a_m['subtype']}",
                         f"{b_m['marker_id']} {b_m['cell_type']}/{b_m['subtype']}",
                         reason,
                         "人工判定是否同一证据重复：保留更具体/更高证据级行，另一行移入 audit_exclusions",
                         "待处理"]
                    )
                    c_groups += 1

        priority = "高" if (a_vocab > 0 or c_groups > 0) else ("中" if (a_pattern > 0 or b_up + b_recover > 0) else "低")
        overview.append(
            [paper_id, titles.get(paper_id, ""), len(rows), "有",
             a_vocab, a_pattern, b_up + b_recover, c_groups, priority, "待复核", "", ""]
        )

    # ---------- 输出工作簿 ----------
    wb = openpyxl.Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="4472C4")
    wrap = Alignment(wrap_text=True, vertical="top")

    def fill_sheet(ws, header, data, widths):
        ws.append(header)
        for c in ws[1]:
            c.font = head_font
            c.fill = head_fill
            c.alignment = Alignment(vertical="center")
        for row in data:
            ws.append(row)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = wrap

    ws = wb.active
    ws.title = "复核总览"
    fill_sheet(
        ws,
        ["paper_id", "论文标题", "marker行数", "review_md", "A漏提候选(词表)", "A漏提候选(模式)",
         "B证据类型候选", "C语义重复组", "优先级", "复核状态", "复核人", "备注"],
        sorted(overview, key=lambda r: {"高": 0, "中": 1, "低": 2, }.get(r[8], 3)),
        [26, 40, 9, 10, 13, 13, 12, 11, 7, 9, 8, 20],
    )

    fill_sheet(
        wb.create_sheet("A_漏提候选"),
        ["paper_id", "论文标题", "候选基因", "跨篇词表命中", "其他收录篇目", "证据句（原文，[章节]前缀）", "收录状态", "建议动作", "状态"],
        gate_a_rows,
        [26, 30, 12, 11, 24, 60, 24, 40, 8],
    )
    fill_sheet(
        wb.create_sheet("B_证据类型候选"),
        ["marker_id", "paper_id", "cell_type", "gene", "当前evidence_type", "候选方向", "source_locator", "source_context", "状态"],
        gate_b_rows,
        [10, 26, 28, 10, 16, 28, 24, 50, 8],
    )
    fill_sheet(
        wb.create_sheet("C_语义重复候选"),
        ["paper_id", "gene", "行1", "行2", "重复原因", "建议动作", "状态"],
        gate_c_rows,
        [26, 10, 40, 40, 24, 44, 8],
    )

    d_rows = []
    checks = ["聚类清单对账（正文/图注 cluster 数 vs 总表细胞类型数）",
              "marker 归属核对（同基因多细胞归属是否与原文语境一致）",
              "物种一致性（行内 species vs 任务表物种；鼠/鼠鼠基因写法）",
              "基因写法核对（保留原文大小写；大鼠首字母大写）",
              "跨篇一致性（细胞命名与四层分类口径）"]
    for row in sorted(overview, key=lambda r: {"高": 0, "中": 1, "低": 2}.get(r[8], 3)):
        for chk in checks:
            d_rows.append([row[0], row[1][:40], row[8], chk, "待处理", "", ""])
    fill_sheet(
        wb.create_sheet("D_人工清单"),
        ["paper_id", "论文标题", "优先级", "检查项", "状态", "结论", "备注"],
        d_rows,
        [26, 30, 7, 52, 8, 26, 20],
    )

    out = OUT_DIR / "recheck_plan.xlsx"
    wb.save(out)
    print(f"saved {out}")
    print(f"papers: {len(overview)}, gate A rows: {len(gate_a_rows)}, gate B rows: {len(gate_b_rows)}, gate C rows: {len(gate_c_rows)}")
    n_high = sum(1 for r in overview if r[8] == "高")
    print(f"priority: 高={n_high}, 中={sum(1 for r in overview if r[8]=='中')}, 低={sum(1 for r in overview if r[8]=='低')}")


if __name__ == "__main__":
    main()
