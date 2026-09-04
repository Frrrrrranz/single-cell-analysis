"""Batch 1 复核落表（recheck-2026-09-02 → 2026-09-03 应用）

应用五篇 verdict（marker提取/audits/recheck-2026-09-02/batch1_work/*_verdict.md）：
1. DOI_10.1038_s41586-020-2922-4（人肺图谱）    补录 8 / 升级 28 / 移除 34 / 标签修正 3 / 拼写修正 2
2. DOI_10.1038_s41588-022-01243-4（人肺空间图谱）补录 5 / 升级 28 / 移除 4 / 标签修正 7 / locator 修正 3 / is_pns 修正 72
3. DOI_10.1038_s41591-024-03215-z（乳腺癌 MBC）  升级 8 / 移除 2 / subtype 修正 1
4. DOI_10.1101_2025.09.26.678707（人肾图谱 v2） 升级 12 / is_pns 修正 48（verdict D 门结论"肾脏非 PNS"，落表核对时发现全表误设 true）
5. DOI_10.7554_elife.71752（人 DRG）            补录 19（A门 8 + D门追加 8 + 中等置信 3）
6. 试点遗留：M01516（PMID_35115729 Pdgfra）升级 author_declared

净变化：1882 + 32 - 40 = 1874 行
"""

from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(r"d:\OneDrive\Desktop\组\marker提取")
XLSX = ROOT / "表单" / "our_markers.xlsx"
BACKUP = ROOT / "audited-extraction" / "recovery" / "our_markers_pre_batch1_recheck_backup_2026-09-03.xlsx"

BATCH = "B20260903-BATCH1-RECHECK"
METHOD = "recheck_batch1_2026-09-03"
TODAY = datetime(2026, 9, 3)
AUDIT_MODEL = "recheck_pipeline_v2_main_agent"
REMOVE_TAG = "batch1_recheck_removed_2026-09-03"

P_LUNG = "DOI_10.1038_s41586-020-2922-4"
P_SPAT = "DOI_10.1038_s41588-022-01243-4"
P_BRCA = "DOI_10.1038_s41591-024-03215-z"
P_KIDN = "DOI_10.1101_2025.09.26.678707"
P_DRG = "DOI_10.7554_elife.71752"
P_PILOT = "PMID_35115729"

PAPER_META = {
    P_LUNG: {"task_no": 4, "dataset": "CELLxGENE:8c42cfd0-0b0a-46d5-910c-fc833d83c45e"},
    P_SPAT: {"task_no": 5, "dataset": "CELLxGENE:093d3bfe-6f0f-4ac0-a7a1-829f94d0a49f"},
    P_BRCA: {"task_no": 29, "dataset": "CELLxGENE:71513028-e4a0-4c62-a435-16598791690a"},
    P_KIDN: {"task_no": 39, "dataset": "CELLxGENE:9ac3d74d-1d94-4ff9-abf9-fe5abe4bd2a6"},
    P_DRG: {"task_no": 3, "dataset": "CELLxGENE:d973ea15-3722-4bc8-a0e8-1d956d26e203"},
}

# ---------- 共用原句 ----------

S_LUNG_FIBRO = (
    "Two clusters expressed classical fibroblast markers (BSG and COL1A2) (Fig. 1e) but one "
    "(SPINT2+FGFR4+GPC3+) localized to alveoli ('alveolar fibroblasts') and the other "
    "(SFRP2+PI16+SERPINF1+) to vascular adventitia and nearby airways ('adventitial fibroblasts') "
    "(Fig. 1f, Extended Data Fig. 4a-d)."
)
S_LUNG_PREPEND = (
    "Clearly distinct clusters were annotated ... we prepended a representative marker gene to their "
    "'canonical' identity (for example, IGSF21+ dendritic, EREG+ dendritic, and TREM2+ dendritic)."
)
S_LUNG_MYOFIBRO = "One cluster (WIF1+FGF18+ASPN+) is classical myofibroblasts."
S_LUNG_AT2SIG = (
    "AT2-signalling selective markers include Wnt ligands, receptors and transcription factors "
    "(for example, WNT5A, LRP5 and TCF7L2 highlighted green) (Extended Data Fig. 3i)."
)
S_EDF5G = (
    "HPA antibody staining of (g) non-myelinating Schwann cell markers (CADM, GRIK2, NCAM1, ITGB4 "
    "and L1CAM) (Extended Data Fig. 5g)."
)
S_NERVE4 = (
    "we identified the following four new clusters relating to airway peripheral nerves: myelinating "
    "Schwann cells (mSchwann) (NFASC, NCMAP, MBP and PRX), ... nonmyelinating Schwann cells (nmSchwann) "
    "(NGFR, SCN7A, CHD2, L1CAM and NCAM1); nmSchwann and mSchwann cell marker genes were enriched in "
    "cell adhesion and myelination gene sets."
)
S_SMG_MARKERS = (
    "smFISH staining for mucous (MUC5B), serous (LPO) and duct (MIA) cell markers in human bronchi "
    "sections (Extended Data Fig. 7d)."
)
S_IRVENPERI = (
    "IR-Ven-Peri markers CCL21 and CCL19 localize adjacent to the venous vessel marker ACKR1 (Fig. 3g); "
    "smFISH staining for IR-Ven-peri (CCL21, CCL19), venous endothelia (ACKR1) and smooth muscle (ACTA2) "
    "markers (Extended Data Fig. 6f)."
)
S_BRCA_PANEL = (
    "Canonical cell-type-specific markers (for example, EPCAM for epithelial cells, CD19 for B cells, "
    "CD4 for T helper cells, CD8 for cytotoxic T lymphocytes, CD56 for NK cells and CD14 for "
    "macrophages) (Methods, Gene panel design for MERFISH and ExSeq)."
)
S_KIDN_FR = (
    "Failed repair (fr) states were further marked by the expression of PROM1 (frPT and frTAL), ROBO2 "
    "and MEG3 (frPT), or ITGB8 and TMPRSS4 (frTAL) (Results, 'Resolved versus unresolved epithelial repair')."
)
S_KIDN_HBEGF = (
    "We also found another inflammatory population marked by expression of growth factors HBEGF and "
    "AREG, as well as proinflammatory genes PLAUR, IL1B, OSM and CXCL8 (Results, "
    "'Clinicopathologically-linked immune subtypes')."
)
S_DRG_NONNEUR = (
    "Clusters not expressing high levels of neuronal or somatosensory genes like SNAP25, SCN9A, SCN10A, "
    "PIEZO2, NEFH, etc. but instead expressing elevated levels of markers of non-neuronal cells including "
    "PRP1, MBP, QKI, LPAR1, and APOE were tagged as non-neuronal and were removed to allow reclustering "
    "of 'purified' human DRG neurons."
)
S_DRG_H12 = (
    "A second larger group of human neurons H12 is marked by NTRK3 and the voltage-gated ion channel "
    "SCN1A, but is only weakly positive for NEFH, expresses moderate levels of PIEZO2 (Figure 2B, "
    "Figure 1-figure supplement 5) and appears distinct from any potential mouse counterpart."
)

# ---------- 补录行（32 行） ----------


def new_row(marker_id: str, paper: str, cell_type: str, gene: str, evidence: str,
            locator: str, ctx: str, is_pns: str, four_layer, subtype=None, note=None) -> dict:
    return {
        "marker_id": marker_id,
        "task_no": PAPER_META[paper]["task_no"],
        "dataset_id": PAPER_META[paper]["dataset"],
        "paper_id": paper,
        "document_id": paper,
        "document_role": "primary",
        "ct_id": None,
        "subtype_id": None,
        "cell_type": cell_type,
        "subtype": subtype,
        "species": "human",
        "is_pns_cell": is_pns,
        "gene_symbol": gene,
        "original_symbol": gene,
        "evidence_type": evidence,
        "marker_polarity": "positive",
        "candidate_class": "formal_candidate",
        "source_locator": locator,
        "source_context": ctx,
        "review_status": "approved",
        "review_method": METHOD,
        "notes": note,
        "source_file": f"marker提取/review_md/{paper}.md",
        "imported_at": TODAY,
        "audit_status": "recheck_include",
        "normalization_status": "exact",
        "citation_verified": True,
        "audit_model": AUDIT_MODEL,
        "audit_notes": "Batch1 复核轮 2026-09-03 补录",
        "four_layer_category": four_layer,
        "recovery_source": "batch1_recheck",
    }


S_LUNG_GOBLET = (
    "canonical goblet cell markers MUC5B and MUC5AC and transcription factor SPDEF in mouse (left) "
    "and human (right) goblet cells (Extended Data Fig. 11a legend)."
)
S_LUNG_SEROUS = (
    "Dot plots of expression of serous cell markers LTF, LYZ, BPIFBP1 and HP showing switched expression "
    "(type 3 change) from mouse airway epithelial cells to human serous cells, which mice lack "
    "(Extended Data Fig. 12g legend)."
)

ADD_ROWS = [
    # --- 人肺图谱（8） ---
    new_row("M01886", P_LUNG, "Goblet cells", "MUC5B", "author_declared",
            "Extended Data Fig. 11a legend", S_LUNG_GOBLET, "false", "outside",
            note="Batch1 复核补录（A 门候选 4）：'canonical goblet cell markers' 双门槛满足；Gob（goblet）为 Fig.1a 注释类型，此前整簇漏提"),
    new_row("M01887", P_LUNG, "Goblet cells", "MUC5AC", "author_declared",
            "Extended Data Fig. 11a legend", S_LUNG_GOBLET, "false", "outside",
            note="Batch1 复核补录（A 门候选 5）：同 MUC5B 句"),
    new_row("M01888", P_LUNG, "Serous cells", "LTF", "author_declared",
            "Extended Data Fig. 12g legend", S_LUNG_SEROUS, "false", "outside",
            note="Batch1 复核补录（A 门候选 6）：'serous cell markers' 双门槛满足；Ser（serous）为 Fig.1a 注释类型，此前整簇漏提"),
    new_row("M01889", P_LUNG, "Serous cells", "LYZ", "author_declared",
            "Extended Data Fig. 12g legend", S_LUNG_SEROUS, "false", "outside",
            note="Batch1 复核补录（A 门候选 25）"),
    new_row("M01890", P_LUNG, "Serous cells", "BPIFBP1", "author_declared",
            "Extended Data Fig. 12g legend", S_LUNG_SEROUS, "false", "outside",
            note="Batch1 复核补录（A 门候选 26，存疑标注）：'serous cell markers' 措辞成立故收录；BPIFBP1 非标准 HGNC 符号（疑为 BPIFA1/BPIFB1 之 PDF 转写变体），保留原文拼写，待图版核对"),
    new_row("M01891", P_LUNG, "Serous cells", "HP", "author_declared",
            "Extended Data Fig. 12g legend", S_LUNG_SEROUS, "false", "outside",
            note="Batch1 复核补录（D 门整簇，同 A 门候选 6 句第 4 基因）"),
    new_row("M01892", P_LUNG, "Bronchial endothelial cell", "PLVAP", "annotation_marker",
            "Results §Cell markers, regulators and interactions, p.621",
            "Thus, bronchial endothelial cells are distinct from their counterparts in the pulmonary "
            "circulation, distinguished by matrix (VWA1 and HSPG2), fenestrated morphology (PLVAP) and "
            "cell cycle-associated (MYC and HBEGF) genes.",
            "false", "outside",
            note="Batch1 复核补录（A 门候选 31）：'distinguished by' 区分性 marker 措辞；旧'非 PNS 范围外'排除理由已失效。VWA1/HSPG2/HBEGF 未入本轮候选，记 D 门待扩清单"),
    new_row("M01893", P_LUNG, "Bronchial vessel cells", "MYC", "author_declared",
            "Extended Data Fig.3k legend",
            "smFISH for general endothelial marker CLDN5 (red, centre), bronchial vessel-specific markers "
            "MYC (green) and Bro1-specific marker ACKR1 (red, right) on serial sections of bronchial "
            "vessel cells (arrowheads).",
            "false", "outside",
            note="M00160 归属修正补录（D2b）：MYC 为 bronchial vessel（Bro1/Bro2）specific marker 而非泛支气管内皮 marker，'bronchial vessel-specific markers' 措辞"),
    # --- 人肺空间图谱（5） ---
    new_row("M01894", P_SPAT, "nonmyelinating Schwann cells (nmSchwann)", "CADM", "author_declared",
            "Extended Data Fig. 5g legend", S_EDF5G, "true", None,
            note="Batch1 复核补录（A 门候选 26 / B2 恢复）：'non-myelinating Schwann cell markers' 明确措辞；CADM 非标准 HGNC 符号（疑为 CADM1 之论文写法），PDF 原文（p.25）即写作 CADM，保留原文拼写，待图版核对"),
    new_row("M01895", P_SPAT, "smooth muscle cells", "ACTA2", "author_declared",
            "Extended Data Fig. 6f legend", S_IRVENPERI, "false", "outside",
            note="Batch1 复核补录（D 门整簇）：smooth muscle 此前整簇漏提；EDF6f 'smooth muscle (ACTA2) markers' 措辞"),
    new_row("M01896", P_SPAT, "secretory goblet/club cells", "SCGB1A1", "figure_labeled",
            "Extended Data Fig. 7e legend",
            "smFISH staining of secretory goblet/club (SCGB1A1), ciliated (FOXJ1) and duct "
            "(ALDH1A3/RARRES1) transformed expression separated by location in the single nuclei "
            "RNA-seq data in human bronchus section (Extended Data Fig. 7e).",
            "false", "outside",
            note="Batch1 复核补录（D 门整簇）：气道上皮参考类型整簇漏提；EDF7e smFISH 面板图级区分（无 marker 措辞，定 figure_labeled）"),
    new_row("M01897", P_SPAT, "ciliated cells", "FOXJ1", "figure_labeled",
            "Extended Data Fig. 7e legend",
            "smFISH staining of secretory goblet/club (SCGB1A1), ciliated (FOXJ1) and duct "
            "(ALDH1A3/RARRES1) transformed expression separated by location in the single nuclei "
            "RNA-seq data in human bronchus section (Extended Data Fig. 7e).",
            "false", "outside",
            note="Batch1 复核补录（D 门整簇）：同 M01896"),
    new_row("M01898", P_SPAT, "basal cells", "KRT14", "author_declared",
            "Extended Data Fig. 7a,i legends",
            "Marker gene expression dot plot for airway (KRT14), duct (ALDH1A3) and myoepithelium "
            "(FHOD3) in human bronchi epithelial cells (Extended Data Fig. 7a); smFISH staining for "
            "muscle (TAGLN), basal epithelia (KRT14) (Extended Data Fig. 7i).",
            "false", "outside",
            note="Batch1 复核补录（D 门整簇）：EDF7a 'Marker gene expression dot plot for airway (KRT14)' + EDF7i 'basal epithelia (KRT14)'"),
    # --- 人 DRG（19） ---
    new_row("M01899", P_DRG, "H12", "NTRK3", "author_declared",
            "Results p.6; Figure 2B; Figure 1-figure supplement 5", S_DRG_H12, "true", None,
            subtype="human-specific putative mechanosensor class",
            note="Batch1 复核补录（A 门候选 1）：'is marked by NTRK3' 明确 marker 措辞"),
    new_row("M01900", P_DRG, "peptidergic nociceptors", "TAC1", "author_declared",
            "Results p.10, Figure 5A legend; Results p.6; Figure 1C",
            "Peptidergic nociceptors marked by expression of TAC1 (blue) and additional SCN10A-positive "
            "cells are also present in this region of the ganglion (Figure 5A legend); ... these three "
            "markers each labeled a large group of neurons (Figure 2-figure supplement 1).",
            "true", None,
            subtype="covering H1/H2/H3/H5/H6",
            note="Batch1 复核补录（A 门候选 6，组级行）：图注 'marked by expression of TAC1' + 主文 'the same three markers'（TAC1/NEFH/OSMR）"),
    new_row("M01901", P_DRG, "H10", "MRGPRX1", "figure_labeled",
            "Results p.11; Figure 5B legend",
            "the human chloroquine responsive receptor MRGPRX1 localized selectively to H10 neurons "
            "(Figure 5B); Figure 5B legend: genes that distinguish H10 and H11 and mark specific sets "
            "of mouse NP1-3 neurons.",
            "true", None,
            subtype="nonpeptidergic nociceptor / candidate pruriceptor",
            note="Batch1 复核补录（A 门候选 15）：Figure 5B 图级区分证据，与现有 M00010（JAK1/H11）同口径"),
    new_row("M01902", P_DRG, "non-neuronal cells", "PRP1", "author_declared",
            "Methods: Single nuclear capture, sequencing, and data analysis (p.16)", S_DRG_NONNEUR,
            "false", None,
            note="Batch1 复核补录（A 门候选 16）：'markers of non-neuronal cells including...' 用于 'tagged as non-neuronal' 识别；原文即写作 PRP1（疑为 PRPH/peripherin 非规范写法），保留原文拼写；该群未细分亚型，按白名单 is_pns_cell=false"),
    new_row("M01903", P_DRG, "non-neuronal cells", "MBP", "author_declared",
            "Methods: Single nuclear capture, sequencing, and data analysis (p.16)", S_DRG_NONNEUR,
            "false", None,
            note="Batch1 复核补录（A 门候选 2）：同 PRP1 句"),
    new_row("M01904", P_DRG, "non-neuronal cells", "QKI", "author_declared",
            "Methods: Single nuclear capture, sequencing, and data analysis (p.16)", S_DRG_NONNEUR,
            "false", None,
            note="Batch1 复核补录（A 门候选 17）"),
    new_row("M01905", P_DRG, "non-neuronal cells", "LPAR1", "author_declared",
            "Methods: Single nuclear capture, sequencing, and data analysis (p.16)", S_DRG_NONNEUR,
            "false", None,
            note="Batch1 复核补录（A 门候选 18）"),
    new_row("M01906", P_DRG, "non-neuronal cells", "APOE", "author_declared",
            "Methods: Single nuclear capture, sequencing, and data analysis (p.16)", S_DRG_NONNEUR,
            "false", None,
            note="Batch1 复核补录（A 门候选 3）"),
    new_row("M01907", P_DRG, "H12", "SCN1A", "author_declared",
            "Results p.6; Figure 2B; Figure 1-figure supplement 5", S_DRG_H12, "true", None,
            subtype="human-specific putative mechanosensor class",
            note="Batch1 复核补录（D 门追加）：同句 'marked by NTRK3 and the voltage-gated ion channel SCN1A' 双对象"),
    new_row("M01908", P_DRG, "H15", "PVALB", "annotation_marker",
            "Results p.8; Figure 4B; Figure 4-figure supplement 1",
            "putative proprioceptive neurons (H15) were distinguished by their expression of NEFH, "
            "PIEZO2, and PVALB and lack of NTRK2 (Figure 4B, Figure 4-figure supplement 1); Putative "
            "proprioceptors, highlighted by double arrowheads, expressing PIEZO2 (green) and PVALB "
            "(red), but not NTRK2 (blue) were typically highly clustered in the ganglion.",
            "true", None,
            subtype="putative proprioceptive neurons",
            note="Batch1 复核补录（D 门追加）：'distinguished by their expression of' 识别句 + Figure 4B"),
    new_row("M01909", P_DRG, "H15", "NEFH", "annotation_marker",
            "Results p.8; Figure 4B; Figure 4-figure supplement 1",
            "putative proprioceptive neurons (H15) were distinguished by their expression of NEFH, "
            "PIEZO2, and PVALB and lack of NTRK2 (Figure 4B, Figure 4-figure supplement 1).",
            "true", None,
            subtype="putative proprioceptive neurons",
            note="Batch1 复核补录（D 门追加）：同基因不同 cell_type 新行，与 H3/H6/human DRG neurons 行不冲突"),
    new_row("M01910", P_DRG, "H15", "PIEZO2", "annotation_marker",
            "Results p.8; Figure 4B; Figure 4-figure supplement 1",
            "putative proprioceptive neurons (H15) were distinguished by their expression of NEFH, "
            "PIEZO2, and PVALB and lack of NTRK2 (Figure 4B, Figure 4-figure supplement 1).",
            "true", None,
            subtype="putative proprioceptive neurons",
            note="Batch1 复核补录（D 门追加）"),
    new_row("M01911", P_DRG, "H4", "SCN10A", "annotation_marker",
            "Results p.6; Figure 2B; Figure 1-figure supplement 5",
            "we designated H4 as c-nociceptors because of their expression of nociception-related "
            "SCN10A and NTRK1 and low level of NEFH (Figure 2B, Figure 1-figure supplement 5).",
            "true", None,
            subtype="c-nociceptors",
            note="Batch1 复核补录（D 门追加）：'designated ... because of their expression of' 命名依据句"),
    new_row("M01912", P_DRG, "H4", "NTRK1", "annotation_marker",
            "Results p.6; Figure 2B; Figure 1-figure supplement 5",
            "we designated H4 as c-nociceptors because of their expression of nociception-related "
            "SCN10A and NTRK1 and low level of NEFH (Figure 2B, Figure 1-figure supplement 5).",
            "true", None,
            subtype="c-nociceptors",
            note="Batch1 复核补录（D 门追加）"),
    new_row("M01913", P_DRG, "H10", "PIEZO2", "annotation_marker",
            "Results p.11; Figure 1C; Figure 2-figure supplement 2",
            "H10 cells are also distinguished from H11 and mouse pruriceptors by their prominent "
            "expression of the stretch-gated ion channel PIEZO2 (Figure 1C, Figure 2-figure supplement 2).",
            "true", None,
            subtype="nonpeptidergic nociceptor / candidate pruriceptor",
            note="Batch1 复核补录（D 门追加）：'distinguished by their prominent expression of' 区分句"),
    new_row("M01914", P_DRG, "H11", "SST", "figure_labeled",
            "Figure 5C legend; Results p.11",
            "probed for expression of genes that distinguish H11 (SST, blue) from H10 cells (PIEZO2, "
            "green) (Figure 5C legend); Multiplexed ISH showed that SST divides the OSMR-positive cells "
            "into two intermingling types (Figure 5C).",
            "true", None,
            subtype="candidate pruriceptor",
            note="Batch1 复核补录（D 门追加）：Figure 5C 图注逐基因区分指定，与 M00010 同口径"),
    new_row("M01915", P_DRG, "peptidergic nociceptors", "CALCA", "author_declared",
            "Results p.4-5; Figure 1C",
            "In the human DRG dataset, TAC1 (substance P), CALCA and CALCB (CGRP), and ADCYAP1 (PACAP), "
            "are expressed in several transcriptomic classes (H1, H2, H3, H5, and H6, Figure 1C); Some "
            "of these (including H3 and H6) also express peptidergic markers; Figure 1C legend: relative "
            "expression level of diagnostic markers.",
            "true", None,
            subtype="covering H1/H2/H3/H5/H6",
            note="Batch1 复核补录（D 门追加，中等置信）：'peptidergic markers' 回指性措辞 + Figure 1C diagnostic markers dotplot；待人工复核确认"),
    new_row("M01916", P_DRG, "peptidergic nociceptors", "CALCB", "author_declared",
            "Results p.4-5; Figure 1C",
            "In the human DRG dataset, TAC1 (substance P), CALCA and CALCB (CGRP), and ADCYAP1 (PACAP), "
            "are expressed in several transcriptomic classes (H1, H2, H3, H5, and H6, Figure 1C).",
            "true", None,
            subtype="covering H1/H2/H3/H5/H6",
            note="Batch1 复核补录（D 门追加，中等置信）：同 CALCA；待人工复核确认"),
    new_row("M01917", P_DRG, "peptidergic nociceptors", "ADCYAP1", "author_declared",
            "Results p.4-5; Figure 1C",
            "In the human DRG dataset, TAC1 (substance P), CALCA and CALCB (CGRP), and ADCYAP1 (PACAP), "
            "are expressed in several transcriptomic classes (H1, H2, H3, H5, and H6, Figure 1C).",
            "true", None,
            subtype="covering H1/H2/H3/H5/H6",
            note="Batch1 复核补录（D 门追加，中等置信）：同 CALCA；待人工复核确认"),
]

# ---------- 升级行（77 行：evidence_type → author_declared） ----------
# ctx/locator 来自 verdict 已核实的完整原句；note 为升级理由

UPGRADES: dict[str, dict] = {
    # --- 人肺图谱 B 门（10） ---
    "M00120": {"ctx": S_LUNG_FIBRO, "locator": "Results §New lung cell types; Fig.1e,f legends",
               "note": "Batch1 复核 B 门升级：(SFRP2+PI16+SERPINF1+) 组合签名命名 adventitial fibroblasts，'classical fibroblast markers' 语境"},
    "M00127": {"ctx": S_LUNG_FIBRO, "locator": "Results §New lung cell types; Fig.1e,f legends",
               "note": "Batch1 复核 B 门升级：(SPINT2+FGFR4+GPC3+) 组合签名命名 alveolar fibroblasts"},
    "M00129": {"ctx": S_LUNG_FIBRO, "locator": "Results §New lung cell types; Fig.1e,f legends",
               "note": "Batch1 复核 B 门升级：同 M00127"},
    "M00133": {"ctx": "Fig.1c legend: Dot plot of AT2 marker expression (10x dataset); One cluster "
                      "(WIF1+HHIP+CA2+) expressed higher levels of some canonical AT2 markers (SFTPA1, "
                      "SFTPC and ETV5).",
               "locator": "Results §New lung cell types; Fig.1c legend",
               "note": "Batch1 复核 B 门升级：'AT2 marker expression' dot plot + (WIF1+HHIP+CA2+) 签名"},
    "M00135": {"ctx": "AT2 selective markers include negative regulators of Hedgehog and Wnt signalling "
                      "pathways (for example, HHIP and WIF1, highlighted red) (Extended Data Fig. 3i).",
               "locator": "Results §New lung cell types; Extended Data Fig.3i legend",
               "note": "Batch1 复核 B 门升级：'AT2 selective markers include HHIP and WIF1'"},
    "M00180": {"ctx": S_LUNG_PREPEND + "; EREG+ dendritic cell marker EREG (red) and general dendritic "
                      "cell marker GPR183 (white) (m) (Extended Data Fig. 4m).",
               "locator": "Methods: Cell clustering, doublet calling, and annotation; Extended Data Fig.4m legend",
               "note": "Batch1 复核 B 门升级：'prepended a representative marker gene' + 图注 'EREG+ dendritic cell marker EREG'"},
    "M00184": {"ctx": S_LUNG_PREPEND + "; cell markers IGSF21 (red) and GPR34 (white) (l) "
                      "(Extended Data Fig. 4l).",
               "locator": "Methods: Cell clustering, doublet calling, and annotation; Extended Data Fig.4l legend",
               "note": "Batch1 复核 B 门升级：同 M00180"},
    "M00194": {"ctx": S_LUNG_MYOFIBRO + " Myofibroblast and fibromyocyte marker ASPN (red) "
                      "(Extended Data Fig. 4f).",
               "locator": "Results §New lung cell types; Extended Data Fig.4f legend",
               "note": "Batch1 复核 B 门升级：图注 'myofibroblast and fibromyocyte marker ASPN'"},
    "M00196": {"ctx": S_LUNG_MYOFIBRO, "locator": "Results §New lung cell types; Fig.1e legend",
               "note": "Batch1 复核 B 门升级：(WIF1+FGF18+ASPN+) 组合签名命名 classical myofibroblasts"},
    "M00224": {"ctx": S_LUNG_PREPEND + "; TREM2+ dendritic cell markers TREM2 (red) and CHI3L1 (white) "
                      "(n) (Extended Data Fig. 4n).",
               "locator": "Methods: Cell clustering, doublet calling, and annotation; Extended Data Fig.4n legend",
               "note": "Batch1 复核 B 门升级：同 M00180"},
    # --- 人肺图谱 D 门顺带（18） ---
    "M00118": {"ctx": "fibroblast-selective markers Pi16 (white) and Serpinf1 (red) (adventitial "
                      "fibroblast probes, Extended Data Fig. 4d).",
               "locator": "Results §New lung cell types; Fig.1e legend; Extended Data Fig.4d legend",
               "note": "Batch1 复核 D 门顺带升级：EDF4 图注 'fibroblast-selective markers Pi16'（M00121 拼写变体行已并入本行）"},
    "M00119": {"ctx": "adventitial fibroblast marker SERPINF1 (red, right) (Fig. 1f); fibroblast-selective "
                      "markers Pi16 (white) and Serpinf1 (red) (Extended Data Fig. 4d).",
               "locator": "Fig.1f legend; Extended Data Fig.4d legend",
               "note": "Batch1 复核 D 门顺带升级：Fig.1f 'adventitial fibroblast marker SERPINF1'（M00122 已并入）"},
    "M00128": {"ctx": "alveolar fibroblast marker GPC3 (red, left) (Fig. 1f).",
               "locator": "Fig.1f legend",
               "note": "Batch1 复核 D 门顺带升级：Fig.1f 'alveolar fibroblast marker GPC3'"},
    "M00131": {"ctx": "alveolar fibroblast-selective markers Slc7a10 (white) and Fgfr4 (red) "
                      "(Extended Data Fig. 4b).",
               "locator": "Extended Data Fig.4b legend",
               "note": "Batch1 复核 D 门：拼写修正 SLC7A10 + species mouse→human（原鼠式写法系图注原始拼写所致）+ 升级 author_declared（'alveolar fibroblast-selective markers'）"},
    "M00137": {"ctx": "shared AT2 and AT2-signalling marker SFTPC (white) and specific AT2 marker WIF1 "
                      "(red puncta) (Fig. 1d); AT2 marker SFTPC (Fig. 4d).",
               "locator": "Fig.1d legend; Fig. 4d legend",
               "note": "Batch1 复核 D 门顺带升级（M00141 标签分裂行已并入，Fig.4d 证据并入）"},
    "M00138": {"ctx": "specific AT2 marker WIF1 (red puncta) (Fig. 1d); AT2 selective markers include ... "
                      "HHIP and WIF1, highlighted red (Extended Data Fig. 3i).",
               "locator": "Fig.1d legend; Extended Data Fig.3i legend",
               "note": "Batch1 复核 D 门顺带升级：'specific AT2 marker WIF1'"},
    "M00142": {"ctx": "shared AT2 and AT2-signalling marker SFTPC (white) (Fig. 1d).",
               "locator": "Fig.1d legend",
               "note": "Batch1 复核 D 门顺带升级 + 标签修正 AT2-signalling cell (AT2-s)"},
    "M00144": {"ctx": S_LUNG_AT2SIG, "locator": "Extended Data Fig.3i legend",
               "note": "Batch1 复核 D 门顺带升级：'AT2-signalling selective markers include WNT5A, LRP5 and TCF7L2'"},
    "M00145": {"ctx": S_LUNG_AT2SIG, "locator": "Extended Data Fig.3i legend",
               "note": "Batch1 复核 D 门顺带升级：同 M00144"},
    "M00146": {"ctx": S_LUNG_AT2SIG, "locator": "Extended Data Fig.3i legend",
               "note": "Batch1 复核 D 门顺带升级：同 M00144"},
    "M00153": {"ctx": "basal cells (marked by KRT5, red) (Extended Data Fig. 3e); basal marker KRT5 "
                      "(Extended Data Fig. 3g).",
               "locator": "Extended Data Fig.3e,g legends",
               "note": "Batch1 复核 D 门顺带升级：'marked by KRT5' / 'basal marker KRT5'"},
    "M00161": {"ctx": "Bro1-specific marker ACKR1 (red, right) (Extended Data Fig. 3k).",
               "locator": "Extended Data Fig.3k legend",
               "note": "Batch1 复核 D 门顺带升级：'Bro1-specific marker ACKR1'"},
    "M00174": {"ctx": "general dendritic cell marker GPR183 (white) (m) (Extended Data Fig. 4m).",
               "locator": "Extended Data Fig.4m legend",
               "note": "Batch1 复核 D 门顺带升级：'general dendritic cell marker GPR183'"},
    "M00182": {"ctx": "myofibroblast and fibromyocyte marker ASPN (red) (Extended Data Fig. 4f).",
               "locator": "Extended Data Fig.4f legend",
               "note": "Batch1 复核 D 门顺带升级：'myofibroblast and fibromyocyte marker ASPN'"},
    "M00185": {"ctx": "cell markers IGSF21 (red) and GPR34 (white) (l) (Extended Data Fig. 4l).",
               "locator": "Extended Data Fig.4l legend",
               "note": "Batch1 复核 D 门顺带升级：EDF4l 'cell markers IGSF21 and GPR34'"},
    "M00195": {"ctx": S_LUNG_MYOFIBRO, "locator": "Results §New lung cell types; Fig.1e legend",
               "note": "Batch1 复核 D 门顺带升级：(WIF1+FGF18+ASPN+) 组合签名"},
    "M00212": {"ctx": "Alveolar section of human lung probed by smFISH for pericyte marker COX4I2 "
                      "(Extended Data Fig. 5d).",
               "locator": "Extended Data Fig.4h legend; Extended Data Fig.5d legend",
               "note": "Batch1 复核 D 门顺带升级：EDF5d 'pericyte marker COX4I2'"},
    "M00225": {"ctx": "TREM2+ dendritic cell markers TREM2 (red) and CHI3L1 (white) (n) "
                      "(Extended Data Fig. 4n).",
               "locator": "Extended Data Fig.4n legend",
               "note": "Batch1 复核 D 门顺带升级：EDF4n 'TREM2+ dendritic cell markers ... CHI3L1'"},
    # --- 人肺空间图谱（28） ---
    "M00264": {"ctx": "We identified a previously undefined cluster expressing monocyte (CD14) and "
                      "macrophage markers, termed macro-intermediate (Extended Data Fig. 8b).",
               "locator": "Results, section 'Myeloid cells show previously undescribed phenotypes'; Extended Data Fig. 8b",
               "note": "Batch1 复核 B 门升级：'monocyte (CD14) and macrophage markers' + termed 命名用途"},
    "M00024": {"ctx": S_EDF5G, "locator": "Extended Data Fig. 5g legend",
               "note": "Batch1 复核 B2 门升级（B① supplementary_marker 升级）：EDF5g 'non-myelinating Schwann cell markers' 图注"},
    "M00025": {"ctx": S_EDF5G, "locator": "Extended Data Fig. 5g legend",
               "note": "Batch1 复核 B2 门升级：同 M00024"},
    "M00294": {"ctx": S_SMG_MARKERS + " serous cells (LPO+RARRES1-APRILhigh) (Fig. 5g; Extended Data Fig. 10d,e).",
               "locator": "Results (Fig. 5g; Extended Data Fig. 10d,e); Fig. 4b legend; Extended Data Fig. 7d legend",
               "note": "Batch1 复核 C 门合并升级：M00297 重复行已并入，locator 合并"},
    "M00013": {"ctx": S_NERVE4, "locator": "Results, section 'Four distinct cell types in airway peripheral nerves'; Extended Data Fig. 5a,b",
               "note": "Batch1 复核 D 门顺带升级：四新簇命名句 'myelinating Schwann cells (mSchwann) (NFASC, NCMAP, MBP and PRX)'"},
    "M00014": {"ctx": S_NERVE4, "locator": "Results, section 'Four distinct cell types in airway peripheral nerves'; Extended Data Fig. 5a,b",
               "note": "Batch1 复核 D 门顺带升级：同 M00013"},
    "M00015": {"ctx": S_NERVE4, "locator": "Results, section 'Four distinct cell types in airway peripheral nerves'; Extended Data Fig. 5a,b",
               "note": "Batch1 复核 D 门顺带升级：同 M00013"},
    "M00016": {"ctx": S_NERVE4, "locator": "Results, section 'Four distinct cell types in airway peripheral nerves'; Extended Data Fig. 5a,b",
               "note": "Batch1 复核 D 门顺带升级：同 M00013"},
    "M00017": {"ctx": S_NERVE4, "locator": "Results, section 'Four distinct cell types in airway peripheral nerves'; Extended Data Fig. 5a,b",
               "note": "Batch1 复核 D 门顺带升级：'nonmyelinating Schwann cells (nmSchwann) (NGFR, SCN7A, CHD2, L1CAM and NCAM1)'"},
    "M00018": {"ctx": S_NERVE4, "locator": "Results, section 'Four distinct cell types in airway peripheral nerves'; Extended Data Fig. 5a,b",
               "note": "Batch1 复核 D 门顺带升级：同 M00017"},
    "M00019": {"ctx": S_NERVE4, "locator": "Results, section 'Four distinct cell types in airway peripheral nerves'; Extended Data Fig. 5a,b",
               "note": "Batch1 复核 D 门顺带升级：同 M00017"},
    "M00020": {"ctx": S_NERVE4 + " " + S_EDF5G,
               "locator": "Results, section 'Four distinct cell types in airway peripheral nerves'; Extended Data Fig. 5a,b; Extended Data Fig. 5g legend",
               "note": "Batch1 复核 D 门顺带升级：命名句 + EDF5g 'non-myelinating Schwann cell markers (... L1CAM)'"},
    "M00021": {"ctx": S_NERVE4 + " " + S_EDF5G,
               "locator": "Results, section 'Four distinct cell types in airway peripheral nerves'; Extended Data Fig. 5a,b; Extended Data Fig. 5g legend",
               "note": "Batch1 复核 D 门顺带升级：同 M00020"},
    "M00023": {"ctx": "Nerve-associated cell type markers have distinct locations in the airway nerve "
                      "bundles identified by smFISH staining (Fig. 2i); non-myelinating (SCN7A, SOX10) "
                      "Schwann cell ... specific genes (Extended Data Fig. 5j).",
               "locator": "Fig. 2i legend; Extended Data Fig. 5j",
               "note": "Batch1 复核 D 门顺带升级：Fig. 2i 'Nerve-associated cell type markers' smFISH + EDF5j"},
    "M00235": {"ctx": "endoneurial NAF marker (USP54) (Extended Data Fig. 5h).",
               "locator": "Extended Data Fig. 5h caption",
               "note": "Batch1 复核 D 门顺带升级：'endoneurial NAF marker (USP54)'"},
    "M00279": {"ctx": "perineurial NAF markers (SLC22A3 and SORBS1) (Extended Data Fig. 5i).",
               "locator": "Extended Data Fig. 5i caption",
               "note": "Batch1 复核 D 门顺带升级：'perineurial NAF markers'"},
    "M00281": {"ctx": "perineurial NAF markers (SLC22A3 and SORBS1) (Extended Data Fig. 5i).",
               "locator": "Extended Data Fig. 5i caption",
               "note": "Batch1 复核 D 门顺带升级：同 M00279"},
    "M00246": {"ctx": "Dot plot of IR-fibro marker genes that overlap with Fibroblast reticular cell "
                      "and fDC markers (Fig. 2c legend).",
               "locator": "Fig. 2c legend",
               "note": "Batch1 复核 D 门顺带升级：Fig. 2c 'IR-fibro marker genes'"},
    "M00247": {"ctx": "Dot plot of IR-fibro marker genes that overlap with Fibroblast reticular cell "
                      "and fDC markers (Fig. 2c legend).",
               "locator": "Fig. 2c legend",
               "note": "Batch1 复核 D 门顺带升级：同 M00246"},
    "M00248": {"ctx": "Dot plot of IR-fibro marker genes that overlap with Fibroblast reticular cell "
                      "and fDC markers (Fig. 2c legend).",
               "locator": "Fig. 2c legend",
               "note": "Batch1 复核 D 门顺带升级：同 M00246"},
    "M00249": {"ctx": "Dot plot of IR-fibro marker genes that overlap with Fibroblast reticular cell "
                      "and fDC markers (Fig. 2c legend).",
               "locator": "Fig. 2c legend",
               "note": "Batch1 复核 D 门顺带升级：同 M00246"},
    "M00251": {"ctx": S_IRVENPERI, "locator": "Fig. 3g legend; Extended Data Fig. 6f legend",
               "note": "Batch1 复核 D 门顺带升级：Fig. 3g 'IR-Ven-Peri markers CCL21 and CCL19'"},
    "M00253": {"ctx": S_IRVENPERI, "locator": "Fig. 3g legend; Extended Data Fig. 6f legend",
               "note": "Batch1 复核 D 门顺带升级：同 M00251"},
    "M00284": {"ctx": "B lineage markers (IgD, IgA2 and IgG) (Fig. 4g legend); We also detected IgD+ "
                      "naive B cells and CD3+ CD4+ T helper cells in the human SMG (Fig. 4g).",
               "locator": "Methods, COVID-19 data analysis; Fig. 4g legend",
               "note": "Batch1 复核 D 门顺带升级：Fig. 4g 'B lineage markers (IgD, IgA2 and IgG)'（IgA2 对应基因 IGHA2），用于 SMG 中 IgA2+/IgG- 细胞识别"},
    "M00285": {"ctx": "B lineage markers (IgD, IgA2 and IgG) (Fig. 4g legend); We also detected IgD+ "
                      "naive B cells and CD3+ CD4+ T helper cells in the human SMG (Fig. 4g).",
               "locator": "Methods, COVID-19 data analysis; Fig. 4g legend",
               "note": "Batch1 复核 D 门顺带升级：同 M00284"},
    "M00292": {"ctx": S_SMG_MARKERS,
               "locator": "Results, section 'Identification of duct cells in airway SMG'; Extended Data Fig. 7d legend",
               "note": "Batch1 复核 D 门顺带升级：EDF7d 'duct (MIA) cell markers'"},
    "M00296": {"ctx": S_SMG_MARKERS, "locator": "Fig. 4b legend; Extended Data Fig. 7d legend",
               "note": "Batch1 复核 D 门顺带升级 + 标签修正 SMG mucous cells：EDF7d 'mucous (MUC5B) cell markers'"},
    "M00298": {"ctx": S_IRVENPERI, "locator": "Fig. 3g legend; Extended Data Fig. 6f legend",
               "note": "Batch1 复核 D 门顺带升级：Fig. 3g 'the venous vessel marker ACKR1' + EDF6f"},
    # --- 乳腺癌（8） ---
    "M01304": {"ctx": S_BRCA_PANEL, "locator": "Methods, Gene panel design for MERFISH and ExSeq, Canonical cell-type-specific markers; Fig. 5b",
               "note": "Batch1 复核 B 门升级：'Canonical cell-type-specific markers ... CD19 for B cells' + 面板覆盖 all major cell types + Fig. 5b marker 分区"},
    "M01305": {"ctx": "a small population of potentially regulatory B cells expressing FOXP3 in addition "
                      "to the typical B cell marker FCRL5 (Methods, De novo cell type annotation of the "
                      "cell-segmented MERFISH data).",
               "locator": "Methods, De novo cell type annotation of the cell-segmented MERFISH data",
               "note": "Batch1 复核 B 门升级：'the typical B cell marker FCRL5'，语境即 MERFISH de novo 注释流程"},
    "M01308": {"ctx": S_BRCA_PANEL, "locator": "Methods, Gene panel design for MERFISH and ExSeq, Gene collection, Canonical cell-type-specific markers",
               "note": "Batch1 复核 B 门升级；注：原文写作 CD8（未区分 CD8A/CD8B），行内 CD8A 属蛋白/家族→基因映射"},
    "M01309": {"ctx": S_BRCA_PANEL, "locator": "Methods, Gene panel design for MERFISH and ExSeq, Canonical cell-type-specific markers; Fig. 5b",
               "note": "Batch1 复核 B 门升级：'Canonical cell-type-specific markers ... EPCAM for epithelial cells'"},
    "M01311": {"ctx": "Macrophage co-localization phenotypes (Fig. 4c,e) were neither specifically "
                      "enriched nor depleted with expression of CD163, a key macrophage marker, with the "
                      "three representative samples showing predominantly CD163+ macrophages (Fig. 4f).",
               "locator": "Results, Fig. 4f; Fig. 5a,b; Methods, Gene panel design",
               "note": "Batch1 复核 B 门升级：'CD163, a key macrophage marker'；M01314 重复行已并入本行"},
    "M01313": {"ctx": S_BRCA_PANEL, "locator": "Methods, Gene panel design; Fig. 5b",
               "note": "Batch1 复核 B 门升级：'Canonical cell-type-specific markers ... CD14 for macrophages'"},
    "M01316": {"ctx": S_BRCA_PANEL, "locator": "Methods, Gene panel design; Fig. 5b",
               "note": "Batch1 复核 B 门升级；注：原文写作 CD56（蛋白名），NCAM1 未在原文文本出现，属蛋白→基因映射"},
    "M01317": {"ctx": S_BRCA_PANEL, "locator": "Methods, Gene panel design; Fig. 5b",
               "note": "Batch1 复核 B 门升级：'Canonical cell-type-specific markers ... CD4 for T helper cells'"},
    # --- 肾脏（12） ---
    "M01549": {"ctx": S_KIDN_FR, "locator": "Results, 'Resolved versus unresolved epithelial repair'",
               "note": "Batch1 复核 B 门升级：'states were further marked by the expression of'（MEG3 括注 frPT）"},
    "M01550": {"ctx": S_KIDN_FR, "locator": "Results, 'Resolved versus unresolved epithelial repair'",
               "note": "Batch1 复核 B 门升级：同句（PROM1 括注 frPT and frTAL）"},
    "M01551": {"ctx": S_KIDN_FR, "locator": "Results, 'Resolved versus unresolved epithelial repair'",
               "note": "Batch1 复核 B 门升级：同句（ROBO2 括注 frPT）"},
    "M01552": {"ctx": S_KIDN_FR, "locator": "Results, 'Resolved versus unresolved epithelial repair'",
               "note": "Batch1 复核 B 门升级：同句（ITGB8 括注 frTAL）"},
    "M01553": {"ctx": S_KIDN_FR, "locator": "Results, 'Resolved versus unresolved epithelial repair'",
               "note": "Batch1 复核 B 门升级：同句（PROM1 括注 frPT and frTAL）"},
    "M01554": {"ctx": S_KIDN_FR, "locator": "Results, 'Resolved versus unresolved epithelial repair'",
               "note": "Batch1 复核 B 门升级：同句（TMPRSS4 括注 frTAL）"},
    "M01566": {"ctx": S_KIDN_HBEGF, "locator": "Results, 'Clinicopathologically-linked immune subtypes'",
               "note": "Batch1 复核 B 门升级：'population marked by expression of growth factors HBEGF and AREG'"},
    "M01567": {"ctx": S_KIDN_HBEGF, "locator": "Results, 'Clinicopathologically-linked immune subtypes'",
               "note": "Batch1 复核 B 门升级：'marked by expression of ... proinflammatory genes PLAUR, IL1B, OSM and CXCL8' 整体收录"},
    "M01568": {"ctx": S_KIDN_HBEGF, "locator": "Results, 'Clinicopathologically-linked immune subtypes'",
               "note": "Batch1 复核 B 门升级：细胞群以 HBEGF+ 命名，marker 身份与注释用途最强"},
    "M01569": {"ctx": S_KIDN_HBEGF, "locator": "Results, 'Clinicopathologically-linked immune subtypes'",
               "note": "Batch1 复核 B 门升级：同 M01567"},
    "M01570": {"ctx": S_KIDN_HBEGF, "locator": "Results, 'Clinicopathologically-linked immune subtypes'",
               "note": "Batch1 复核 B 门升级：同 M01567"},
    "M01571": {"ctx": S_KIDN_HBEGF, "locator": "Results, 'Clinicopathologically-linked immune subtypes'",
               "note": "Batch1 复核 B 门升级：同 M01567"},
    # --- 试点遗留（1） ---
    "M01516": {"ctx": "We detected EFs in the Mpz-Sun1 atlas (Fig. 2b) and found they express the "
                      "fibroblast marker Pdgfra as well as the stem cell markers Cd34 and nmSC marker "
                      "Ngfr (Fig. 2c).",
               "locator": "Results p.4; Fig. 2b-c",
               "note": "试点遗留升级（elife.71752 与 s41588-022-01243-4 两篇 verdict 复核确认）：'the fibroblast marker Pdgfra' 直接归属 EFs；M01517 Cd34 维持 annotation_marker（'stem cell markers' 措辞指向干性身份）"},
}

# ---------- 字段修正（标签/拼写/subtype/locator，不涉及 evidence_type） ----------

FIELD_FIXES: dict[str, dict] = {
    # 人肺图谱：标签修正 3 + 拼写修正 1（M00131 拼写在 UPGRADES note 中说明，此处只改字段）
    "M00140": {"cell_type": "Alveolar type 2 (AT2)",
               "audit_notes": "Batch1 复核 D6 标签修正：AT2 → Alveolar type 2 (AT2)（标签分裂统一）"},
    "M00143": {"cell_type": "AT2-signalling cell (AT2-s)",
               "audit_notes": "Batch1 复核 D6 标签修正：与 M00144-146 统一写法；polarity=negative 保留（图注 SFTPC+ WIF1- 支持阴性区分特征）"},
    "M00131": {"gene_symbol": "SLC7A10", "species": "human",
               "audit_notes": "Batch1 复核 D5 拼写修正：Slc7a10 → SLC7A10；species mouse → human（原鼠式写法系 EDF4b 图注原始拼写所致）"},
    "M00192": {"gene_symbol": "ELN", "species": "human",
               "audit_notes": "Batch1 复核 D5 拼写修正：Eln → ELN；species mouse → human（原鼠式写法系图注原始拼写所致）；evidence_type 维持 figure_labeled"},
    # 人肺空间图谱：标签修正 + locator 修正
    "M00240": {"cell_type": "IR-fibro",
               "source_locator": "preamble (Introduction/Abstract); Fig. 2c legend; Extended Data Fig. 3a legend",
               "audit_notes": "Batch1 复核 D6 标签修正 → IR-fibro（短名统一）；EDF3a smFISH 证据 locator 并入（M00244 已并入）"},
    "M00241": {"cell_type": "IR-fibro",
               "source_locator": "preamble (Introduction/Abstract); Fig. 2c legend; Extended Data Fig. 3a legend",
               "audit_notes": "Batch1 复核 D6 标签修正 → IR-fibro；同 M00240（M00245 已并入）"},
    "M00236": {"cell_type": "endoneurial NAF",
               "audit_notes": "Batch1 复核 D6 标签修正：endoneurial nerve-associated fibroblasts (NAF) → endoneurial NAF（与 M00235 短名统一）"},
    "M00237": {"cell_type": "endoneurial NAF",
               "audit_notes": "Batch1 复核 D6 标签修正：同 M00236"},
    "M00295": {"cell_type": "SMG serous cells",
               "audit_notes": "Batch1 复核 D6 标签修正：SMG serous cell → SMG serous cells（复数风格统一）；polarity=negative 保留（serous cells LPO+RARRES1-）"},
    "M00282": {"source_locator": "Results, section 'Colocalization of IgA plasma cells with the SMG'; Extended Data Fig. 9c legend (panels d,e)",
               "audit_notes": "Batch1 复核 D 门 locator 修正：原 EDF 9b 有误，图注实为 EDF 9c (panels d,e) 'B plasma marker MZB1'"},
    "M00228": {"source_locator": "Results, section 'Colocalization of IgA plasma cells with the SMG' / Fig. 4g; Fig. 6 schematic",
               "audit_notes": "Batch1 复核 D6：Fig. 6 schematic 证据并入 locator（M00227 语义重复行已移除）"},
    # 乳腺癌：subtype 修正
    "M01315": {"subtype": "CD163+ state",
               "audit_notes": "Batch1 复核 D 门建议修正：原 subtype 'CD163+/CD68+/APOE+/HIF1A+ state' 嵌入非作者声明 marker（APOE/HIF1A 见 A/B2 门判定不录），简化为 CD163+ state"},
    # 肾脏：αSMA 备注
    "M01578": {"notes_append": "原文用 αSMA（蛋白名），ACTA2 为其 HGNC 基因符号（蛋白名→基因符号转换，可接受）",
               "audit_notes": "Batch1 复核 D 门第 4 项备注"},
}

# ---------- is_pns_cell 系统性修正 ----------
# 空间图谱：M00013-M00025（Schwann 13 行）保持 true，其余 72 行 → false
# 肾脏：全表 48 行 → false（verdict D 门结论"肾脏非 PNS"；落表核对时发现全表误设 true，verdict 摘要漏记此项）

SPATIAL_PNS_KEEP = {f"M{i:05d}" for i in range(13, 26)} | {"M01894"}

# ---------- 移除行（40 行 → audit_exclusions） ----------

R_FLOW = ("流式抗体面板归属错位（D2 门）：Methods 血免疫细胞 FACS 染色基因（CD3/CD4/CD8/CD14/CD19/CD47/CD56/CD235a 面板）"
          "被批量错挂到非归属细胞；正确归属行已存在。保留/正确归属：{sup}")

REMOVALS: dict[str, tuple[str, str]] = {
    # --- 人肺图谱：流式错位 19 ---
    "M00147": (R_FLOW.format(sup="classical monocytes M00168（CD14）"), "M00168"),
    "M00150": (R_FLOW.format(sup="CD4+ T cells M00162（CD4）"), "M00162"),
    "M00152": (R_FLOW.format(sup="NK M00199/M00203（NCAM1）"), "M00199"),
    "M00155": (R_FLOW.format(sup="classical monocytes M00168（CD14）"), "M00168"),
    "M00156": (R_FLOW.format(sup="B cells M00148（CD19）"), "M00148"),
    "M00159": (R_FLOW.format(sup="NK M00199/M00203（NCAM1）"), "M00199"),
    "M00197": (R_FLOW.format(sup="classical monocytes M00168（CD14）"), "M00168"),
    "M00198": (R_FLOW.format(sup="B cells M00148（CD19）"), "M00148"),
    "M00200": (R_FLOW.format(sup="classical monocytes M00168（CD14）"), "M00168"),
    "M00201": (R_FLOW.format(sup="B cells M00148（CD19）"), "M00148"),
    "M00202": (R_FLOW.format(sup="CD4+ T cells M00162（CD4）"), "M00162"),
    "M00204": (R_FLOW.format(sup="basophils/eosinophils M00154（CCR3）"), "M00154"),
    "M00205": (R_FLOW.format(sup="classical monocytes M00168（CD14）"), "M00168"),
    "M00206": (R_FLOW.format(sup="B cells M00148（CD19）"), "M00148"),
    "M00211": (R_FLOW.format(sup="NK M00199/M00203（NCAM1）"), "M00199"),
    "M00217": (R_FLOW.format(sup="classical monocytes M00168（CD14）"), "M00168"),
    "M00218": (R_FLOW.format(sup="B cells M00148（CD19）"), "M00148"),
    "M00220": ("流式抗体面板归属错位（D2 门）：ITGAM/CD11b 为髓系 marker，挂 T cells 错误；本文无髓系 ITGAM 行", ""),
    "M00221": (R_FLOW.format(sup="NK M00199/M00203（NCAM1）"), "M00199"),
    # --- 人肺图谱：C 门 7 ---
    "M00164": ("C 门语义重复：与 M00168（classical monocytes/CD14）同源流式抗体行同基因，保留更具体标签行", "M00168"),
    "M00165": ("C 门归属错误：CD19 为 B cell marker，挂 classical and nonclassical monocytes 错误；正确归属 B cells 已有 M00148", "M00148"),
    "M00166": ("C 门归属错误：CD4 为 T cell marker，挂 classical and nonclassical monocytes 错误；正确归属 CD4+ T cells 已有 M00162", "M00162"),
    "M00169": ("C 门归属错误：CD19 为 B cell marker，挂 classical monocytes 错误（流式阴性对照行）；正确归属 B cells 已有 M00148", "M00148"),
    "M00167": ("C 门归属错误：NCAM1 (CD56) 为 NK marker，挂 classical and nonclassical monocytes 错误；正确归属 NK 已有 M00199/M00203", "M00199"),
    "M00170": ("C 门归属错误：NCAM1 为 NK marker，挂 classical monocytes 错误；正确归属 NK 已有 M00199/M00203", "M00199"),
    "M00219": ("C 门语义重复：与 M00162（CD4+ T cells/CD4）同源同基因，保留更具体标签行", "M00162"),
    # --- 人肺图谱：双标签重复 3 ---
    "M00171": ("D 门双标签重复：'Dendritic cell (subtype EREG+)' 与 M00180（EREG+ dendritic）规范化后同一细胞同一基因同一证据（prepended marker 句）；保留作者命名标签行 M00180", "M00180"),
    "M00172": ("D 门双标签重复：与 M00184（IGSF21+ dendritic）重复；保留 M00184", "M00184"),
    "M00173": ("D 门双标签重复：与 M00224（TREM2+ dendritic）重复；保留 M00224", "M00224"),
    # --- 人肺图谱：拼写变体重复 3 ---
    "M00121": ("D5 拼写变体重复：Pi16（EDF4d 鼠式写法）修正后与 M00118（Adventitial fibroblast/PI16）同键；EDF4 'fibroblast-selective markers Pi16' 措辞并入 M00118 升级依据", "M00118"),
    "M00122": ("D5 拼写变体重复：Serpinf1 与 M00119（AdvF/SERPINF1）同键", "M00119"),
    "M00130": ("D5 拼写变体重复：Fgfr4（图注拼讹 Frfr4/Fgfr4）与 M00127（AlvF/FGFR4）同键", "M00127"),
    # --- 人肺图谱：标签分裂重复 1 + 归属移除 1 ---
    "M00141": ("D6 标签分裂重复：'AT2' 与 'Alveolar type 2 (AT2)' 统一标签后与 M00137 同键；Fig.4d 'AT2 marker SFTPC' 措辞并入 M00137 升级依据", "M00137"),
    "M00160": ("D2b 归属移除：MYC 为 bronchial vessel（Bro1/Bro2）specific marker 而非泛支气管内皮 marker（EDF3k 'bronchial vessel-specific markers MYC'）；正确归属以新行 M01893（Bronchial vessel cells/MYC/author_declared）补录", "M01893"),
    # --- 人肺空间图谱：4 ---
    "M00297": ("C 门语义重复：SMG-serous/LPO 与 M00294（SMG serous cells/LPO）同一细胞群同一基因；M00294 升级 author_declared 并合并 locator", "M00294"),
    "M00244": ("D6 命名重复：M00240/M00241 重命名为 IR-fibro 后与 IR-fibro/CCL19 同键；Fig. 2c legend 证据并入 M00240", "M00240"),
    "M00245": ("D6 命名重复：M00240/M00241 重命名为 IR-fibro 后与 IR-fibro/CCL21 同键；Fig. 2c legend 证据并入 M00241", "M00241"),
    "M00227": ("D6 语义重复：Fig. 6 schematic 的 'CD4 T cells' 与 Fig. 4g/Results 的 'CD3+ CD4+ T helper cells' 为同一群（GAIN 内 CD4 T 细胞）；保留正文+IHC 证据更强的 M00228，locator 并入", "M00228"),
    # --- 乳腺癌：2 ---
    "M01306": ("C 门重复+归属偏移：FCRL5 的 marker 声明（'the typical B cell marker FCRL5'）指向 B cells 全体而非 B regulatory cells 亚群（过度具体化）；与 M01305 同一证据句语义重复，保留 M01305；B regulatory cells 定义性 marker 为 FOXP3（M01307）", "M01305"),
    "M01314": ("D 门重复（机械扫描未标记，复核新发现）：与 M01311（Macrophage/CD163）同基因同细胞类型（仅大小写/单复数差异）、同出 'CD163, a key macrophage marker' 句；保留 locator 更全的 M01311", "M01311"),
}

# ---------- import_log ----------

LOG_NOTES = {
    P_LUNG: ("Batch1 复核落表：补录 8（Goblet MUC5B/MUC5AC、Serous LTF/LYZ/BPIFBP1/HP、PLVAP、MYC 归属补录）；"
             "升级 28（B 门 10 + D 门顺带 18）；移除 34（流式错位 19 + C 门 7 + 双标签重复 3 + 拼写变体 3 + "
             "标签分裂 1 + 归属移除 1）；标签修正 3；拼写修正 2（M00131 SLC7A10、M00192 ELN）；行数 113→87"),
    P_SPAT: ("Batch1 复核落表：补录 5（CADM、smooth muscle ACTA2、SCGB1A1、FOXJ1、KRT14）；升级 28；移除 4"
             "（LPO 重复、IR-fibro 命名重复 2、CD4 T cell 语义重复）；标签修正 7；locator 修正 3；"
             "is_pns_cell 系统性修正 72 行（全表误设 true，仅 Schwann 13 行保持 true）；行数 85→86"),
    P_BRCA: ("Batch1 复核落表：升级 8（CD19/FCRL5/CD8A/EPCAM/CD163/CD14/NCAM1/CD4 → author_declared）；"
             "移除 2（M01306 FCRL5 归属偏移、M01314 CD163 重复）；M01315 subtype 简化为 CD163+ state；"
             "行数 15→13"),
    P_KIDN: ("Batch1 复核落表：升级 12（frPT/frTAL 6 + moMAC-HBEGF+ 6）；is_pns_cell 修正 48 行"
             "（verdict D 门结论'肾脏非 PNS'，落表核对时发现全表误设 true，verdict 摘要漏记此项）；"
             "M01578 αSMA→ACTA2 备注补充；行数 48 不变"),
    P_DRG: ("Batch1 复核落表：补录 19（A 门 8：NTRK3/TAC1/MRGPRX1/PRP1/MBP/QKI/LPAR1/APOE；"
            "D 门追加 8：SCN1A/PVALB/NEFH/PIEZO2-H15/SCN10A/NTRK1/PIEZO2-H10/SST；"
            "中等置信 3：CALCA/CALCB/ADCYAP1）；B 门 6 条维持不升级；行数 10→29"),
    P_PILOT: ("Batch1 试点遗留处置：M01516（Pdgfra/endoneurial fibroblasts (EFs)）升级 author_declared"
              "（'the fibroblast marker Pdgfra'）；M01517 Cd34 维持 annotation_marker；无行数变化"),
}

LOG_IMPORTED = {P_LUNG: 8, P_SPAT: 5, P_BRCA: 0, P_KIDN: 0, P_DRG: 19, P_PILOT: 0}


def as_bool_str(v) -> str:
    return "true" if str(v).strip().lower() == "true" else "false"


def main() -> None:
    if not BACKUP.exists():
        shutil.copy2(XLSX, BACKUP)
        print(f"backup -> {BACKUP.name}")
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["markers"]
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}

    def iter_data():
        return ws.iter_rows(min_row=2)

    # 1) 追加 32 行（唯一键 paper_id+cell_type+gene_symbol 查重）
    added = 0
    existing_keys = {
        (r[idx["paper_id"]].value, (r[idx["cell_type"]].value or "").strip().lower(), (r[idx["gene_symbol"]].value or "").strip().lower())
        for r in iter_data() if r[idx["paper_id"]].value
    }
    for row_data in ADD_ROWS:
        key = (row_data["paper_id"], row_data["cell_type"].strip().lower(), row_data["gene_symbol"].strip().lower())
        if key in existing_keys:
            print(f"  skip existing key: {row_data['cell_type']} + {row_data['gene_symbol']}")
            continue
        ws.append([row_data.get(h) for h in header])
        existing_keys.add(key)
        added += 1
        print(f"  + {row_data['marker_id']}: {row_data['cell_type']} / {row_data['gene_symbol']}")
    print(f"[1] added {added}/{len(ADD_ROWS)} rows")

    # 2) 升级 77 行
    upgraded = 0
    for row in iter_data():
        mid = row[idx["marker_id"]].value
        if mid in UPGRADES:
            up = UPGRADES[mid]
            old_ev = row[idx["evidence_type"]].value
            row[idx["evidence_type"]].value = "author_declared"
            row[idx["source_context"]].value = up["ctx"]
            row[idx["source_locator"]].value = up["locator"]
            row[idx["notes"]].value = up["note"]
            old_an = row[idx["audit_notes"]].value or ""
            row[idx["audit_notes"]].value = f"{old_an}; {METHOD} evidence_type 升级（{old_ev} → author_declared）".lstrip("; ")
            rm = row[idx["review_method"]].value or ""
            row[idx["review_method"]].value = f"{rm}; {METHOD}".lstrip("; ")
            upgraded += 1
            print(f"  ↑ {mid} ({old_ev} → author_declared)")
    print(f"[2] upgraded {upgraded}/{len(UPGRADES)} rows")

    # 3) 字段修正
    fixed = 0
    for row in iter_data():
        mid = row[idx["marker_id"]].value
        if mid not in FIELD_FIXES:
            continue
        fx = FIELD_FIXES[mid]
        for col in ("cell_type", "subtype", "gene_symbol", "species", "source_locator"):
            if col in fx:
                row[idx[col]].value = fx[col]
        if "notes_append" in fx:
            old = row[idx["notes"]].value or ""
            row[idx["notes"]].value = f"{old}；{fx['notes_append']}".lstrip("；")
        if "audit_notes" in fx:
            old_an = row[idx["audit_notes"]].value or ""
            row[idx["audit_notes"]].value = f"{old_an}; {fx['audit_notes']}".lstrip("; ")
        fixed += 1
        print(f"  ✓ {mid}: {', '.join(k for k in fx if k != 'audit_notes')}")
    print(f"[3] field fixes {fixed}/{len(FIELD_FIXES)} rows")

    # 4) is_pns_cell 系统性修正
    pns_fixed = 0
    for row in iter_data():
        mid = row[idx["marker_id"]].value
        paper = row[idx["paper_id"]].value
        if paper == P_SPAT and mid not in SPATIAL_PNS_KEEP:
            if as_bool_str(row[idx["is_pns_cell"]].value) == "true":
                row[idx["is_pns_cell"]].value = "false"
                old_an = row[idx["audit_notes"]].value or ""
                row[idx["audit_notes"]].value = (f"{old_an}; {METHOD} is_pns_cell 修正 true→false"
                                                 "（提取期系统性错误，白名单规则）").lstrip("; ")
                pns_fixed += 1
        elif paper == P_KIDN:
            if as_bool_str(row[idx["is_pns_cell"]].value) == "true":
                row[idx["is_pns_cell"]].value = "false"
                old_an = row[idx["audit_notes"]].value or ""
                row[idx["audit_notes"]].value = (f"{old_an}; {METHOD} is_pns_cell 修正 true→false"
                                                 "（肾脏非 PNS，落表核对时发现的系统性提取错误）").lstrip("; ")
                pns_fixed += 1
    print(f"[4] is_pns_cell fixed {pns_fixed} rows (expect 72+48=120)")

    # 5) 移除 40 行 → audit_exclusions（先归档，再按行号降序删除）
    ex = wb["audit_exclusions"]
    removed_rows = []
    for row in list(iter_data()):
        mid = row[idx["marker_id"]].value
        if mid in REMOVALS:
            removed_rows.append((row[0].row, {h: row[idx[h]].value for h in header}))
    for _, values in removed_rows:
        reason, sup = REMOVALS[values["marker_id"]]
        reason = f"{reason}。（superseded by: {sup}）" if sup else f"{reason}。"
        ex.append([
            values["paper_id"], values["task_no"], values["cell_type"], values["subtype"],
            values["species"], values["original_symbol"], values["gene_symbol"],
            values["normalization_status"], values["evidence_type"], values["marker_polarity"],
            "exclude", reason, values["source_locator"], values["source_context"],
            None, values["citation_verified"], AUDIT_MODEL, values["marker_id"], REMOVE_TAG,
        ])
    for row_no in sorted((r[0] for r in removed_rows), reverse=True):
        ws.delete_rows(row_no)
    print(f"[5] removed {len(removed_rows)}/{len(REMOVALS)} rows -> audit_exclusions")

    # 6) import_log（6 条）
    log = wb["import_log"]
    log_header = [c.value for c in log[1]]
    title_map = {}
    for r in log.iter_rows(min_row=2, values_only=True):
        if r[2]:
            title_map[r[2]] = r[4]
    for paper, meta in PAPER_META.items():
        log_row = {h: None for h in log_header}
        log_row.update({
            "batch_id": BATCH, "task_no": meta["task_no"], "paper_id": paper,
            "document_id": paper, "paper_title": title_map.get(paper),
            "review_method": METHOD, "imported_count": LOG_IMPORTED[paper],
            "source_file": f"marker提取/audits/recheck-2026-09-02/batch1_work/{paper}_verdict.md",
            "imported_at": TODAY, "notes": LOG_NOTES[paper],
        })
        log.append([log_row.get(h) for h in log_header])
    log_row = {h: None for h in log_header}
    log_row.update({
        "batch_id": BATCH, "task_no": 38, "paper_id": P_PILOT, "document_id": P_PILOT,
        "paper_title": title_map.get(P_PILOT), "review_method": METHOD,
        "imported_count": 0, "source_file": "marker提取/audits/recheck-2026-09-02/batch1_work/DOI_10.7554_elife.71752_verdict.md",
        "imported_at": TODAY, "notes": LOG_NOTES[P_PILOT],
    })
    log.append([log_row.get(h) for h in log_header])
    print("[6] import_log +6 entries")

    # 7) 说明与统计
    st = wb["说明与统计"]
    st.append(["Batch 1 复核轮说明（recheck-2026-09-02，五篇范例，应用日期 2026-09-03）"])
    st.append([
        "五篇 verdict 全量落表：补录 32（肺图谱 8 / 空间图谱 5 / DRG 19）；升级 77（含试点遗留 M01516）；"
        "移除 40（肺图谱 34 / 空间图谱 4 / 乳腺癌 2，均入 audit_exclusions）；标签/拼写/subtype/locator 修正 15；"
        "is_pns_cell 系统性修正 120 行（空间图谱 72 + 肾脏 48）。总行数 1882 → 1874。"
        "判定依据：marker提取/audits/recheck-2026-09-02/batch1_work/*_verdict.md。"
    ])

    # 8) 校验（唯一键断言限定本批论文；批次外遗留重复仅告警，移交后续批次处理）
    n = sum(1 for _ in iter_data())
    assert n == 1874, f"row count mismatch: {n} != 1874"
    batch_papers = set(PAPER_META) | {P_PILOT}
    dup = {}
    legacy = {}
    for r in iter_data():
        if not r[idx["paper_id"]].value:
            continue
        paper = r[idx["paper_id"]].value
        k = (paper, (r[idx["cell_type"]].value or "").strip().lower(),
             (r[idx["gene_symbol"]].value or "").strip().lower())
        target = dup if paper in batch_papers else legacy
        target.setdefault(k, 0)
        target[k] += 1
    dups = {k: v for k, v in dup.items() if v > 1}
    assert not dups, f"unique key violations (batch papers): {dups}"
    legacy_dups = {k: v for k, v in legacy.items() if v > 1}
    if legacy_dups:
        print(f"[8] WARNING: {len(legacy_dups)} legacy duplicate keys outside batch 1 "
              "(pre-existing, deferred to later batches):")
        for k, v in sorted(legacy_dups.items()):
            print(f"    {v}x {k}")
    print(f"[8] verification OK: {n} rows, no unique-key duplicates within batch 1")

    wb.save(XLSX)
    print("saved.")


if __name__ == "__main__":
    main()
