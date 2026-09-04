"""从 our_markers.xlsx 生成同学模板格式的交付表（our_markers_by_cell.xlsx）。

行粒度：一行 = 一篇论文的一个作者细胞标签，Marker 按作者顺序用中文分号串联。
缺口列（谱系/发育阶段/疾病/代码网址）留空待补；生成后做行数与基因数对账。
"""

from __future__ import annotations

import argparse
import json
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
MARKER_DIR = PROJECT_ROOT / "marker提取"
MASTER_XLSX = MARKER_DIR / "表单" / "our_markers.xlsx"
METADATA_DIR = MARKER_DIR / "article_metadata" / "output"
DEFAULT_OUTPUT = MARKER_DIR / "表单" / "our_markers_by_cell.xlsx"

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
LOGGER = logging.getLogger(__name__)

EVIDENCE_LABELS = {
    "author_declared": "作者声明Marker",
    "annotation_marker": "注释用Marker",
    "figure_labeled": "图中标注Marker",
    "supplementary_marker": "补充材料Marker",
}

TISSUE_ZH = {
    "lung": "肺",
    "bronchus": "支气管",
    "lung epithelium": "肺上皮",
    "alveolar sac": "肺泡",
    "islet of Langerhans": "胰岛",
    "blood": "血液",
    "kidney": "肾脏",
    "liver": "肝脏",
    "bone marrow": "骨髓",
    "choroid plexus": "脉络丛",
    "carina of trachea": "气管隆突",
    "nasopharynx": "鼻咽",
    "urinary bladder": "膀胱",
    "dome of urinary bladder": "膀胱顶",
    "prostate gland": "前列腺",
    "peripheral zone of prostate": "前列腺外周带",
    "abdomen": "腹部",
    "anterior wall of left ventricle": "左心室前壁",
    "dorsal root ganglion": "背根神经节",
    "mammary gland": "乳腺",
    "embryo": "胚胎",
    "adipose tissue": "脂肪组织",
}

SPECIES_ZH = {"human": "人", "mouse": "小鼠", "rat": "大鼠", "other": "其他", "unknown": "未知"}

MARKER_HEADERS = ["文章标题", "DOI", "谱系", "组织/区域", "发育阶段", "疾病/状态", "作者细胞名称", "Marker", "依据"]
PANEL_HEADERS = ["文章标题", "补充图/主图面板", "谱系", "组织/区域", "发育阶段", "疾病/状态", "作者注释代码/数据网址"]


def doi_from_paper_id(paper_id: str) -> str:
    if paper_id.startswith("DOI_"):
        rest = paper_id[4:]
        return rest.replace("_", "/", 1) if "_" in rest else rest
    if paper_id.startswith("PMID_"):
        return f"PMID: {paper_id[5:]}"
    return paper_id


def load_paper_info() -> dict[str, dict]:
    wb = load_workbook(MASTER_XLSX, read_only=True)
    ws = wb["audit_summary"]
    rows = list(ws.iter_rows(values_only=True))
    idx = {n: i for i, n in enumerate(rows[0])}
    papers: dict[str, dict] = {}
    for r in rows[1:]:
        papers[r[idx["paper_id"]]] = {
            "title": (r[idx["paper_title"]] or "").strip(),
            "tissue": (r[idx["tissue"]] or "").strip(),
        }

    # 历史冻结行（如 elife.71752）不在终审范围，标题从 import_log 兜底。
    log_rows = list(wb["import_log"].iter_rows(values_only=True))
    log_idx = {n: i for i, n in enumerate(log_rows[0])}
    for r in log_rows[1:]:
        pid = r[log_idx["paper_id"]]
        if pid not in papers:
            papers[pid] = {"title": "", "tissue": ""}
        if not papers[pid]["title"]:
            papers[pid]["title"] = (r[log_idx["paper_title"]] or "").strip()
    wb.close()

    for meta_path in METADATA_DIR.glob("*_metadata.json"):
        pid = meta_path.name[: -len("_metadata.json")]
        if pid in papers and papers[pid]["title"]:
            continue
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            continue
        identity = meta.get("article_identity") or {}
        if pid not in papers:
            papers[pid] = {"title": "", "tissue": ""}
        if not papers[pid]["title"]:
            papers[pid]["title"] = (identity.get("title_original") or "").strip()
    return papers


def main() -> None:
    parser = argparse.ArgumentParser(description="生成同学模板格式的交付表")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    papers = load_paper_info()

    wb_src = load_workbook(MASTER_XLSX, read_only=True)
    ws = wb_src["markers"]
    rows = list(ws.iter_rows(values_only=True))
    idx = {n: i for i, n in enumerate(rows[0])}
    data = rows[1:]
    wb_src.close()

    groups: dict[tuple, dict] = {}
    paper_order: list[str] = []
    for r in data:
        pid = r[idx["paper_id"]]
        if pid not in paper_order:
            paper_order.append(pid)
        cell_type = (r[idx["cell_type"]] or "").strip() or "未命名细胞群"
        subtype = (r[idx["subtype"]] or "").strip()
        key = (pid, cell_type, subtype)
        g = groups.setdefault(
            key,
            {
                "markers": [],
                "evidence": [],
                "locators": [],
                "species": {},
            },
        )
        symbol = (r[idx["gene_symbol"]] or r[idx["original_symbol"]]).strip()
        if r[idx["marker_polarity"]] == "negative":
            symbol += "（阴性）"
        if symbol not in g["markers"]:
            g["markers"].append(symbol)
        label = EVIDENCE_LABELS.get(r[idx["evidence_type"]], r[idx["evidence_type"]])
        if label not in g["evidence"]:
            g["evidence"].append(label)
        loc = (r[idx["source_locator"]] or "").strip()
        if loc and loc not in g["locators"]:
            g["locators"].append(loc)
        sp = SPECIES_ZH.get(r[idx["species"]] or "unknown", "未知")
        g["species"][sp] = g["species"].get(sp, 0) + 1

    # dict 保持插入顺序，即总表导入顺序（marker_id 首次出现顺序）。
    sorted_keys = list(groups)

    out = Workbook()
    ws_m = out.active
    ws_m.title = "Marker表"
    ws_m.append(MARKER_HEADERS)
    marker_total = 0
    species_counter: dict[str, int] = {}
    for key in sorted_keys:
        pid, cell_type, subtype = key
        g = groups[key]
        info = papers.get(pid, {"title": "", "tissue": ""})
        cell_name = f"{cell_type}（{subtype}）" if subtype else cell_type
        tissue = TISSUE_ZH.get(info["tissue"], info["tissue"])
        依据 = "证据：" + "＋".join(g["evidence"])
        if g["locators"]:
            依据 += "；定位：" + "、".join(g["locators"])
        ws_m.append([info["title"], doi_from_paper_id(pid), "", tissue, "", "", cell_name, "；".join(g["markers"]), 依据])
        marker_total += len(g["markers"])
        for sp, n in g["species"].items():
            species_counter[sp] = species_counter.get(sp, 0) + n

    ws_p = out.create_sheet("图表索引")
    ws_p.append(PANEL_HEADERS)
    for pid in paper_order:
        panel_keys = [k for k in sorted_keys if k[0] == pid]
        if not panel_keys:
            continue
        info = papers[pid]
        tissue = TISSUE_ZH.get(info["tissue"], info["tissue"])
        locators: list[str] = []
        for k in panel_keys:
            for loc in groups[k]["locators"]:
                if loc not in locators:
                    locators.append(loc)
        for loc in locators:
            ws_p.append([info["title"], loc, "", tissue, "", "", ""])

    ws_doc = out.create_sheet("说明")
    doc_rows = [
        ["文献 Marker 提取模板（交付版）", ""],
        ["项目", "内容"],
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["数据来源", "our_markers.xlsx（基因级总表，1883 条 = 97 条终审冻结 + 1786 条 2026-09-01 恢复轮追加）"],
        ["行粒度", "一行 = 一篇论文的一个作者细胞标签；Marker 去重后按出现顺序用中文分号串联"],
        ["Marker定义", "仅收录作者作为 Marker 呈现、且实际用于细胞群识别、命名或归类的基因；两个条件缺一不可。"],
        ["证据规则", "证据优先级：文章图表与图注 > 补充表和方法 > 作者官方代码/数据 > 外部知识。"],
        ["缺失值规则", "文章未说明的信息留空，不根据常识补写。"],
        ["待补列", "谱系、发育阶段、疾病/状态、作者注释代码/数据网址——原始管线未提取，当前留空。"],
        ["阴性标注", "作者标注为阴性/不表达的 Marker 以“（阴性）”后缀标注（共 95 条）。"],
        ["物种说明", "模板未设物种列；本表数据为 人 1498 / 小鼠 318 / 其他 67（按基因条数）。"],
        ["图表索引说明", "“文章标题”列为多论文合并而加；其余列与同学模板一致。"],
        ["Marker总数", str(marker_total)],
        ["面板—作者标签记录数", str(len(sorted_keys))],
        ["论文数", str(len({k[0] for k in sorted_keys}))],
        ["版本规则", "本表由脚本从总表自动导出，可随时重建；总表 our_markers.xlsx 为唯一真源。"],
    ]
    for r in doc_rows:
        ws_doc.append(r)

    header_fill = PatternFill("solid", fgColor="EAF1F8")
    header_font = Font(bold=True)
    for sheet, widths in (
        (ws_m, [40, 26, 10, 14, 10, 10, 34, 52, 60]),
        (ws_p, [40, 40, 10, 14, 10, 10, 26]),
        (ws_doc, [18, 100]),
    ):
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")
        for i, w in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = w
        sheet.freeze_panes = "A2"
    ws_m.freeze_panes = "C2"
    for row in ws_m.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column in (7, 8, 9)))
    for row in ws_p.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column in (1, 2)))

    args.output.parent.mkdir(parents=True, exist_ok=True)
    out.save(args.output)

    total_rows = len(data)
    LOGGER.info(
        "生成 %s：细胞标签 %d 行 / 论文 %d 篇 / Marker %d 条（对账 %s）",
        args.output,
        len(sorted_keys),
        len({k[0] for k in sorted_keys}),
        marker_total,
        "一致" if marker_total == total_rows else f"不一致！总表 {total_rows}",
    )
    if marker_total != total_rows:
        raise SystemExit(1)
    LOGGER.info("物种分布: %s", " / ".join(f"{k} {v}" for k, v in sorted(species_counter.items())))


if __name__ == "__main__":
    main()
