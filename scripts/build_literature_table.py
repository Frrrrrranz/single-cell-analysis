"""从数据集_按物种组织分类.xlsx 构建文献-细胞类型-工具一览表"""
import openpyxl
from openpyxl import Workbook

src = 'D:/OneDrive/Desktop/组/db/数据集_按物种组织分类.xlsx'
out = 'D:/OneDrive/Desktop/组/db/文献_细胞类型_工具一览表.xlsx'

wb = openpyxl.load_workbook(src, data_only=True)
ws = wb['全部']
headers = [c.value for c in ws[1]]

# 列索引
col_map = {h: i for i, h in enumerate(headers)}

# 需要的列
output_cols = [
    'Publication_Title',         # 论文标题
    '中文标题',                   # 中文标题  
    'Year',                       # 年份
    'PMID/DOI',                   # PMID/DOI
    'Species',                    # 物种
    '组织系统大类',                # 组织分类
    '主组织',                     # 主组织
    '细胞分层',                   # L1/L2/L3/L4
    '分层依据(具体细胞)',           # 具体细胞类型
    '外周相关细胞(规范)',           # 规范化的外周相关细胞
    'Technology',                 # 技术/工具
    'Marker_Genes',               # Marker 基因
    '代表Dataset_ID',             # Dataset ID
]

# 创建输出工作簿
out_wb = Workbook()
out_ws = out_wb.active
out_ws.title = '文献-细胞类型-工具'

# 写表头
out_ws.append(output_cols)

# 写数据
row_count = 0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    # 跳过空行
    if not any(c is not None for c in row):
        continue
    
    data_row = []
    for col_name in output_cols:
        val = row[col_map[col_name]]
        if val is None:
            data_row.append('NA')
        else:
            data_row.append(str(val).strip())
    out_ws.append(data_row)
    row_count += 1

# 调整列宽
for col in out_ws.columns:
    max_len = 0
    col_letter = col[0].column_letter
    for cell in col:
        if cell.value:
            # 中文字符算2个宽度
            cell_len = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
            max_len = max(max_len, min(cell_len, 60))
    out_ws.column_dimensions[col_letter].width = max_len + 2

out_wb.save(out)
wb.close()
out_wb.close()

print(f'已生成: {out}')
print(f'总行数（含表头）: {row_count + 1}')
print(f'实际文献数: {row_count}')