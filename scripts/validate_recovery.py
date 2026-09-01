"""B-lite 恢复轮专项验收（2026-09-01）。

检查项：
1. 现有 97 行冻结零变动（对照 frozen_baseline.json）；
2. 追加行全部满足 include 不变量（正式证据/符号唯一/引用核验/物种确定）；
3. marker_id 无重复、去重键无重复；
4. 池记录全部有去向（追加/合并/重复/降级），无悬空；
5. verify JSON 与总表、diff CSV 三方对账一致；
6. 排除表 outcome 覆盖完整。
"""

from __future__ import annotations

import csv
import json
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from build_recovery_pool import FORMAL_EVIDENCE_TYPES
from run_recovery_verify import load_candidates

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
MARKER_DIR = PROJECT_ROOT / "marker提取"
AUDIT_DIR = MARKER_DIR / "audited-extraction" / "markers"
RECOVERY_DIR = MARKER_DIR / "audited-extraction" / "recovery"
MASTER_XLSX = MARKER_DIR / "表单" / "our_markers.xlsx"

FOUR_LAYERS = {"L1", "L2", "L3", "L4", "outside", "unknown"}
RECOVERY_SOURCES = {"A_exclude", "A_downgraded", "B_unaudited", "new_finding"}

errors: list[str] = []


def check(condition: bool, message: str) -> None:
    if not condition:
        errors.append(message)


def dedup_key(paper_id, cell_type, subtype, species, gene_symbol, polarity) -> tuple:
    return (
        paper_id,
        str(cell_type or "").strip().lower(),
        str(subtype or "").strip().lower(),
        species,
        str(gene_symbol or "").strip(),
        polarity or "unknown",
    )


def main() -> None:
    # ---- 加载 ----
    verify_paths = sorted(RECOVERY_DIR.glob("*_verify.json"))
    check(len(verify_paths) == 40, f"verify JSON 数量 {len(verify_paths)} != 40")
    verifies = {
        p.name.replace("_verify.json", ""): json.loads(p.read_text(encoding="utf-8"))
        for p in verify_paths
    }

    diff_path = RECOVERY_DIR / "recovery_diff.csv"
    baseline_path = RECOVERY_DIR / "frozen_baseline.json"
    check(diff_path.exists(), "缺少 recovery_diff.csv")
    check(baseline_path.exists(), "缺少 frozen_baseline.json")
    diff_rows = list(csv.DictReader(diff_path.open(encoding="utf-8-sig"))) if diff_path.exists() else []
    baseline = json.loads(baseline_path.read_text(encoding="utf-8")) if baseline_path.exists() else {}

    wb = load_workbook(MASTER_XLSX, read_only=True)
    ws = wb["markers"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    col = {name: i for i, name in enumerate(headers)}
    data_rows = [r for r in rows[1:] if r[col["marker_id"]]]
    wb.close()

    # ---- 1. 冻结零变动 ----
    frozen_count = baseline.get("frozen_count", 97)
    orig_width = len(baseline.get("headers", headers))
    for i in range(frozen_count):
        current = [
            str(v) if v is not None else None for v in (data_rows[i][:orig_width] if i < len(data_rows) else [])
        ]
        saved = [str(v) if v is not None else None for v in baseline["rows"][i]]
        check(current == saved, f"冻结行 {i + 1} 与基线不一致（marker_id={baseline['rows'][i][0]}）")

    # ---- 2/3. 追加行与全表唯一性 ----
    seen_ids: set = set()
    seen_keys: set = set()
    recovery_rows = 0
    frozen_status = Counter()
    for row in data_rows:
        marker_id = row[col["marker_id"]]
        check(marker_id not in seen_ids, f"marker_id 重复: {marker_id}")
        seen_ids.add(marker_id)
        status = row[col["audit_status"]]
        frozen_status[status] += 1
        key = dedup_key(
            row[col["paper_id"]],
            row[col["cell_type"]],
            row[col.get("subtype", -1)],
            row[col["species"]],
            row[col["gene_symbol"]],
            row[col["marker_polarity"]],
        )
        if status == "recovery_include":
            recovery_rows += 1
            check(key not in seen_keys, f"恢复行去重键重复: {key}")
            check(row[col["evidence_type"]] in FORMAL_EVIDENCE_TYPES, f"{marker_id}: 非正式证据")
            check(
                row[col.get("normalization_status", -1)] in ("exact", "alias_resolved"),
                f"{marker_id}: normalization_status 无效",
            )
            check(row[col.get("citation_verified", -1)] is True, f"{marker_id}: 引用未核验")
            check(row[col["species"]] != "unknown", f"{marker_id}: 物种未知")
            check(
                row[col.get("four_layer_category", -1)] in FOUR_LAYERS,
                f"{marker_id}: four_layer_category 无效",
            )
            check(
                row[col.get("recovery_source", -1)] in RECOVERY_SOURCES,
                f"{marker_id}: recovery_source 无效",
            )
        seen_keys.add(key)

    # ---- 4/5. 池记录与 verify 对账 ----
    diff_actions = Counter(r["action"] for r in diff_rows)
    appended_ids = {r["marker_id"] for r in diff_rows if r["action"] == "appended"}
    check(len(appended_ids) == len([r for r in diff_rows if r["action"] == "appended"]), "diff 存在重复 marker_id")
    check(recovery_rows == diff_actions.get("appended", 0), 
          f"总表恢复行 {recovery_rows} != diff appended {diff_actions.get('appended', 0)}")

    # verify include 全部有去向
    pool_total = 0
    for paper_id, data in verifies.items():
        verifications = {v.get("candidate_index"): v for v in data.get("verifications", [])}
        candidates = load_candidates(paper_id)
        pool_total += len(candidates)
        check(
            sorted(verifications) == sorted(c["candidate_index"] for c in candidates),
            f"{paper_id}: verify 候选数量与池不一致",
        )
        for c in candidates:
            v = verifications.get(c["candidate_index"])
            if v is None:
                errors.append(f"{paper_id}: 候选 {c['candidate_index']} 缺少复核结果")
            elif v.get("decision") == "include":
                found = any(
                    r["paper_id"] == paper_id
                    and r["new_decision"] == "include"
                    and r["original_symbol"] == v.get("original_symbol")
                    and r["cell_type"].strip().lower() == str(v.get("cell_type", "")).strip().lower()
                    for r in diff_rows
                )
                check(found, f"{paper_id}: include 候选 {v.get('original_symbol')} 无 diff 记录")
        for nf in data.get("new_findings", []):
            if nf.get("decision") == "include":
                found = any(
                    r["paper_id"] == paper_id
                    and r["source"] == "new_finding"
                    and r["original_symbol"] == nf.get("original_symbol")
                    for r in diff_rows
                )
                check(found, f"{paper_id}: 新发现 include {nf.get('original_symbol')} 无 diff 记录")

    # ---- 6. 排除表 outcome 覆盖 ----
    wb_ex = load_workbook(MASTER_XLSX, read_only=True)
    ws_ex = wb_ex["audit_exclusions"]
    ex_rows = list(ws_ex.iter_rows(values_only=True))
    ex_headers = list(ex_rows[0])
    wb_ex.close()
    check("recovery_outcome" in ex_headers, "audit_exclusions 缺 recovery_outcome 列")
    outcomes = Counter(
        r[ex_headers.index("recovery_outcome")] for r in ex_rows[1:] if r[0]
    )
    no_outcome = sum(1 for r in ex_rows[1:] if r[0] and not r[ex_headers.index("recovery_outcome")])
    check(no_outcome == 0, f"{no_outcome} 条排除行缺少 recovery_outcome")

    print(f"总表行数: {len(data_rows)}（冻结 {frozen_count} + 恢复 {recovery_rows}）")
    print(f"冻结状态分布: {dict(frozen_status)}")
    print(f"diff 动作分布: {dict(diff_actions)}")
    print(f"排除 outcome 分布: {dict(outcomes)}")
    print(f"池候选总数（送检）: {pool_total}")

    if errors:
        print(f"\n验证失败：{len(errors)} 项")
        for e in errors[:50]:
            print(f"  - {e}")
        sys.exit(1)
    print("\n恢复轮专项验证全部通过。")


if __name__ == "__main__":
    main()
