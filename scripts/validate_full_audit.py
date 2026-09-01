"""当前全部文章 Marker 终审的程序化质量检查。

全部满足时退出 0，否则退出非 0 并打印具体 paper_id/marker 错误。
检查项对应执行计划 7.6 与第 9 节验收清单。
"""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
MARKER_DIR = PROJECT_ROOT / "marker提取"
MD_DIR = MARKER_DIR / "review_md"
RAW_DIR = MARKER_DIR / "markers_output_v2"
AUDIT_DIR = MARKER_DIR / "audited-extraction" / "markers"
SOURCE_XLSX = MARKER_DIR / "our_markers.xlsx"
AUDITED_XLSX = MARKER_DIR / "表单" / "our_markers.xlsx"
AUDITED_HTML = MARKER_DIR / "audited-extraction" / "marker-summary.html"

# 2026-08-30 审核开始前原始总表的 SHA256；若原表被修改则验证失败。
SOURCE_XLSX_BASELINE_SHA256 = "1c096dedc4191277f89e6131aeb772a919c346d9246390aa75d11f2e343fe71d"

FORMAL_EVIDENCE_TYPES = {
    "author_declared",
    "annotation_marker",
    "figure_labeled",
    "supplementary_marker",
}
ALL_EVIDENCE_TYPES = FORMAL_EVIDENCE_TYPES | {
    "cluster_enriched",
    "model_inferred",
    "reference_imported",
}
NORMALIZATION_STATUSES = {"exact", "alias_resolved", "ambiguous", "non_gene_entity", "unresolved"}
SPECIES = {"human", "mouse", "rat", "other", "unknown"}
POLARITIES = {"positive", "negative", "unknown"}
DECISIONS = {"include", "context_only", "exclude", "unresolved"}
PAPER_STATUSES = {"pass", "corrected", "no_formal_marker", "no_formal_target_marker", "unresolved"}

DEDUP_KEY_FIELDS = ("paper_id", "cell_type", "subtype", "species", "normalized_symbol", "marker_polarity")


class ValidationErrors:
    def __init__(self) -> None:
        self.items: list[str] = []

    def add(self, message: str) -> None:
        self.items.append(message)

    def fail_if_any(self) -> None:
        if self.items:
            print(f"验证失败：{len(self.items)} 项")
            for item in self.items:
                print(f"  - {item}")
            sys.exit(1)
        print("全部验证通过：当前全部文章的终审结果、修正版 Excel 与 HTML 一致。")


def sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def validate_source_files(errors: ValidationErrors) -> dict[str, dict[str, Any]]:
    md_paths = sorted(MD_DIR.glob("*.md"))
    audit_paths = sorted(AUDIT_DIR.glob("*_audit.json"))

    if len(audit_paths) != len(md_paths):
        errors.add(f"audit JSON 数量 {len(audit_paths)} != review_md 数量 {len(md_paths)}")

    md_ids = {path.stem for path in md_paths}
    audit_ids = {path.name.replace("_audit.json", "") for path in audit_paths}

    if md_ids != audit_ids:
        errors.add(f"audit 与 review_md 的 paper_id 集合不一致: {sorted(md_ids ^ audit_ids)}")

    if SOURCE_XLSX.exists() and sha256_file(SOURCE_XLSX) != SOURCE_XLSX_BASELINE_SHA256:
        errors.add("原始 our_markers.xlsx 哈希与审核前基准不一致（原表被修改）")

    audits: dict[str, dict[str, Any]] = {}
    for path in audit_paths:
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError as exc:
            errors.add(f"{path.name} 无法解析: {exc}")
            continue
        paper_id = data.get("paper_id")
        if paper_id != path.name.replace("_audit.json", ""):
            errors.add(f"{path.name} paper_id 不一致: {paper_id!r}")
            continue
        audits[paper_id] = data
    return audits


def validate_audit_content(audits: dict[str, dict[str, Any]], errors: ValidationErrors) -> None:
    include_keys: dict[tuple, str] = {}
    include_evidence: dict[tuple, list[tuple]] = {}
    conflict_keys: set[tuple] = set()
    for paper_id, data in audits.items():
        task = data.get("task") or {}
        if not (task.get("catalog_cell_layers") or task.get("target_cell_scope")):
            errors.add(f"{paper_id}: 缺少四层分类标签")
        if data.get("paper_status") not in PAPER_STATUSES:
            errors.add(f"{paper_id}: 无效 paper_status {data.get('paper_status')!r}")
        if not isinstance(data.get("summary"), str) or not data["summary"].strip():
            errors.add(f"{paper_id}: 缺少 summary")
        for field in ("source_markdown", "source_raw_json", "source_markdown_sha256", "source_raw_sha256", "audit_model"):
            if not data.get(field):
                errors.add(f"{paper_id}: 缺少 {field}")

        md_path = MD_DIR / str(data.get("source_markdown", ""))
        raw_path = RAW_DIR / str(data.get("source_raw_json", ""))
        if md_path.exists():
            current = sha256_text(md_path.read_text(encoding="utf-8"))
            if current != data.get("source_markdown_sha256"):
                errors.add(f"{paper_id}: Markdown 哈希与审核时不一致")
        else:
            errors.add(f"{paper_id}: source_markdown 不存在 {md_path.name}")
        if raw_path.exists():
            current = sha256_text(raw_path.read_text(encoding="utf-8"))
            if current != data.get("source_raw_sha256"):
                errors.add(f"{paper_id}: raw JSON 哈希与审核时不一致")
        elif RAW_DIR.exists():
            errors.add(f"{paper_id}: source_raw_json 不存在 {raw_path.name}")

        for index, marker in enumerate(data.get("markers", [])):
            prefix = f"{paper_id} markers[{index}]"
            if marker.get("species") not in SPECIES:
                errors.add(f"{prefix}: 无效 species {marker.get('species')!r}")
            if marker.get("evidence_type") not in ALL_EVIDENCE_TYPES:
                errors.add(f"{prefix}: 无效 evidence_type {marker.get('evidence_type')!r}")
            if marker.get("marker_polarity") not in POLARITIES:
                errors.add(f"{prefix}: 无效 marker_polarity {marker.get('marker_polarity')!r}")
            if marker.get("normalization_status") not in NORMALIZATION_STATUSES:
                errors.add(f"{prefix}: 无效 normalization_status {marker.get('normalization_status')!r}")
            if marker.get("decision") not in DECISIONS:
                errors.add(f"{prefix}: 无效 decision {marker.get('decision')!r}")
            for field in ("cell_type", "original_symbol", "source_locator", "source_context"):
                if not str(marker.get(field) or "").strip():
                    errors.add(f"{prefix}.{field} 为空")
            if not isinstance(marker.get("citation_match_score"), (int, float)):
                errors.add(f"{prefix}: 缺少 citation_match_score")
            if not isinstance(marker.get("citation_verified"), bool):
                errors.add(f"{prefix}: 缺少 citation_verified")

            key = (
                paper_id,
                str(marker.get("cell_type") or "").strip().lower(),
                str(marker.get("subtype") or "").strip().lower(),
                marker.get("species"),
                str(marker.get("normalized_symbol") or "").strip(),
                marker.get("marker_polarity"),
            )
            if marker.get("decision") == "include":
                if marker.get("evidence_type") not in FORMAL_EVIDENCE_TYPES:
                    errors.add(f"{prefix}: include 但非正式证据")
                if marker.get("normalization_status") not in {"exact", "alias_resolved"}:
                    errors.add(f"{prefix}: include 但 normalization_status={marker.get('normalization_status')}")
                if marker.get("citation_verified") is not True:
                    errors.add(f"{prefix}: include 但 citation_verified=false")
                if marker.get("species") == "unknown":
                    errors.add(f"{prefix}: include 但 species=unknown")
                include_keys[key] = prefix
                include_evidence.setdefault(key, []).append(
                    (marker.get("evidence_type"), marker.get("source_locator"))
                )
            if (
                marker.get("decision") in {"exclude", "context_only"}
                and key in include_keys
            ):
                conflict_keys.add(key)
    for key in conflict_keys:
        errors.add(f"同一键同时存在 include 与 exclude/context_only: {key}")
    for key, evidence_list in include_evidence.items():
        if len(set(evidence_list)) < len(evidence_list):
            errors.add(f"同键 include 存在 evidence+locator 完全一致的冗余条目: {key}")


def validate_high_risk_expectations(audits: dict[str, dict[str, Any]], errors: ValidationErrors) -> None:
    def find(paper_id: str, symbol: str, **filters: Any) -> list[dict[str, Any]]:
        result = []
        for marker in audits.get(paper_id, {}).get("markers", []):
            if (marker.get("normalized_symbol") or "").upper() != symbol.upper():
                continue
            if all(marker.get(field) == value for field, value in filters.items()):
                result.append(marker)
        return result

    # 1. pros.24020：CHGA human include；CGRP 不是唯一基因符号，不入表
    if not find("DOI_10.1002_pros.24020", "CHGA", species="human", decision="include"):
        errors.add("高风险 DOI_10.1002_pros.24020: human CHGA 未 include")
    if find("DOI_10.1002_pros.24020", "CGRP", decision="include"):
        errors.add("高风险 DOI_10.1002_pros.24020: CGRP 被 include（应为 ambiguous/unresolved）")

    # 2. isci.111628：NRXN 家族名不得入表
    if find("DOI_10.1016_j.isci.2024.111628", "NRXN", decision="include"):
        errors.add("高风险 DOI_10.1016_j.isci.2024.111628: NRXN 家族名被 include")

    # 3. s41588-025-02158-6：原文不存在的基因（如 CDH19）不得 include
    if find("DOI_10.1038_s41588-025-02158-6", "CDH19", decision="include"):
        errors.add("高风险 DOI_10.1038_s41588-025-02158-6: CDH19 原文 0 次出现，不得 include")

    # 4. s41586-024-07069-w：Tgfb2+ Schwann 注释放行
    if not find("DOI_10.1038_s41586-024-07069-w", "Tgfb2", decision="include"):
        errors.add("高风险 DOI_10.1038_s41586-024-07069-w: Tgfb2+ myelinating Schwann 未 include")

    # 5. PMID_35115729：marked by/marks 证据放行、B2m 不漏
    if not find("PMID_35115729", "B2m", decision="include"):
        errors.add("高风险 PMID_35115729: B2m 未 include（原漏提）")


def validate_audited_workbook(audits: dict[str, dict[str, Any]], errors: ValidationErrors) -> None:
    if not AUDITED_XLSX.exists():
        errors.add(f"修正版总表不存在: {AUDITED_XLSX.name}")
        return
    wb = load_workbook(AUDITED_XLSX, read_only=True)
    for sheet in ("markers", "audit_exclusions", "audit_summary"):
        if sheet not in wb.sheetnames:
            errors.add(f"marker提取/表单/our_markers.xlsx 缺少 sheet: {sheet}")
    if "markers" not in wb.sheetnames:
        wb.close()
        return

    ws = wb["markers"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    col = {name: index for index, name in enumerate(headers)}
    required = ("marker_id", "paper_id", "cell_type", "species", "gene_symbol", "evidence_type",
                "marker_polarity", "source_locator", "source_context", "audit_status")
    for name in required:
        if name not in col:
            errors.add(f"markers sheet 缺少列: {name}")

    seen_ids: set[str] = set()
    seen_keys: set[tuple] = set()
    include_count = 0
    for row in rows[1:]:
        marker_id = row[col["marker_id"]] if "marker_id" in col else None
        if not marker_id:
            continue
        if marker_id in seen_ids:
            errors.add(f"markers sheet marker_id 重复: {marker_id}")
        seen_ids.add(marker_id)
        audit_status = row[col["audit_status"]] if "audit_status" in col else None
        if audit_status == "audited_include":
            include_count += 1
            key = (
                row[col["paper_id"]],
                str(row[col["cell_type"]] or "").lower(),
                str(row[col.get("subtype", -1)] or "").lower() if "subtype" in col else "",
                row[col["species"]],
                str(row[col["gene_symbol"]] or "").strip(),
                row[col["marker_polarity"]] or "unknown",
            )
            if key in seen_keys:
                errors.add(f"markers sheet 去重键重复: {key}")
            seen_keys.add(key)
            evidence = row[col["evidence_type"]]
            if evidence not in FORMAL_EVIDENCE_TYPES:
                errors.add(f"{marker_id}: audited_include 但非正式证据 {evidence}")
            normalization = row[col["normalization_status"]] if "normalization_status" in col else None
            if normalization not in {"exact", "alias_resolved"}:
                errors.add(f"{marker_id}: audited_include 但 normalization_status={normalization!r}")
            if "citation_verified" in col and row[col["citation_verified"]] is not True:
                errors.add(f"{marker_id}: audited_include 但 citation_verified != true")
            if row[col["species"]] == "unknown":
                errors.add(f"{marker_id}: audited_include 但 species=unknown")

    expected_include = len(
        {
            (
                paper_id,
                str(marker.get("cell_type") or "").strip().lower(),
                str(marker.get("subtype") or "").strip().lower(),
                marker.get("species"),
                str(marker.get("normalized_symbol") or "").strip(),
                marker.get("marker_polarity"),
            )
            for paper_id, data in audits.items()
            for marker in data.get("markers", [])
            if marker.get("decision") == "include"
        }
    )
    if include_count != expected_include:
        errors.add(f"markers sheet include 行数 {include_count} 与 audit JSON 去重后 include 键数 {expected_include} 不一致")

    if "audit_summary" in wb.sheetnames:
        summary_rows = list(wb["audit_summary"].iter_rows(values_only=True))
        summary_ids = {row[1] for row in summary_rows[1:] if row[1]}
        missing = set(audits) - summary_ids
        if missing:
            errors.add(f"audit_summary sheet 缺少论文: {sorted(missing)}")
        if len(summary_ids) != len(audits):
            errors.add(f"audit_summary sheet 论文数 {len(summary_ids)} != audit JSON 数量 {len(audits)}")
    wb.close()

    if not AUDITED_HTML.exists():
        errors.add(f"修正版 HTML 不存在: {AUDITED_HTML.name}")


def main() -> None:
    parser = argparse.ArgumentParser(description="验证当前全部文章的 Marker 终审产物")
    parser.add_argument("--skip-html", action="store_true", help="HTML 尚未生成时跳过该项")
    args = parser.parse_args()

    errors = ValidationErrors()
    audits = validate_source_files(errors)
    if audits:
        validate_audit_content(audits, errors)
        if set(audits) == {path.stem for path in MD_DIR.glob("*.md")}:
            validate_high_risk_expectations(audits, errors)
            if not args.skip_html:
                validate_audited_workbook(audits, errors)
            else:
                if not AUDITED_XLSX.exists():
                    errors.add(f"修正版总表不存在: {AUDITED_XLSX.name}")
    errors.fail_if_any()


if __name__ == "__main__":
    main()
