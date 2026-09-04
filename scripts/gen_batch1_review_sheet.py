"""生成 Batch 1 同学审阅表（dual 规范 9 列）。

数据源：marker提取/表单/our_markers.xlsx markers sheet（Batch 1 复核落表后状态）
输出：marker提取/.temp/batch1_review_2026-09-02.xlsx

列映射（plan marker-recheck-pipeline-v2 步骤 4）：
- 文章标题 ← article_metadata article_identity.title_original（缺失留空，不推断）
- DOI ← paper_id
- 谱系/组织/区域/发育阶段/疾病/状态 ← 留空（总表无该列，按用户指示不推断）
- 作者细胞名称 ← cell_type（subtype 非空时括注）
- Marker ← 同组 gene_symbol 合并（marker_id 顺序）
- 依据 ← 同组 source_locator 去重合并
分组粒度：paper_id + species + cell_type + subtype（生物学背景含物种，人鼠不合并）
"""

from __future__ import annotations

import json
from collections import OrderedDict
from pathlib import Path

import openpyxl

ROOT = Path(r"d:\OneDrive\Desktop\组\marker提取")
XLSX = ROOT / "表单" / "our_markers.xlsx"
META_DIR = ROOT / "article_metadata" / "output"
OUT_DIR = ROOT / ".temp"
OUT = OUT_DIR / "batch1_review_2026-09-02.xlsx"

PAPERS = [
    "DOI_10.1038_s41586-020-2922-4",
    "DOI_10.1038_s41588-022-01243-4",
    "DOI_10.1038_s41591-024-03215-z",
    "DOI_10.1101_2025.09.26.678707",
    "DOI_10.7554_elife.71752",
]

HEADERS = ["文章标题", "DOI", "谱系", "组织/区域", "发育阶段", "疾病/状态", "作者细胞名称", "Marker", "依据"]


def load_titles() -> dict[str, str]:
    titles = {}
    for pid in PAPERS:
        f = META_DIR / f"{pid}_metadata.json"
        if f.exists():
            d = json.loads(f.read_text(encoding="utf-8"))
            t = (d.get("article_identity") or {}).get("title_original") or ""
            if t:
                titles[pid] = t
    return titles


def main() -> None:
    titles = load_titles()
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb["markers"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(header)}

    groups: "OrderedDict[tuple, list]" = OrderedDict()
    skipped = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        pid = r[idx["paper_id"]]
        if pid not in PAPERS:
            continue
        status = str(r[idx["review_status"]] or "").lower()
        if status != "approved":
            skipped += 1
            continue
        key = (pid, r[idx["species"]], (r[idx["cell_type"]] or "").strip(), (r[idx["subtype"]] or "").strip())
        groups.setdefault(key, []).append({
            "mid": r[idx["marker_id"]],
            "gene": (r[idx["gene_symbol"]] or "").strip(),
            "locator": (r[idx["source_locator"]] or "").strip(),
        })

    OUT_DIR.mkdir(exist_ok=True)
    out = openpyxl.Workbook()
    sheet = out.active
    sheet.title = "Marker表"
    sheet.append(HEADERS)

    rows_written = 0
    for key in sorted(groups, key=lambda k: (PAPERS.index(k[0]), k[2].lower(), k[3].lower())):
        pid, _species, cell_type, subtype = key
        entries = sorted(groups[key], key=lambda e: e["mid"])
        name = f"{cell_type} ({subtype})" if subtype else cell_type
        markers = ", ".join(e["gene"] for e in entries)
        parts = []
        for e in entries:
            for part in str(e["locator"]).split(";"):
                part = part.strip()
                if part:
                    parts.append(part)
        locators = list(OrderedDict.fromkeys(parts))
        sheet.append([
            titles.get(pid, ""), pid, "", "", "", "",
            name, markers, "; ".join(locators),
        ])
        rows_written += 1

    note = out.create_sheet("说明")
    note.append(["Batch 1 复核审阅表"])
    note.append([])
    note.append(["项目", "内容"])
    note.append(["数据来源", "marker提取/表单/our_markers.xlsx（Batch 1 复核落表后，2026-09-03）"])
    note.append(["涵盖论文", "5 篇：人肺图谱 / 人肺空间图谱 / 乳腺癌 MBC / 人肾图谱 v2 / 人 DRG"])
    note.append(["行粒度", "一行 = 一个作者细胞名称（cell_type+subtype）；同组 Marker 合并，依据列为全部出处"])
    note.append(["留空说明", "谱系/组织/区域/发育阶段/疾病/状态在总表中无对应数据列，按规范留空，不做推断"])
    note.append(["标题留空", "DOI_10.7554_elife.71752 无 article_metadata，标题留空"])
    note.append(["生成脚本", "scripts/gen_batch1_review_sheet.py"])

    for col, width in zip("ABCDEFGHI", (48, 34, 8, 10, 10, 12, 34, 46, 60)):
        sheet.column_dimensions[col].width = width

    out.save(OUT)
    print(f"papers: {len(PAPERS)}, groups: {rows_written}, skipped non-approved: {skipped}")
    print(f"titles loaded: {len(titles)}/5")
    print(f"saved -> {OUT}")


if __name__ == "__main__":
    main()
