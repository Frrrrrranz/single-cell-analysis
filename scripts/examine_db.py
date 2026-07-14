"""检查数据库 Excel 文件的结构和内容"""
import openpyxl

files = [
    "d:/OneDrive/Desktop/组/db/pns-scrna.xlsx",
    "d:/OneDrive/Desktop/组/db/数据集_按物种组织分类.xlsx",
    "d:/OneDrive/Desktop/组/db/导师检索_四层_研究级_collection去重.xlsx",
]

for fp in files:
    print(f"\n{'='*80}")
    print(f"FILE: {fp.split('/')[-1]}")
    print('='*80)
    wb = openpyxl.load_workbook(fp, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n  Sheet: [{sn}]  Rows: {ws.max_row}  Cols: {ws.max_column}")
        # Print header row
        headers = [c.value for c in ws[1]]
        print(f"  Headers: {headers}")
        # Print first 5 data rows
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=min(6, ws.max_row), values_only=True)):
            print(f"  Row {i+2}: {row}")
        # Print last 2 rows
        if ws.max_row > 6:
            print(f"  ... (skipping to last 2 rows)")
            for row in ws.iter_rows(min_row=ws.max_row-1, max_row=ws.max_row, values_only=True):
                print(f"  Row {ws.max_row-1}: {row}")
    wb.close()