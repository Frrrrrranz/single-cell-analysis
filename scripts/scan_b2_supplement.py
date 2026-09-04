"""步骤 0.5：B 门补充扫描（2026-09-02 复核轮）。

两个数据源（均为机械过滤，只产候选，不落表）：
- audit_exclusions 中 recovery_outcome ∈ 4 类降级/未决结局行，
  source_context 含 marker 措辞 → B② 身份恢复候选；
- markers 表 evidence_type=supplementary_marker 行，
  source_context 含 marker 措辞 → B① 升级候选（不受定位符守卫，人工判定）。

输出：recheck_plan.xlsx 增补 B2_补充候选 sheet（不改动既有 sheet）。
"""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(r"d:\OneDrive\Desktop\组\marker提取")
XLSX = ROOT / "表单" / "our_markers.xlsx"
PLAN = ROOT / "audits" / "recheck-2026-09-02" / "recheck_plan.xlsx"

MARKER_LANGUAGE = re.compile(
    r"\bmarkers?\b|\bmarked\s+by\b|\bmarks\b|markers?\s+highlighting|"
    r"characteri[sz]ed\s+by|defined\s+by|\bspecified\s+by\b|\bsignature\b",
    re.IGNORECASE,
)

RECOVERY_POOL = {
    "recheck_context_only",
    "recheck_unresolved",
    "new_finding_recheck_unresolved",
    "new_finding_recheck_context_only",
}

SRC_EXCL = "B②身份恢复（exclusions降级行）"
SRC_SUPP = "B①补充（supplementary_marker升级）"
ACT_EXCL = "人工核对：满足双重门槛（marker身份+注释用途）则恢复入表（新发marker_id），否则维持排除并注明"
ACT_SUPP = "人工核对：作者marker措辞明确则升级author_declared，否则维持supplementary_marker"


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    mrows = list(wb["markers"].iter_rows(values_only=True))
    mh = list(mrows[0])
    erows = list(wb["audit_exclusions"].iter_rows(values_only=True))
    eh = list(erows[0])
    wb.close()

    i = lambda h, col: h.index(col)
    paper_genes = defaultdict(set)
    for r in mrows[1:]:
        pid = r[i(mh, "paper_id")]
        for key in (r[i(mh, "gene_symbol")], r[i(mh, "original_symbol")]):
            if key:
                paper_genes[pid].add(str(key).casefold())

    b2 = []
    stats = Counter()

    for r in erows[1:]:
        if r[i(eh, "recovery_outcome")] not in RECOVERY_POOL:
            continue
        stats["excl_pool"] += 1
        ctx = str(r[i(eh, "source_context")] or "")
        if not ctx.strip():
            stats["excl_no_ctx"] += 1
            continue
        if not MARKER_LANGUAGE.search(ctx):
            continue
        stats["excl_hit"] += 1
        pid = r[i(eh, "paper_id")]
        gene = r[i(eh, "normalized_symbol")] or r[i(eh, "original_symbol")] or ""
        in_table = "是" if str(gene).casefold() in paper_genes.get(pid, set()) else "否"
        b2.append([
            SRC_EXCL, "", pid, r[i(eh, "cell_type")], r[i(eh, "subtype")],
            gene, r[i(eh, "original_symbol")], r[i(eh, "evidence_type")],
            r[i(eh, "recovery_outcome")], str(r[i(eh, "source_locator")] or "")[:50],
            ctx[:200], in_table, ACT_EXCL, "待处理",
        ])

    for r in mrows[1:]:
        if r[i(mh, "evidence_type")] != "supplementary_marker":
            continue
        stats["sup_total"] += 1
        ctx = str(r[i(mh, "source_context")] or "")
        if not ctx.strip() or not MARKER_LANGUAGE.search(ctx):
            continue
        stats["sup_hit"] += 1
        b2.append([
            SRC_SUPP, r[i(mh, "marker_id")], r[i(mh, "paper_id")],
            r[i(mh, "cell_type")], r[i(mh, "subtype")], r[i(mh, "gene_symbol")],
            r[i(mh, "original_symbol")], "supplementary_marker", "-",
            str(r[i(mh, "source_locator")] or "")[:50], ctx[:200],
            "-", ACT_SUPP, "待处理",
        ])

    wb2 = openpyxl.load_workbook(PLAN)
    if "B2_补充候选" in wb2.sheetnames:
        del wb2["B2_补充候选"]
    ws = wb2.create_sheet("B2_补充候选")
    header = ["来源", "marker_id", "paper_id", "cell_type", "subtype", "gene",
              "original_symbol", "evidence_type", "recovery_outcome",
              "source_locator", "source_context", "总表已有该基因", "建议动作", "状态"]
    widths = [24, 10, 26, 26, 18, 12, 12, 18, 24, 22, 60, 12, 46, 8]
    ws.append(header)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="4472C4")
    for c in ws[1]:
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(vertical="center")
    for row in b2:
        ws.append(row)
    for n, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(n)].width = w
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = wrap
    wb2.save(PLAN)

    print(f"B2 sheet rows: {len(b2)}")
    print(f"exclusions 降级池: {stats['excl_pool']}（无 context: {stats['excl_no_ctx']}），"
          f"marker 措辞命中: {stats['excl_hit']}")
    print(f"supplementary_marker: {stats['sup_total']}，命中: {stats['sup_hit']}")
    by_paper = Counter(r[2] for r in b2)
    print("--- 按论文分布（并入批次用） ---")
    for p, c in by_paper.most_common():
        print(f"  {c:3d}  {p}")
    print("--- 命中样例（前5） ---")
    for r in b2[:5]:
        print(f"  [{r[0][:12]}] {r[2]} | {r[3]} | {r[5]} | {str(r[10])[:90]}")


if __name__ == "__main__":
    main()
