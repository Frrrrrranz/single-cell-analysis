"""Fuzzy match our 9 papers to literature table"""
import openpyxl

wb = openpyxl.load_workbook(
    r"D:\OneDrive\Desktop\组\db\文献_细胞类型_工具一览表.xlsx",
    data_only=False
)
ws = wb.active

# Our papers with keywords to search in titles
our_papers = {
    "CELL-STEM-CELL.30.20.2023": ["stem cell"],
    "COMMUN-BIOL.5.1105.2022": [],
    "ENEURO.0066-20.2020": ["peripheral nerve", "toma", "eneuro"],
    "FRONT-CELL-NEUROSCI.15.624826.2021": [],
    "GLIA.69.188.2020": ["glia"],
    "J-NEUROINFLAMM.22.205.2025": ["neuroinflamm"],
    "NATURE.587.619.2020": ["nature"],
    "PNAS.117.9466.2020": ["pnas"],
    "SCI-IMMUNOL.8.adf9988.2023": ["early human lung", "sciimmunol"],
}

print("=== Searching literature table for our papers ===")
for row_idx in range(2, ws.max_row + 1):
    title = str(ws.cell(row=row_idx, column=3).value or "").lower()
    source = str(ws.cell(row=row_idx, column=1).value or "")
    cat = str(ws.cell(row=row_idx, column=2).value or "")
    doi = str(ws.cell(row=row_idx, column=6).value or "").lower()

    # Check fill
    fill = ws.cell(row=row_idx, column=1).fill
    fg = fill.fgColor
    highlighted = fg and fg.rgb not in (None, "00000000", "0")

    for pdf_name, keywords in our_papers.items():
        pdf_lower = pdf_name.lower()
        matched = False

        # Try matching parts of the PDF name
        parts = pdf_lower.replace(".full.pdf", "").split(".")
        for part in parts:
            if part in title or part in doi:
                matched = True
                break

        # Try keywords
        if not matched:
            for kw in keywords:
                if kw in title:
                    matched = True
                    break

        if matched:
            flag = "[HIGHLIGHTED - other]" if highlighted else "[UNHIGHLIGHTED - OURS]"
            print(f"  {pdf_name}")
            print(f"    -> Row {row_idx}: {source} | {cat} | {ws.cell(row=row_idx, column=3).value[:70]}")
            print(f"    {flag}")
            break

print()
# Also check if there's a separate tracking mechanism
wb2 = openpyxl.load_workbook(r"D:\OneDrive\Desktop\组\db\pns-scrna.xlsx", data_only=False)
print("=== pns-scrna.xlsx papers sheet ===")
ws2 = wb2["papers"]
h2 = [c.value for c in ws2[1]]
print(f"Headers: {h2}")
# Check cell fills in paper_id column or status column
for row_idx in range(2, ws2.max_row + 1):
    pid = ws2.cell(row=row_idx, column=1).value
    status = ws2.cell(row=row_idx, column=13).value
    title = ws2.cell(row=row_idx, column=2).value
    fill = ws2.cell(row=row_idx, column=1).fill
    fg = fill.fgColor
    highlighted = fg and fg.rgb not in (None, "00000000", "0")
    flag = " [HIGHLIGHTED]" if highlighted else ""
    print(f"  {pid} | {status} | {str(title)[:60]}{flag}")

wb.close()
wb2.close()
