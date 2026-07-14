"""Identify which papers are ours (unhighlighted) vs other student's (highlighted)"""
import openpyxl

wb = openpyxl.load_workbook(
    r"D:\OneDrive\Desktop\组\db\文献_细胞类型_工具一览表.xlsx",
    data_only=True
)
ws = wb.active

# Our 9 downloaded papers from papers/ directory
our_papers = [
    "CELL-STEM-CELL.30.20.2023",
    "COMMUN-BIOL.5.1105.2022",
    "ENEURO.0066-20.2020",
    "FRONT-CELL-NEUROSCI.15.624826.2021",
    "GLIA.69.188.2020",
    "J-NEUROINFLAMM.22.205.2025",
    "NATURE.587.619.2020",
    "PNAS.117.9466.2020",
    "SCI-IMMUNOL.8.adf9988.2023",
]

print("=== Our 9 PDFs in literature table ===")
for row_idx in range(2, ws.max_row + 1):
    title = str(ws.cell(row=row_idx, column=3).value or "")
    doi = str(ws.cell(row=row_idx, column=6).value or "")
    source = str(ws.cell(row=row_idx, column=1).value or "")
    category = str(ws.cell(row=row_idx, column=2).value or "")

    for pdf_name in our_papers:
        # Match by DOI or title keywords
        pdf_upper = pdf_name.upper()
        if pdf_upper in title.upper() or pdf_upper[:10].upper() in doi.upper() or pdf_upper[:10].upper() in title.upper():
            # Check if highlighted
            wb2 = openpyxl.load_workbook(
                r"D:\OneDrive\Desktop\组\db\文献_细胞类型_工具一览表.xlsx",
                data_only=False
            )
            ws2 = wb2.active
            fill = ws2.cell(row=row_idx, column=1).fill
            fg = fill.fgColor
            highlighted = fg and fg.rgb not in (None, "00000000", "0")
            wb2.close()

            flag = "[HIGHLIGHTED - other student]" if highlighted else "[UNHIGHLIGHTED - OURS]"
            print(f"  {pdf_name}")
            print(f"    Row {row_idx}: {source} | {category} | {title[:60]}")
            print(f"    DOI: {doi}")
            print(f"    {flag}")
            break
    else:
        continue

print()
print("=== Summary ===")
# Count total unhighlighted papers
wb2 = openpyxl.load_workbook(
    r"D:\OneDrive\Desktop\组\db\文献_细胞类型_工具一览表.xlsx",
    data_only=False
)
ws2 = wb2.active
unhighlighted_count = 0
for row_idx in range(2, ws2.max_row + 1):
    fill = ws2.cell(row=row_idx, column=1).fill
    fg = fill.fgColor
    if fg and fg.rgb in (None, "00000000", "0"):
        unhighlighted_count += 1
wb2.close()

print(f"Total unhighlighted papers in table: {unhighlighted_count}")
print(f"Our downloaded PDFs: {len(our_papers)}")

wb.close()
