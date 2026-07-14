import openpyxl
import os
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

XLSX_PATH = r"D:\OneDrive\Desktop\组\db\cellxgene\cellxgene_filtered\pns_papers_summary.xlsx"
HTML_PATH = r"D:\OneDrive\Desktop\组\db\cellxgene\cellxgene_filtered\manual_download_helper.html"
DOWNLOADS_DIR = r"D:\OneDrive\Desktop\组\db\cellxgene\cellxgene_filtered\downloads"

def clean_str(s):
    if not s:
        return ""
    return re.sub(r'[^a-z0-9]', '', str(s).lower())

def main():
    print("=== 开始同步 CellxGene 文献状态至 Excel ===")
    
    if not os.path.exists(XLSX_PATH):
        print(f"❌ 找不到 Excel 汇总表: {XLSX_PATH}")
        return
        
    if not os.path.exists(HTML_PATH):
        print(f"❌ 找不到 HTML 辅助文件: {HTML_PATH}")
        return
        
    # 1. 扫描当前 downloads 文件夹下真实存在的文件
    existing_files = os.listdir(DOWNLOADS_DIR)
    existing_files_lower = {f.lower(): f for f in existing_files if f.endswith(".pdf")}
    print(f"当前归库 downloads 目录中共有 {len(existing_files_lower)} 个已就位的 PDF 文件。")
    
    # 2. 从 HTML 辅助文件中解析文献定义 {doi/pmid -> 规范文件名}
    with open(HTML_PATH, "r", encoding="utf-8") as f:
        html_content = f.read()
        
    pattern = r'<div class="card"[^>]*data-title="([^"]*)"[^>]*data-doi="([^"]*)"[^>]*data-pmid="([^"]*)"[^>]*>.*?<span class="paper-index">文献 #(\d+)</span>.*?<div class="filename-value">([^<]*)</div>'
    matches = re.findall(pattern, html_content, re.S)
    
    # 建立以 DOI 后缀、PMID 映射至规范文件名的字典
    paper_to_filename = {}
    for title, doi, pmid, idx, filename in matches:
        idx_num = int(idx)
        paper_to_filename[idx_num] = {
            "idx": idx_num,
            "doi": doi.strip(),
            "pmid": pmid.strip(),
            "filename": filename.strip()
        }

    # 3. 载入 Excel 并修改状态
    wb = openpyxl.load_workbook(XLSX_PATH)
    sh = wb.active
    
    # 解析表头列名映射
    header = [cell.value for cell in sh[1]]
    col_map = {}
    for col_idx, col_name in enumerate(header, 1):
        if col_name:
            col_map[col_name.strip()] = col_idx
            
    # 兼容中英文表头名
    doi_col = col_map.get("doi")
    pmid_col = col_map.get("PMID")
    
    success_col = None
    for name, idx in col_map.items():
        if "是否下载" in name or "success" in name.lower() or "下载成功" in name:
            success_col = idx
            break
            
    path_col = None
    for name, idx in col_map.items():
        if "PDF路径" in name or "path" in name.lower() or "路径" in name:
            path_col = idx
            break
            
    if not success_col or not path_col:
        print("❌ 无法在 Excel 表头中找到 '是否下载成功' 或 '新PDF路径' 等列，请检查表头名。")
        print(f"当前的表头为: {header}")
        return
        
    updated_count = 0
    not_found_count = 0
    already_y_count = 0
    
    for r_idx in range(2, sh.max_row + 1):
        doi_val = str(sh.cell(row=r_idx, column=doi_col).value or "").strip()
        pmid_val = str(sh.cell(row=r_idx, column=pmid_col).value or "").strip()
        
        # 寻找本行在 HTML 里的匹配文献
        matched_paper = None
        for idx_num, paper in paper_to_filename.items():
            # 优先用 DOI 后缀或 PMID 匹配
            if paper["doi"] and paper["doi"] != "-" and paper["doi"] != "无" and doi_val:
                if clean_str(paper["doi"]) == clean_str(doi_val):
                    matched_paper = paper
                    break
            if paper["pmid"] and paper["pmid"] != "-" and paper["pmid"] != "无" and pmid_val:
                if paper["pmid"] == pmid_val:
                    matched_paper = paper
                    break
                    
        if matched_paper:
            target_filename = matched_paper["filename"]
            # 检查这个文件是否真实存在
            if target_filename.lower() in existing_files_lower:
                real_filename = existing_files_lower[target_filename.lower()]
                full_path = os.path.join(DOWNLOADS_DIR, real_filename)
                
                # 获取原状态
                curr_status = sh.cell(row=r_idx, column=success_col).value
                
                # 写入状态和物理绝对路径
                sh.cell(row=r_idx, column=success_col, value="Y")
                sh.cell(row=r_idx, column=path_col, value=full_path)
                
                if curr_status != "Y":
                    print(f"📝 行 {r_idx} [文献 #{matched_paper['idx']}] 状态更新：N -> Y")
                    updated_count += 1
                else:
                    already_y_count += 1
            else:
                sh.cell(row=r_idx, column=success_col, value="N")
                sh.cell(row=r_idx, column=path_col, value=None)
                not_found_count += 1
        else:
            # 如果在 HTML 里没找到对应文献，但根据文件名模糊查询 downloads 下的 PMID 
            found_by_pmid = False
            if pmid_val and pmid_val != "-" and pmid_val != "无" and len(pmid_val) > 4:
                for fname_lower, real_name in existing_files_lower.items():
                    if pmid_val in fname_lower:
                        full_path = os.path.join(DOWNLOADS_DIR, real_name)
                        sh.cell(row=r_idx, column=success_col, value="Y")
                        sh.cell(row=r_idx, column=path_col, value=full_path)
                        updated_count += 1
                        found_by_pmid = True
                        break
            if not found_by_pmid:
                not_found_count += 1

    wb.save(XLSX_PATH)
    print(f"\nExcel 同步更新完成！")
    print(f"   新更新状态为 Y 并且填入路径的有: {updated_count} 篇")
    print(f"   原本就已经是 Y 状态的有: {already_y_count} 篇")
    print(f"   依然处于缺失 (N) 状态的有: {not_found_count} 篇")

if __name__ == "__main__":
    main()
