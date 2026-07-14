import openpyxl

wb = openpyxl.load_workbook(
    r"D:\OneDrive\Desktop\组\db\文献_细胞类型_工具一览表.xlsx",
    data_only=False
)
ws = wb.active
h = [c.value for c in ws[1]]

# Check cell fills/colors
print("Row | 来源 | 分类 | 标题(前50) | Fill Color")
for row_idx in range(2, ws.max_row + 1):
    a = ws.cell(row=row_idx, column=1).value  # 来源
    b = ws.cell(row=row_idx, column=2).value  # 分类
    c = ws.cell(row=row_idx, column=3).value  # 标题
    fill = ws.cell(row=row_idx, column=1).fill
    fg_color = fill.fgColor
    if fg_color:
        color = fg_color.rgb if fg_color.rgb else "none"
    else:
        color = "none"
    highlighted = color not in (None, "00000000", "0", "FFFFFF", "ffffffff")
    flag = " [HIGHLIGHTED]" if highlighted else ""
    print(f"  {row_idx:2d} | {str(a or ''):12s} | {str(b or ''):4s} | {str(c or '')[:50]:50s} | {color}{flag}")

print()
print(f"Total rows: {ws.max_row - 1}")

wb.close()
