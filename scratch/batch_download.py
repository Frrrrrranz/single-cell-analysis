import openpyxl
import requests
from bs4 import BeautifulSoup
import re
import sys
import os
import time
import logging
from scansci_pdf.sources import download as scansciDownload

# 配置日志规范，同时输出到控制台和日志文件
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(r"D:\OneDrive\Desktop\组\paper_search\batch_download.log", encoding="utf-8")
    ]
)

sys.stdout.reconfigure(encoding='utf-8')

# 文件路径配置
EXCEL_PATH = r"D:\OneDrive\Desktop\组\paper_search\爬取失败的论文_补充信息-整理.xlsx"
DOWNLOAD_DIR = r"D:\OneDrive\Desktop\组\paper_search\downloads"
HELPER_HTML_PATH = r"D:\OneDrive\Desktop\组\paper_search\manual_download_helper.html"

# 备用下载源配置
SCIHUB_MIRRORS = [
    "https://sci-hub.ru",
    "https://sci-hub.st",
    "https://sci-hub.se"
]

LIBGEN_MIRRORS = [
    "https://libgen.is",
    "https://libgen.rs",
    "https://libgen.st"
]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}

def clean_title(title):
    if title is None:
        return ""
    return str(title).replace("\n", " ").replace("\r", " ").strip()

def download_file(url, save_path):
    """
    下载文件并验证其是否为合法的 PDF 文件（以 %PDF 开头）
    """
    try:
        r = requests.get(url, headers=HEADERS, timeout=30)
        if r.status_code == 200:
            content = r.content
            if not content.startswith(b'%PDF'):
                logging.warning(f"下载的文件内容头部不是 %PDF (前缀为: {content[:10]}), 链接: {url}")
                return False
            
            # 写入文件
            with open(save_path, 'wb') as f:
                f.write(content)
            logging.info(f"成功保存 PDF: {os.path.basename(save_path)}")
            return True
        else:
            logging.warning(f"下载请求返回状态码异常: {r.status_code}, 链接: {url}")
    except Exception as e:
        logging.error(f"下载文件时发生错误: {e}, 链接: {url}")
    return False

def tryScansciPdf(doiStr: str, savePath: str) -> bool:
    """
    使用 scansci-pdf 强大的多源并行赛跑引擎下载文献，并增加 60 秒的绝对外层超时保护以防假死
    """
    if not doiStr:
        return False
    import concurrent.futures
    try:
        downloadDir = os.path.dirname(savePath)
        # NOTE: 针对 scansci-pdf 接口可能在某些假死 TCP 连接上挂起的情况，使用 ThreadPoolExecutor 强行外加 60 秒的硬性超时限制
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(
                scansciDownload,
                doiStr,
                output_dir=downloadDir,
                scihub_enabled=True,
                rename=False
            )
            try:
                result = future.result(timeout=60)
            except concurrent.futures.TimeoutError:
                logging.warning(f"scansci-pdf 接口调用严重超时（超出了 60 秒的外层硬限制），判定为失败，避免假死。")
                return False
        
        if result and result.get("success"):
            tempPdfPath = result.get("file")
            if tempPdfPath and os.path.exists(tempPdfPath):
                # NOTE: 校验文件魔数，防止下载到 HTML 拦截页或损坏 of 错误页
                with open(tempPdfPath, 'rb') as f:
                    pdfHeader = f.read(4)
                if pdfHeader.startswith(b'%PDF'):
                    if tempPdfPath != savePath:
                        if os.path.exists(savePath):
                            try:
                                os.remove(savePath)
                            except Exception:
                                pass
                        os.rename(tempPdfPath, savePath)
                    logging.info(f"成功保存 PDF: {os.path.basename(savePath)} (来源: {result.get('source')})")
                    return True
                else:
                    logging.warning(f"下载的文件内容头部不是 %PDF (前缀为: {pdfHeader}), 链接: {tempPdfPath}")
                    try:
                        os.remove(tempPdfPath)
                    except Exception:
                        pass
        else:
            logging.warning(f"scansci-pdf 接口下载失败: {result.get('error', 'unknown error') if result else 'no response'}")
    except Exception as e:
        logging.error(f"scansci-pdf 模块运行发生异常: {e}")
    return False

def try_pmc(pmid, save_path):
    """
    第二级：通过 PMID 转换到 PMCID，然后从 PubMed Central 下载 PDF 全文
    """
    if not pmid:
        return False
        
    conv_url = f"https://www.ncbi.nlm.nih.gov/pmc/utils/idconv/v1.0/?ids={pmid}&format=json&tool=pmc_downloader&email=pmc_downloader@example.com"
    logging.info(f"正在尝试 PMC 转换 API 获取 PMID: {pmid}")
    try:
        r = requests.get(conv_url, timeout=15)
        if r.status_code == 200:
            data = r.json()
            records = data.get("records", [])
            if records:
                pmcid = records[0].get("pmcid")
                if pmcid:
                    logging.info(f"[PMC] 查找到对应的 PMCID: {pmcid} (PMID: {pmid})")
                    pdf_url = f"https://www.ncbi.nlm.nih.gov/pmc/articles/{pmcid}/pdf/"
                    logging.info(f"[PMC] 尝试下载 PDF 链接: {pdf_url}")
                    return download_file(pdf_url, save_path)
                else:
                    logging.info(f"[PMC] PMID: {pmid} 未能关联到 PMC 全文 ID")
        else:
            logging.warning(f"[PMC] 请求 API 失败，状态码: {r.status_code}")
    except Exception as e:
        logging.error(f"[PMC] 请求发生异常: {e}")
    return False

def generateManualHtml(failedPapers: list[dict[str, Any]], count: int) -> None:
    """
    为下载失败的论文生成高度美观、功能丰富的 HTML 辅助下载页面
    """
    import urllib.parse
    cardsList: list[str] = []
    
    for idx, paper in enumerate(failedPapers, 1):
        title: str = paper['title']
        doi: str = paper['doi']
        pmid: str = paper['pmid']
        filename: str = paper['filename']
        rowNum: int = paper['rowNum']
        reason: str = paper.get('reason', '自动下载失败')
        
        doiClean: str = doi
        if doiClean.startswith("http"):
            doiClean = doiClean.split("doi.org/")[-1].strip()
            
        scholarQuery: str = urllib.parse.quote(doiClean if doiClean != "-" else title)
        scholarLink: str = f"https://scholar.google.com/scholar?q={scholarQuery}"
        scihubLink: str = f"https://sci-hub.st/{doiClean}" if doiClean != "-" else "https://sci-hub.st/"
        doiLink: str = f"https://doi.org/{doiClean}" if doiClean != "-" else "#"
        
        # NOTE: 构造 PubMed 检索链接。如果有 PMID，直接访问其页面；如果是 "无"，则通过标题检索
        if pmid and pmid != "无" and pmid != "-":
            pubmedLink: str = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
            pubmedBtnText: str = "PubMed 直达"
            pubmedBtnClass: str = "btn-pubmed-direct"
        else:
            pubmedTitleQuery: str = urllib.parse.quote(title)
            pubmedLink = f"https://pubmed.ncbi.nlm.nih.gov/?term={pubmedTitleQuery}"
            pubmedBtnText = "PubMed 搜标题"
            pubmedBtnClass = "btn-pubmed-search"
            
        doiDisplay: str = doiClean if doiClean != "-" else "无"
        pmidDisplay: str = pmid if pmid != "-" else "无"
        
        btnDoiHtml: str = f'<a href="{doiLink}" target="_blank" class="btn btn-doi">打开 DOI 页面</a>' if doiClean != "-" else ''
        btnScihubHtml: str = f'<a href="{scihubLink}" target="_blank" class="btn btn-scihub">Sci-Hub 下载</a>' if doiClean != "-" else ''
        btnPubmedHtml: str = f'<a href="{pubmedLink}" target="_blank" class="btn {pubmedBtnClass}">{pubmedBtnText}</a>'
        
        # 给特定的原因增加特定的 class 以使用不同颜色
        reasonClass: str = "reason-other"
        if "版权" in reason or "闭源" in reason:
            reasonClass = "reason-copyright"
        elif "摘要" in reason:
            reasonClass = "reason-abstract"
        elif "新文" in reason or "202" in reason:
            reasonClass = "reason-new"
        elif "图书" in reason or "章节" in reason:
            reasonClass = "reason-book"
        elif "撤稿" in reason:
            reasonClass = "reason-retracted"
            
        cardHtml: str = f"""
        <div class="card" data-reason="{reason}" data-title="{title.lower()}" data-doi="{doiClean.lower()}" data-pmid="{pmidDisplay.lower()}">
            <div class="card-header">
                <div class="header-left">
                    <span class="paper-index">文献 #{idx}</span>
                    <span class="paper-reason {reasonClass}">{reason}</span>
                </div>
                <span class="paper-row-num">Excel 原始行号: {rowNum}</span>
            </div>
            <h2 class="paper-title" onclick="copyText('{title.replace("'", "\\'")}')" title="点击复制标题">{title}</h2>
            <div class="meta-info">
                <span class="meta-item">DOI: <strong onclick="copyText('{doiClean}')" class="clickable-meta" title="点击复制 DOI">{doiDisplay}</strong></span>
                <span class="meta-item">PMID: <strong onclick="copyText('{pmidDisplay}')" class="clickable-meta" title="点击复制 PMID">{pmidDisplay}</strong></span>
            </div>
            <div class="filename-box" onclick="copyFilename(this, '{filename.replace("'", "\\'")}')" title="点击复制规范文件名">
                <div class="filename-box-title">需保存的规范文件名（点击一键复制）：</div>
                <div class="filename-value">{filename}</div>
                <div class="copy-indicator">点击复制</div>
            </div>
            <div class="action-links">
                <a href="{scholarLink}" target="_blank" class="btn btn-scholar">谷歌学术 检索</a>
                {btnPubmedHtml}
                {btnScihubHtml}
                {btnDoiHtml}
            </div>
        </div>
        """
        cardsList.append(cardHtml)
        
    htmlTemplate: str = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>陈禹臣 - 文献手动下载辅助系统 (精美升级版)</title>
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700&family=Noto+Sans+SC:wght@300;400;500;700&display=swap" rel="stylesheet">
    <style>
        :root {
            --bg-color: #0b0f19;
            --container-bg: #111827;
            --card-bg: rgba(31, 41, 55, 0.6);
            --border-color: rgba(255, 255, 255, 0.08);
            --text-primary: #f3f4f6;
            --text-secondary: #9ca3af;
            --accent-blue: #38bdf8;
            --accent-purple: #a855f7;
            --success-color: #10b981;
            --warning-color: #f59e0b;
            --danger-color: #ef4444;
            --indigo-color: #6366f1;
            --pink-color: #ec4899;
        }
        
        * {
            box-sizing: border-box;
            transition: all 0.2s cubic-bezier(0.4, 0, 0.2, 1);
        }

        body {
            font-family: 'Outfit', 'Noto Sans SC', -apple-system, BlinkMacSystemFont, sans-serif;
            background-color: var(--bg-color);
            color: var(--text-primary);
            margin: 0;
            padding: 40px 20px;
            display: flex;
            flex-direction: column;
            align-items: center;
            background-image: 
                radial-gradient(at 10% 10%, rgba(56, 189, 248, 0.08) 0px, transparent 50%),
                radial-gradient(at 90% 90%, rgba(168, 85, 247, 0.08) 0px, transparent 50%);
            background-attachment: fixed;
        }

        .container {
            max-width: 1200px;
            width: 100%;
        }

        header {
            text-align: center;
            margin-bottom: 40px;
        }

        h1 {
            font-size: 2.8rem;
            font-weight: 700;
            background: linear-gradient(135deg, var(--accent-blue), var(--accent-purple));
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            margin: 0 0 12px 0;
            letter-spacing: -0.5px;
        }

        .subtitle {
            color: var(--text-secondary);
            font-size: 1.15rem;
            max-width: 700px;
            margin: 0 auto;
            line-height: 1.6;
        }

        /* 统计看板 */
        .stats-dashboard {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
            gap: 16px;
            margin-bottom: 35px;
        }

        .stat-card {
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 16px;
            padding: 18px;
            text-align: center;
            cursor: pointer;
            backdrop-filter: blur(12px);
            position: relative;
            overflow: hidden;
        }

        .stat-card:hover {
            transform: translateY(-3px);
            border-color: rgba(255, 255, 255, 0.2);
            box-shadow: 0 10px 20px rgba(0, 0, 0, 0.2);
        }

        .stat-card.active {
            border-color: var(--accent-blue);
            background: rgba(56, 189, 248, 0.08);
            box-shadow: 0 0 15px rgba(56, 189, 248, 0.15);
        }

        .stat-card::before {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            width: 4px;
            height: 100%;
            background: var(--text-secondary);
        }

        .stat-card.type-total::before { background: linear-gradient(to bottom, var(--accent-blue), var(--accent-purple)); }
        .stat-card.type-copyright::before { background: var(--warning-color); }
        .stat-card.type-abstract::before { background: var(--pink-color); }
        .stat-card.type-new::before { background: var(--accent-blue); }
        .stat-card.type-book::before { background: var(--indigo-color); }
        .stat-card.type-retracted::before { background: var(--danger-color); }

        .stat-num {
            font-size: 2.2rem;
            font-weight: 700;
            margin-bottom: 4px;
            font-family: 'Outfit', sans-serif;
        }

        .stat-label {
            font-size: 0.85rem;
            color: var(--text-secondary);
            font-weight: 500;
        }

        /* 搜索和过滤工具栏 */
        .controls-bar {
            display: flex;
            flex-wrap: wrap;
            gap: 16px;
            margin-bottom: 25px;
            align-items: center;
            background: rgba(17, 24, 39, 0.8);
            padding: 16px;
            border-radius: 16px;
            border: 1px solid var(--border-color);
            backdrop-filter: blur(10px);
        }

        .search-wrapper {
            flex: 1;
            min-width: 280px;
            position: relative;
        }

        .search-input {
            width: 100%;
            padding: 12px 16px;
            padding-left: 40px;
            background: rgba(0, 0, 0, 0.25);
            border: 1px solid var(--border-color);
            border-radius: 10px;
            color: var(--text-primary);
            font-size: 0.95rem;
        }

        .search-input:focus {
            outline: none;
            border-color: var(--accent-blue);
            box-shadow: 0 0 10px rgba(56, 189, 248, 0.15);
        }

        .search-icon {
            position: absolute;
            left: 14px;
            top: 50%;
            transform: translateY(-50%);
            color: var(--text-secondary);
            pointer-events: none;
        }

        .quick-actions {
            display: flex;
            gap: 12px;
        }

        .btn-action {
            background: rgba(255, 255, 255, 0.05);
            border: 1px solid var(--border-color);
            color: var(--text-primary);
            padding: 10px 18px;
            border-radius: 10px;
            font-size: 0.9rem;
            font-weight: 500;
            cursor: pointer;
            display: inline-flex;
            align-items: center;
            gap: 8px;
        }

        .btn-action:hover {
            background: rgba(255, 255, 255, 0.1);
            border-color: rgba(255, 255, 255, 0.2);
        }

        .btn-action-primary {
            background: linear-gradient(135deg, var(--accent-blue), var(--accent-purple));
            border: none;
            color: #ffffff;
        }

        .btn-action-primary:hover {
            opacity: 0.9;
            transform: translateY(-1px);
        }

        /* 列表容器 */
        .list-container {
            display: flex;
            flex-direction: column;
            gap: 20px;
        }

        /* 文献卡片 */
        .card {
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 20px;
            padding: 26px;
            backdrop-filter: blur(12px);
            position: relative;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.05);
        }

        .card:hover {
            transform: translateY(-2px);
            box-shadow: 0 12px 30px rgba(0, 0, 0, 0.25);
            border-color: rgba(56, 189, 248, 0.25);
        }

        .card-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 16px;
        }

        .header-left {
            display: flex;
            align-items: center;
            gap: 10px;
        }

        .paper-index {
            background: rgba(56, 189, 248, 0.15);
            color: var(--accent-blue);
            padding: 5px 12px;
            border-radius: 8px;
            font-size: 0.85rem;
            font-weight: 600;
            letter-spacing: 0.5px;
        }

        .paper-reason {
            padding: 5px 12px;
            border-radius: 8px;
            font-size: 0.85rem;
            font-weight: 600;
            border: 1px solid transparent;
        }

        /* 原因标记分类颜色 */
        .reason-copyright {
            background: rgba(245, 158, 11, 0.12);
            color: #fbbf24;
            border-color: rgba(245, 158, 11, 0.25);
        }

        .reason-abstract {
            background: rgba(236, 72, 153, 0.12);
            color: #f472b6;
            border-color: rgba(236, 72, 153, 0.25);
        }

        .reason-new {
            background: rgba(56, 189, 248, 0.12);
            color: #60a5fa;
            border-color: rgba(56, 189, 248, 0.25);
        }

        .reason-book {
            background: rgba(99, 102, 241, 0.12);
            color: #818cf8;
            border-color: rgba(99, 102, 241, 0.25);
        }

        .reason-retracted {
            background: rgba(239, 68, 68, 0.12);
            color: #f87171;
            border-color: rgba(239, 68, 68, 0.25);
        }

        .reason-other {
            background: rgba(156, 163, 175, 0.12);
            color: #d1d5db;
            border-color: rgba(156, 163, 175, 0.25);
        }

        .paper-row-num {
            color: var(--text-secondary);
            font-size: 0.85rem;
            font-weight: 500;
        }

        .paper-title {
            font-size: 1.35rem;
            font-weight: 600;
            line-height: 1.45;
            margin: 0 0 16px 0;
            color: var(--text-primary);
            cursor: pointer;
            display: inline-block;
        }

        .paper-title:hover {
            color: var(--accent-blue);
            text-decoration: underline;
        }

        .meta-info {
            display: flex;
            flex-wrap: wrap;
            gap: 20px;
            margin-bottom: 22px;
            font-size: 0.9rem;
            border-bottom: 1px solid rgba(255, 255, 255, 0.04);
            padding-bottom: 16px;
        }

        .meta-item {
            color: var(--text-secondary);
        }

        .clickable-meta {
            color: var(--text-primary);
            cursor: pointer;
            padding: 2px 6px;
            border-radius: 4px;
            background: rgba(255, 255, 255, 0.04);
        }

        .clickable-meta:hover {
            background: rgba(255, 255, 255, 0.1);
            color: var(--accent-blue);
        }

        /* 规范文件名盒 */
        .filename-box {
            background: rgba(0, 0, 0, 0.2);
            padding: 14px 18px;
            border-radius: 12px;
            border: 1px dashed rgba(255, 255, 255, 0.12);
            font-family: 'Consolas', 'Courier New', monospace;
            font-size: 0.92rem;
            margin-bottom: 22px;
            cursor: pointer;
            position: relative;
            display: flex;
            flex-direction: column;
            gap: 6px;
        }

        .filename-box:hover {
            background: rgba(16, 185, 129, 0.05);
            border-color: rgba(16, 185, 129, 0.4);
        }

        .filename-box-title {
            font-size: 0.75rem;
            color: var(--text-secondary);
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }

        .filename-value {
            color: var(--success-color);
            font-weight: 600;
            word-break: break-all;
            padding-right: 70px;
        }

        .copy-indicator {
            position: absolute;
            right: 18px;
            top: 50%;
            transform: translateY(-50%);
            font-size: 0.8rem;
            background: rgba(16, 185, 129, 0.15);
            color: var(--success-color);
            padding: 4px 10px;
            border-radius: 6px;
            font-weight: 600;
            opacity: 0;
            visibility: hidden;
        }

        .filename-box:hover .copy-indicator {
            opacity: 1;
            visibility: visible;
        }

        /* 操作按钮 */
        .action-links {
            display: flex;
            flex-wrap: wrap;
            gap: 12px;
        }

        .btn {
            display: inline-flex;
            align-items: center;
            padding: 10px 20px;
            border-radius: 10px;
            font-size: 0.9rem;
            font-weight: 600;
            text-decoration: none;
            cursor: pointer;
        }

        .btn-scholar {
            background-color: #1e3a8a;
            color: #93c5fd;
        }
        .btn-scholar:hover {
            background-color: #1e40af;
            box-shadow: 0 0 12px rgba(30, 64, 175, 0.4);
        }

        .btn-pubmed-direct {
            background-color: #312e81;
            color: #c7d2fe;
            border: 1px solid rgba(199, 210, 254, 0.2);
        }
        .btn-pubmed-direct:hover {
            background-color: #3730a3;
            box-shadow: 0 0 12px rgba(55, 48, 163, 0.4);
        }

        .btn-pubmed-search {
            background-color: rgba(99, 102, 241, 0.1);
            color: #a5b4fc;
            border: 1px solid rgba(165, 180, 252, 0.3);
        }
        .btn-pubmed-search:hover {
            background-color: rgba(99, 102, 241, 0.2);
        }

        .btn-scihub {
            background-color: #7c2d12;
            color: #fdba74;
        }
        .btn-scihub:hover {
            background-color: #9a3412;
            box-shadow: 0 0 12px rgba(154, 52, 18, 0.4);
        }

        .btn-doi {
            background-color: #064e3b;
            color: #6ee7b7;
        }
        .btn-doi:hover {
            background-color: #065f46;
            box-shadow: 0 0 12px rgba(6, 95, 70, 0.4);
        }

        /* 全局通知 Toast */
        .toast {
            position: fixed;
            bottom: 30px;
            left: 50%;
            transform: translateX(-50%) translateY(100px);
            background: rgba(16, 185, 129, 0.95);
            color: #ffffff;
            padding: 12px 24px;
            border-radius: 50px;
            font-weight: 600;
            font-size: 0.95rem;
            z-index: 1000;
            box-shadow: 0 10px 25px rgba(0, 0, 0, 0.3);
            opacity: 0;
            pointer-events: none;
            transition: all 0.3s cubic-bezier(0.175, 0.885, 0.32, 1.275);
        }

        .toast.show {
            transform: translateX(-50%) translateY(0);
            opacity: 1;
        }

        /* 响应式调整 */
        @media (max-width: 768px) {
            body {
                padding: 20px 10px;
            }
            h1 {
                font-size: 2rem;
            }
            .controls-bar {
                flex-direction: column;
                align-items: stretch;
            }
            .quick-actions {
                flex-direction: column;
            }
            .action-links .btn {
                flex: 1;
                justify-content: center;
                text-align: center;
            }
        }

        .no-results {
            text-align: center;
            padding: 40px;
            color: var(--text-secondary);
            font-size: 1.1rem;
            background: var(--card-bg);
            border-radius: 16px;
            border: 1px dashed var(--border-color);
        }
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>文献手动下载辅助系统 (精美升级版)</h1>
            <div class="subtitle">针对自动下载失败的 <b>__COUNT__ 篇文献</b>，提供多源快速检索与直达通道。请在<b>校园网环境</b>下通过谷歌学术/PubMed下载，并点击复制对应规范的文件名进行命名保存。</div>
        </header>

        <!-- 统计面板（可作为过滤器） -->
        <div class="stats-dashboard">
            <div class="stat-card active type-total" onclick="filterByReason('all')">
                <div class="stat-num">__COUNT__</div>
                <div class="stat-label">全部失败文献</div>
            </div>
            <div class="stat-card type-copyright" onclick="filterByReason('copyright')">
                <div id="count-copyright" class="stat-num">0</div>
                <div class="stat-label">版权闭源</div>
            </div>
            <div class="stat-card type-abstract" onclick="filterByReason('abstract')">
                <div id="count-abstract" class="stat-num">0</div>
                <div class="stat-label">仅有摘要</div>
            </div>
            <div class="stat-card type-new" onclick="filterByReason('new')">
                <div id="count-new" class="stat-num">0</div>
                <div class="stat-label">2026新文献</div>
            </div>
            <div class="stat-card type-book" onclick="filterByReason('book')">
                <div id="count-book" class="stat-num">0</div>
                <div class="stat-label">图书章节</div>
            </div>
            <div class="stat-card type-retracted" onclick="filterByReason('retracted')">
                <div id="count-retracted" class="stat-num">0</div>
                <div class="stat-label">撤稿文献</div>
            </div>
        </div>

        <!-- 搜索与快捷操作 -->
        <div class="controls-bar">
            <div class="search-wrapper">
                <span class="search-icon">🔍</span>
                <input type="text" id="searchInput" class="search-input" placeholder="输入标题、DOI或PMID进行实时搜索..." oninput="performSearch()">
            </div>
            <div class="quick-actions">
                <button class="btn-action btn-action-primary" onclick="copyAllFilteredDOIs()">复制当前列表的所有 DOI</button>
                <button class="btn-action" onclick="resetAllFilters()">重置所有过滤</button>
            </div>
        </div>

        <!-- 列表容器 -->
        <div class="list-container" id="listContainer">
            __CARDS_HTML__
        </div>
    </div>

    <!-- 弹窗通知 -->
    <div id="toast" class="toast">复制成功！</div>

    <script>
        // 初始化统计数据
        document.addEventListener("DOMContentLoaded", function() {
            updateDashboardCounts();
        });

        // 统一提示框
        function showToast(message) {
            const toast = document.getElementById("toast");
            toast.innerText = message;
            toast.classList.add("show");
            setTimeout(() => {
                toast.classList.remove("show");
            }, 2000);
        }

        // 一键复制纯文本
        function copyText(text) {
            navigator.clipboard.writeText(text).then(() => {
                showToast("文本已复制到剪贴板");
            }).catch(err => {
                console.error("复制失败: ", err);
            });
        }

        // 一键复制规范文件名并改变样式
        function copyFilename(element, filename) {
            navigator.clipboard.writeText(filename).then(() => {
                showToast("规范文件名已复制！");
                
                // 增加成功的闪烁微动效
                element.style.background = "rgba(16, 185, 129, 0.15)";
                element.style.borderColor = "var(--success-color)";
                const indicator = element.querySelector('.copy-indicator');
                if (indicator) {
                    indicator.innerText = "已复制";
                    indicator.style.opacity = "1";
                    indicator.style.visibility = "visible";
                }
                
                setTimeout(() => {
                    element.style.background = "rgba(0, 0, 0, 0.2)";
                    element.style.borderColor = "rgba(255, 255, 255, 0.12)";
                    if (indicator) {
                        indicator.innerText = "点击复制";
                        indicator.style.opacity = "";
                        indicator.style.visibility = "";
                    }
                }, 1500);
            }).catch(err => {
                console.error("复制失败: ", err);
            });
        }

        // 各种失败类型的关键字字典，用于前端统计
        const categoryMap = {
            'copyright': ['版权', '闭源'],
            'abstract': ['摘要'],
            'new': ['新文', '2026', '2025'],
            'book': ['图书', '章节'],
            'retracted': ['撤稿']
        };

        // 动态计算 dashboard 上的数量
        function updateDashboardCounts() {
            const cards = document.querySelectorAll(".card");
            let counts = { copyright: 0, abstract: 0, new: 0, book: 0, retracted: 0 };
            
            cards.forEach(card => {
                const reason = card.getAttribute("data-reason") || "";
                
                for (let key in categoryMap) {
                    const matches = categoryMap[key].some(keyword => reason.includes(keyword));
                    if (matches) {
                        counts[key]++;
                        break; // 归入第一匹配类
                    }
                }
            });

            document.getElementById("count-copyright").innerText = counts.copyright;
            document.getElementById("count-abstract").innerText = counts.abstract;
            document.getElementById("count-new").innerText = counts.new;
            document.getElementById("count-book").innerText = counts.book;
            document.getElementById("count-retracted").innerText = counts.retracted;
        }

        let currentActiveReasonType = 'all';

        // 点击仪表盘卡片进行分类过滤
        function filterByReason(type) {
            currentActiveReasonType = type;
            
            // 更新卡片激活状态
            document.querySelectorAll(".stat-card").forEach(sc => sc.classList.remove("active"));
            
            let activeCardClass = `.type-${type}`;
            if (type === 'all') activeCardClass = '.type-total';
            document.querySelector(activeCardClass).classList.add("active");
            
            performSearch(); // 联合搜索框一起过滤
        }

        // 联合搜索与过滤主函数
        function performSearch() {
            const searchQuery = document.getElementById("searchInput").value.toLowerCase().trim();
            const cards = document.querySelectorAll(".card");
            let visibleCount = 0;

            cards.forEach(card => {
                const reason = card.getAttribute("data-reason") || "";
                const title = card.getAttribute("data-title") || "";
                const doi = card.getAttribute("data-doi") || "";
                const pmid = card.getAttribute("data-pmid") || "";
                
                // 1. 判断是否符合当前选中的原因过滤器
                let matchesReason = false;
                if (currentActiveReasonType === 'all') {
                    matchesReason = true;
                } else {
                    const keywords = categoryMap[currentActiveReasonType];
                    matchesReason = keywords.some(keyword => reason.includes(keyword));
                }

                // 2. 判断是否符合搜索文字
                const matchesSearch = title.includes(searchQuery) || 
                                      doi.includes(searchQuery) || 
                                      pmid.includes(searchQuery) ||
                                      reason.toLowerCase().includes(searchQuery);

                if (matchesReason && matchesSearch) {
                    card.style.display = "block";
                    visibleCount++;
                } else {
                    card.style.display = "none";
                }
            });

            // 如果没有匹配结果，显示提示
            let noResultDiv = document.getElementById("noResultsMsg");
            if (visibleCount === 0) {
                if (!noResultDiv) {
                    noResultDiv = document.createElement("div");
                    noResultDiv.id = "noResultsMsg";
                    noResultDiv.className = "no-results";
                    noResultDiv.innerText = "没有找到符合筛选条件的文献。";
                    document.getElementById("listContainer").appendChild(noResultDiv);
                }
            } else {
                if (noResultDiv) {
                    noResultDiv.remove();
                }
            }
        }

        // 复制当前筛选出的所有 DOI
        function copyAllFilteredDOIs() {
            const cards = document.querySelectorAll(".card");
            let dois = [];
            
            cards.forEach(card => {
                if (card.style.display !== "none") {
                    const doi = card.getAttribute("data-doi");
                    if (doi && doi !== "无" && doi !== "-") {
                        dois.push(doi);
                    }
                }
            });

            if (dois.length === 0) {
                showToast("当前过滤列表无有效 DOI 可供复制");
                return;
            }

            const doiText = dois.join("\\n");
            navigator.clipboard.writeText(doiText).then(() => {
                showToast(`已复制 ${dois.length} 个过滤后的 DOI 至剪贴板 (换行分隔)`);
            });
        }

        // 重置所有过滤
        function resetAllFilters() {
            document.getElementById("searchInput").value = "";
            filterByReason('all');
        }
    </script>
</body>
</html>
"""
    finalHtml: str = htmlTemplate.replace("__COUNT__", str(count)).replace("__CARDS_HTML__", "\n".join(cardsList))
    
    with open(HELPER_HTML_PATH, "w", encoding="utf-8") as f:
        f.write(finalHtml)
        
    logging.info(f"成功生成手动下载辅助网页: {HELPER_HTML_PATH}")


def main():
    logging.info("===== 学术文献批量自动下载服务启动 (升级版) =====")
    
    if not os.path.exists(DOWNLOAD_DIR):
        os.makedirs(DOWNLOAD_DIR)
        logging.info(f"已创建下载目录: {DOWNLOAD_DIR}")
        
    try:
        # 使用 openpyxl 加载 Excel，以便完美保留表格的原生格式
        wb = openpyxl.load_workbook(EXCEL_PATH)
        sheet = wb["manual_download_tasks"]
        logging.info(f"成功读取 Excel 文件，工作表名称: manual_download_tasks")
        
        # 解析表头列名，建立列名到列索引的映射（1-indexed）
        header = [cell.value for cell in sheet[1]]
        col_map = {val: idx for idx, val in enumerate(header, 1)}
        
        # 确保关键列均存在
        required_cols = ["学生", "是否下载成功（Y/N）", "DOI", "PMID", "title", "standard_pdf_filename（用于给下载的PDF文件命名）"]
        for col_name in required_cols:
            if col_name not in col_map:
                raise ValueError(f"Excel 工作表中缺少必要的列: {col_name}")
                
        student_col = col_map["学生"]
        success_col = col_map["是否下载成功（Y/N）"]
        doi_col = col_map["DOI"]
        pmid_col = col_map["PMID"]
        title_col = col_map["title"]
        filename_col = col_map["standard_pdf_filename（用于给下载的PDF文件命名）"]
        row_num_col = col_map.get("original_row_number")
        reason_col = col_map.get("失败原因")
        
        total_tasks = 0
        success_count = 0
        failed_papers_info = []
        
        # 统计陈禹臣的任务总数
        for r_idx in range(2, sheet.max_row + 1):
            if sheet.cell(row=r_idx, column=student_col).value == "陈禹臣":
                total_tasks += 1
                
        logging.info(f"其中属于陈禹臣的文献任务共计 {total_tasks} 篇")
        
        current_idx = 0
        for r_idx in range(2, sheet.max_row + 1):
            student_val = sheet.cell(row=r_idx, column=student_col).value
            if student_val != "陈禹臣":
                continue
                
            current_idx += 1
            
            # 获取各列数据
            doi_val = sheet.cell(row=r_idx, column=doi_col).value
            doi_str = str(doi_val).strip() if doi_val is not None else ""
            if doi_str.startswith("http"):
                doi_str = doi_str.split("doi.org/")[-1].strip()
                
            pmid_val = sheet.cell(row=r_idx, column=pmid_col).value
            pmid_str = ""
            if pmid_val is not None:
                try:
                    pmid_str = str(int(float(pmid_val)))
                except:
                    pmid_str = str(pmid_val).strip()
            
            title_val = sheet.cell(row=r_idx, column=title_col).value
            title_clean = clean_title(title_val)
            
            filename_val = sheet.cell(row=r_idx, column=filename_col).value
            filename_str = str(filename_val).strip() if filename_val is not None else ""
            
            orig_row_val = sheet.cell(row=r_idx, column=row_num_col).value if row_num_col else r_idx
            
            if not filename_str:
                logging.error(f"[{current_idx}/{total_tasks}] 行号 {orig_row_val} 的文献规范文件名缺失，跳过。")
                sheet.cell(row=r_idx, column=success_col, value="N")
                continue
                
            save_path = os.path.join(DOWNLOAD_DIR, filename_str)
            
            logging.info(f"\n--- [{current_idx}/{total_tasks}] 开始处理文献 (Excel行号: {orig_row_val}) ---")
            logging.info(f"标题: {title_clean[:70]}...")
            
            # 断点续传：检查本地文件是否已存在且非空
            if os.path.exists(save_path) and os.path.getsize(save_path) > 10 * 1024:
                logging.info(f"本地已存在大于 10KB 的 PDF 文件，跳过下载。")
                sheet.cell(row=r_idx, column=success_col, value="Y")
                success_count += 1
                continue
                
            downloaded = False
            
            # NOTE: 优先通过 scansci-pdf 并行赛跑下载（DOI 驱动）
            if doi_str:
                downloaded = tryScansciPdf(doi_str, save_path)
                
            # NOTE: 如果 DOI 下载失败或者无 DOI，但在 Excel 中有 PMID，回退到 PMC 渠道下载
            if not downloaded and pmid_str:
                downloaded = try_pmc(pmid_str, save_path)
                
            # 回写结果到单元格
            if downloaded:
                sheet.cell(row=r_idx, column=success_col, value="Y")
                success_count += 1
            else:
                logging.warning(f"文献所有六级通道自动下载均告失败。")
                sheet.cell(row=r_idx, column=success_col, value="N")
                reason_val = sheet.cell(row=r_idx, column=reason_col).value if reason_col else "自动下载失败"
                failed_papers_info.append({
                    "title": title_clean,
                    "doi": doi_str if doi_str else "-",
                    "pmid": pmid_str if pmid_str else "-",
                    "filename": filename_str,
                    "rowNum": orig_row_val,
                    "reason": str(reason_val).strip() if reason_val else "自动下载失败"
                })
            
            # 每一篇操作后，阶段性写入磁盘，防止中途异常崩溃
            if current_idx % 5 == 0:
                wb.save(EXCEL_PATH)
                logging.info("已保存当前进度到 Excel 磁盘文件...")
                
            time.sleep(1.5)
            
        # 最终保存 Excel
        logging.info("\n正在保存最终结果到 Excel...")
        wb.save(EXCEL_PATH)
        logging.info("Excel 数据保存成功，且保留了原生的表格格式样式！")
        
        # 生成手动下载辅助 HTML
        if failed_papers_info:
            generateManualHtml(failed_papers_info, len(failed_papers_info))
        else:
            if os.path.exists(HELPER_HTML_PATH):
                try:
                    os.remove(HELPER_HTML_PATH)
                except:
                    pass
            logging.info("太棒了！所有文献均已成功自动下载，无失败项目。")
            
        logging.info(f"\n===== 批量下载完成统计 =====")
        logging.info(f"总任务数: {total_tasks}")
        logging.info(f"成功下载: {success_count} 篇")
        logging.info(f"失败篇数: {len(failed_papers_info)} 篇")
        logging.info(f"下载文件目录: {DOWNLOAD_DIR}")
        
    except Exception as e:
        logging.critical(f"下载流程发生致命错误: {e}")

if __name__ == "__main__":
    main()
