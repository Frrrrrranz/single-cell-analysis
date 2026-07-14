import os
import re
import fitz
import openpyxl
import urllib.parse

DOWNLOAD_DIR = r"D:\OneDrive\Desktop\组\paper_search\downloads"
EXCEL_PATH = r"D:\OneDrive\Desktop\组\paper_search\爬取失败的论文_补充信息-整理.xlsx"
HTML_PATH = r"D:\OneDrive\Desktop\组\paper_search\manual_download_helper.html"

def clean_doi(doi):
    if not doi:
        return ""
    doi = str(doi).strip().lower()
    if "doi.org/" in doi:
        doi = doi.split("doi.org/")[-1].strip()
    return doi

def extract_doi_from_pdf(filepath):
    doi = None
    try:
        with fitz.open(filepath) as doc:
            # 1. 尝试从元数据中提取 DOI 
            meta = doc.metadata or {}
            for key, val in meta.items():
                if val:
                    # 匹配标准的 DOI 格式 (例如 10.xxxx/xxxx)
                    match = re.search(r'10\.\d{4,9}/[a-zA-Z0-9\./\-_]+', str(val))
                    if match:
                        return match.group(0).rstrip(".-,;")
                        
            # 2. 从前 5 页文本中提取 DOI
            text = ""
            for page_idx in range(min(5, len(doc))):
                text += doc[page_idx].get_text()
                
            # 使用正则在文本中检索 DOI 
            match = re.search(r'10\.\d{4,9}/[a-zA-Z0-9\./\-_]+', text)
            if match:
                doi = match.group(0).rstrip(".-,;")
                if doi.startswith("/"):
                    doi = doi[1:]
    except Exception as e:
        print(f"解析 PDF {os.path.basename(filepath)} 发生异常: {e}")
    return doi

def main():
    print("=== 开始全能解析并整理新下载的文件 ===")
    
    # 1. 读取 Excel 数据
    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet = wb["manual_download_tasks"]
    
    header = [cell.value for cell in sheet[1]]
    col_map = {val: idx for idx, val in enumerate(header, 1) if val is not None}
    
    student_col = col_map.get("学生")
    success_col = col_map.get("是否下载成功（Y/N）")
    doi_col = col_map.get("DOI")
    title_col = col_map.get("title")
    reason_col = col_map.get("失败原因")
    pmid_col = col_map.get("PMID")
    
    # 模糊匹配
    filename_col = None
    for col_name, idx in col_map.items():
        if "standard_pdf_filename" in col_name or "规范" in col_name:
            filename_col = idx
            break
            
    id_col = None
    for col_name, idx in col_map.items():
        if "paper_id" in col_name or "ID" in col_name:
            id_col = idx
            break

    # 建立 DOI, PMID -> task 的映射以提供全方位匹配
    cyc_tasks_by_doi = {}
    cyc_tasks_by_pmid = {}
    cyc_all_tasks = []
    
    for r_idx in range(2, sheet.max_row + 1):
        student_val = sheet.cell(row=r_idx, column=student_col).value
        if student_val == "陈禹臣":
            doi_val = sheet.cell(row=r_idx, column=doi_col).value
            pmid_val = str(sheet.cell(row=r_idx, column=pmid_col).value or "").strip()
            filename_val = sheet.cell(row=r_idx, column=filename_col).value
            success_val = sheet.cell(row=r_idx, column=success_col).value
            task = {
                "row_index": r_idx,
                "paper_id": sheet.cell(row=r_idx, column=id_col).value,
                "title": sheet.cell(row=r_idx, column=title_col).value,
                "doi": doi_val,
                "pmid": pmid_val,
                "filename": filename_val,
                "reason": sheet.cell(row=r_idx, column=reason_col).value or "",
                "excel_status": success_val
            }
            cyc_all_tasks.append(task)
            
            cleaned_doi = clean_doi(doi_val)
            if cleaned_doi:
                cyc_tasks_by_doi[cleaned_doi] = task
            if pmid_val and pmid_val != "-" and pmid_val != "无":
                cyc_tasks_by_pmid[pmid_val] = task
                
    # 2. 扫描 downloads 目录
    files = os.listdir(DOWNLOAD_DIR)
    
    # 找出所有非规范命名的 pdf 文件作为待整理对象
    task_filenames = set(t["filename"] for t in cyc_all_tasks)
    new_files = [f for f in files if f.endswith(".pdf") and f not in task_filenames]
    
    print(f"找到 {len(new_files)} 个待整理的非规范命名 PDF 文件: {new_files}")
    
    renamed_count = 0
    fail_match_count = 0
    
    for f in new_files:
        filepath = os.path.join(DOWNLOAD_DIR, f)
        
        # 1) 从 PDF 内容中提取标识
        pdf_doi = extract_doi_from_pdf(filepath)
        cleaned_pdf_doi = clean_doi(pdf_doi)
        
        # 从 PDF 内容中尝试提取 PMID
        pdf_pmid = None
        try:
            with fitz.open(filepath) as doc:
                pdf_text = ""
                for page_idx in range(min(3, len(doc))):
                    pdf_text += doc[page_idx].get_text()
                pmid_match = re.search(r'PMID:\s*(\d+)', pdf_text, re.IGNORECASE)
                if pmid_match:
                    pdf_pmid = pmid_match.group(1).strip()
        except Exception:
            pass
        
        matched_task = None
        
        # 2) 匹配策略 A: DOI 匹配
        if cleaned_pdf_doi:
            if cleaned_pdf_doi in cyc_tasks_by_doi:
                matched_task = cyc_tasks_by_doi[cleaned_pdf_doi]
            else:
                for db_doi, task in cyc_tasks_by_doi.items():
                    if db_doi in cleaned_pdf_doi or cleaned_pdf_doi in db_doi:
                        matched_task = task
                        break
                        
        # 3) 匹配策略 B: 文本中提取的 PMID 匹配
        if not matched_task and pdf_pmid and pdf_pmid in cyc_tasks_by_pmid:
            matched_task = cyc_tasks_by_pmid[pdf_pmid]
            print(f" -> [文本 PMID 匹配成功] 提取到 PMID {pdf_pmid}")
            
        # 4) 匹配策略 C: 使用文件名本身进行匹配
        if not matched_task:
            fname_clean = os.path.splitext(f)[0].lower().strip()
            # C-1: 如果文件名里有纯数字且长度合理，可能是 PMID
            num_match = re.search(r'\d{7,10}', fname_clean)
            if num_match:
                potential_pmid = num_match.group(0)
                if potential_pmid in cyc_tasks_by_pmid:
                    matched_task = cyc_tasks_by_pmid[potential_pmid]
                    print(f" -> [文件名 PMID 匹配成功] 文件名数字 {potential_pmid} 匹配到任务")
            
            # C-2: 文件名包含 PMCID (如 PMC7940196) 
            if not matched_task and "pmc" in fname_clean:
                pmc_match = re.search(r'pmc\d+', fname_clean)
                if pmc_match:
                    # 我们在 PDF 内容里找 PMID 来辅助
                    try:
                        with fitz.open(filepath) as doc:
                            for page in doc[:3]:
                                txt = page.get_text()
                                p_match = re.search(r'PMID:\s*(\d+)', txt, re.IGNORECASE)
                                if p_match:
                                    p_id = p_match.group(1).strip()
                                    if p_id in cyc_tasks_by_pmid:
                                        matched_task = cyc_tasks_by_pmid[p_id]
                                        print(f" -> [PMCID 关联 PMID 成功] PMC 文件中找到 PMID: {p_id}")
                                        break
                    except Exception:
                        pass
                        
            # C-3: 普通子串匹配 (例如 sciadv.aea2538)
            if not matched_task:
                for db_doi, task in cyc_tasks_by_doi.items():
                    db_doi_clean = db_doi.lower()
                    if fname_clean in db_doi_clean or db_doi_clean in fname_clean:
                        matched_task = task
                        print(f" -> [文件名 DOI 模糊匹配成功] 文件名 {f} 匹配到 DOI: {task['doi']}")
                        break
                        
            # C-4: 文件名单词重合度模糊匹配 (针对未下载的陈禹臣任务)
            if not matched_task:
                stop_words = {'the', 'and', 'for', 'with', 'from', 'pdf', 'journal', 'of', 'in', 'on', 'at', 'by', 'an', 'to', 'is', 'are', 'was', 'were', 'or'}
                words = set([w for w in re.split(r'[_ \-()+,]+', fname_clean) if len(w) > 2 and w not in stop_words])
                
                best_match = None
                max_overlap = 0
                for task in cyc_all_tasks:
                    if task["excel_status"] == "Y":
                        continue
                    db_fname = str(task["filename"]).lower() if task["filename"] else ""
                    db_fname_clean = os.path.splitext(db_fname)[0]
                    db_words = set([w for w in re.split(r'[_ \-()+,]+', db_fname_clean) if len(w) > 2 and w not in stop_words])
                    
                    overlap = words.intersection(db_words)
                    if len(overlap) >= 4 and len(overlap) > max_overlap:
                        max_overlap = len(overlap)
                        best_match = task
                if best_match:
                    matched_task = best_match
                    print(f" -> [文件名单词重合匹配成功] 文件名与规范名重合单词数: {max_overlap}，匹配到任务: {matched_task['filename']}")
                        
        print(f"\n文件: {f}")
        print(f" -> 提取到 DOI: {pdf_doi} | 提取到 PMID: {pdf_pmid}")
        
        if matched_task:
            target_filename = matched_task["filename"]
            target_filepath = os.path.join(DOWNLOAD_DIR, target_filename)
            
            try:
                if os.path.exists(target_filepath):
                    os.remove(target_filepath)
                os.rename(filepath, target_filepath)
                print(f" -> [成功匹配] 已改名为: {target_filename}")
                
                # 更新 Excel 记录
                row_idx = matched_task["row_index"]
                sheet.cell(row=row_idx, column=success_col, value="Y")
                sheet.cell(row=row_idx, column=reason_col, value="")
                
                # 更新内存状态
                matched_task["excel_status"] = "Y"
                renamed_count += 1
            except Exception as ex:
                print(f" -> [改名失败] 移动文件时出错: {ex}")
        else:
            print(f" -> [匹配失败] 无法将该文件关联到任何陈禹臣的文献任务")
            fail_match_count += 1
            
    # 保存 Excel
    if renamed_count > 0:
        wb.save(EXCEL_PATH)
        print(f"\nExcel 数据保存完毕！成功更新了 {renamed_count} 篇文献的状态。")
    else:
        print("\n未进行任何状态更新，无需保存 Excel。")
        
    # 3. 重新生成去重并合并的 HTML
    failed_tasks = [t for t in cyc_all_tasks if t["excel_status"] != "Y"]
    
    # 按照 DOI 或 Title 进行去重合并
    grouped_tasks = {}
    for t in failed_tasks:
        doi_key = clean_doi(t["doi"])
        key = doi_key if (doi_key and doi_key != "-" and doi_key != "无") else t["title"].strip().lower()
        
        if key not in grouped_tasks:
            grouped_tasks[key] = {
                "paper_id": t["paper_id"],
                "title": t["title"],
                "doi": t["doi"],
                "pmid": t.get("pmid", "-"),
                "reason": t["reason"],
                "filenames": [t["filename"]]
            }
        else:
            if t["filename"] not in grouped_tasks[key]["filenames"]:
                grouped_tasks[key]["filenames"].append(t["filename"])
                
    print(f"\n重新生成去重后的 HTML... 共有 {len(grouped_tasks)} 篇独立未下载文献（包含 {len(failed_tasks)} 条任务记录）。")
    
    if os.path.exists(HTML_PATH):
        os.remove(HTML_PATH)
        
    html_content = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>陈禹臣 - 文献手动下载辅助列表</title>
    <style>
        .copy-cell {
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 15px;
        }
        .copy-cell span {
            word-break: break-all;
        }
        .copy-cell button {
            white-space: nowrap;
            flex-shrink: 0;
        }
        tr.checked-row {
            text-decoration: line-through;
            color: #999;
            background-color: #fcfcfc;
        }
        .filename-item {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 4px;
            gap: 10px;
            border-bottom: 1px dashed #eee;
            padding-bottom: 4px;
        }
        .filename-item:last-child {
            border-bottom: none;
            margin-bottom: 0;
            padding-bottom: 0;
        }
    </style>
</head>
<body>
    <h1>陈禹臣 - 文献手动下载辅助列表 (共 {0} 篇独立文献待下载)</h1>
    <p>以下文献在批量下载时失败，请人工下载后使用“规范文件名”命名，并放入 downloads 文件夹中。</p>
    
    <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; width: 100%;">
        <thead>
            <tr style="background-color: #f2f2f2;">
                <th>已下载</th>
                <th>序号</th>
                <th>论文ID</th>
                <th>论文名称 (可复制)</th>
                <th>失败原因</th>
                <th>DOI (可复制)</th>
                <th>规范文件名 (可复制，可能有多个版本)</th>
                <th>PubMed链接</th>
                <th>谷歌学术链接</th>
            </tr>
        </thead>
        <tbody>
"""
    html_content = html_content.replace("{0}", str(len(grouped_tasks)))

    for idx, (key, t) in enumerate(grouped_tasks.items(), 1):
        paper_id = str(t["paper_id"])
        title = str(t["title"])
        reason = str(t["reason"]) if t["reason"] else "自动下载失败"
        doi = str(t["doi"]) if t["doi"] else "-"
        pmid = str(t.get("pmid", "-"))
        filenames = t["filenames"]
        
        if pmid and pmid != "无" and pmid != "-":
            pubmed_url = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
            pubmed_link = f'<a href="{pubmed_url}" target="_blank">PubMed直达</a>'
        else:
            pubmed_title_query = urllib.parse.quote(title)
            pubmed_url = f"https://pubmed.ncbi.nlm.nih.gov/?term={pubmed_title_query}"
            pubmed_link = f'<a href="{pubmed_url}" target="_blank">PubMed搜标题</a>'
            
        doi_clean = doi
        if doi_clean.startswith("http"):
            doi_clean = doi_clean.split("doi.org/")[-1].strip()
        
        scholar_query = urllib.parse.quote(doi_clean if doi_clean != "-" else title)
        scholar_url = f"https://scholar.google.com/scholar?q={scholar_query}"
        scholar_link = f'<a href="{scholar_url}" target="_blank">谷歌学术检索</a>'
        
        title_esc = title.replace("'", "\\'")
        doi_esc = doi.replace("'", "\\'")
        id_esc = paper_id.replace("'", "\\'")
        
        # 拼接文件名列表的 HTML
        filenames_html = ""
        for f_idx, fname in enumerate(filenames, 1):
            fname_esc = fname.replace("'", "\\'")
            prefix_label = f"[{f_idx}] " if len(filenames) > 1 else ""
            filenames_html += f"""
                    <div class="filename-item">
                        <span style="font-family: monospace;">{prefix_label}{fname}</span>
                        <button onclick="copyText('{fname_esc}')">复制文件名</button>
                    </div>"""

        html_content += f"""            <tr>
                <td align="center"><input type="checkbox" onchange="toggleRow(this)"></td>
                <td align="center">{idx}</td>
                <td>
                    <div class="copy-cell">
                        <span>{paper_id}</span>
                        <button onclick="copyText('{id_esc}')">复制ID</button>
                    </div>
                </td>
                <td>
                    <div class="copy-cell">
                        <span>{title}</span>
                        <button onclick="copyText('{title_esc}')">复制标题</button>
                    </div>
                </td>
                <td style="color: red;">{reason}</td>
                <td>
                    <div class="copy-cell">
                        <span>{doi}</span>
                        <button onclick="copyText('{doi_esc}')">复制DOI</button>
                    </div>
                </td>
                <td>
                    {filenames_html}
                </td>
                <td align="center">{pubmed_link}</td>
                <td align="center">{scholar_link}</td>
            </tr>
"""

    html_content += """        </tbody>
    </table>

    <script>
        function toggleRow(cb) {
            const row = cb.closest('tr');
            if (cb.checked) {
                row.classList.add('checked-row');
            } else {
                row.classList.remove('checked-row');
            }
        }

        function copyText(text) {
            if (text === '-' || text === '无') {
                alert('无有效数据可复制');
                return;
            }
            navigator.clipboard.writeText(text).then(() => {
                showToast('已成功复制: ' + text);
            }).catch(err => {
                const input = document.createElement('textarea');
                input.value = text;
                document.body.appendChild(input);
                input.select();
                try {
                    document.execCommand('copy');
                    showToast('已成功复制(备用通道): ' + text);
                } catch (e) {
                    alert('复制失败，请手动选择复制');
                }
                document.body.removeChild(input);
            });
        }
        
        function showToast(msg) {
            const div = document.createElement('div');
            div.style.position = 'fixed';
            div.style.bottom = '20px';
            div.style.left = '50%';
            div.style.transform = 'translateX(-50%)';
            div.style.backgroundColor = 'rgba(0,0,0,0.8)';
            div.style.color = '#fff';
            div.style.padding = '10px 20px';
            div.style.borderRadius = '5px';
            div.style.fontSize = '14px';
            div.style.zIndex = '9999';
            div.innerText = msg;
            document.body.appendChild(div);
            setTimeout(() => {
                document.body.removeChild(div);
            }, 1500);
        }
    </script>
</body>
</html>
"""

    with open(HTML_PATH, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"HTML 页面重写成功: {HTML_PATH}")
    
    # 4. 打印最新缺漏状态
    print(f"\n=== 最新下载情况统计 ===")
    print(f"已经就位并成功重命名的文献共：{100 - len(failed_tasks)} 篇")
    print(f"依然缺漏的文献共：{len(failed_tasks)} 篇")
    
    if failed_tasks:
        print("\n--- 依然缺漏的文献列表 (前 15 篇) ---")
        for idx, t in enumerate(failed_tasks[:15], 1):
            print(f" {idx}. ID: {t['paper_id']} | DOI: {t['doi']} | Title: {t['title']}")
        if len(failed_tasks) > 15:
            print(f" ... 还有 {len(failed_tasks) - 15} 篇")

if __name__ == "__main__":
    main()
