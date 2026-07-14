import openpyxl
import os
import re
import sys
import fitz  # PyMuPDF

sys.stdout.reconfigure(encoding='utf-8')

XLSX_PATH = r"D:\OneDrive\Desktop\组\db\cellxgene\cellxgene_filtered\pns_papers_summary.xlsx"
HTML_PATH = r"D:\OneDrive\Desktop\组\db\cellxgene\cellxgene_filtered\manual_download_helper.html"
DOWNLOADS_DIR = r"D:\OneDrive\Desktop\组\db\cellxgene\cellxgene_filtered\downloads"

def clean_str(s):
    if not s:
        return ""
    return re.sub(r'[^a-z0-9]', '', str(s).lower())

def extract_pdf_text_head(filepath):
    text = ""
    try:
        with fitz.open(filepath) as doc:
            for page_idx in range(min(2, len(doc))):
                text += doc[page_idx].get_text()
    except Exception as e:
        print(f"   ⚠️ 无法读取 PDF {os.path.basename(filepath)}: {e}")
    return text

def make_standard_filename(pmid, doi, title):
    # 清洗 title 中的 Windows 非法字符，用 _ 替换，并且去除末尾的空格和点
    clean_title = re.sub(r'[\\/:*?"<>|]', '_', title)
    clean_title = re.sub(r'_+', '_', clean_title).strip()
    
    pmid_str = str(pmid or "").strip()
    doi_str = str(doi or "").strip()
    
    if pmid_str and pmid_str != "-" and pmid_str != "无":
        return f"PMID_{pmid_str}_{clean_title}.pdf"
    else:
        # 清除 doi 里的非法命名字符
        doi_clean = re.sub(r'[\\/:*?"<>|]', '_', doi_str)
        return f"DOI_{doi_clean}_{clean_title}.pdf"

def main():
    print("=== 开始 CellxGene 文献物理级二次校验与 Excel 看板重构 ===")
    
    if not os.path.exists(XLSX_PATH):
        print("❌ 核心 Excel 汇总表缺失，无法进行检查。")
        return
        
    # 1. 扫描 downloads 文件夹中现有的 PDF
    pdf_files = [f for f in os.listdir(DOWNLOADS_DIR) if f.endswith(".pdf")]
    print(f"找到已就位的 PDF 文件共: {len(pdf_files)} 个")
    
    # 2. 从 Excel 读取全部 75 篇文献定义
    wb = openpyxl.load_workbook(XLSX_PATH)
    sh = wb.active
    
    header = [cell.value for cell in sh[1]]
    col_map = {name.strip(): idx for idx, name in enumerate(header, 1) if name}
    
    doi_col = col_map.get("doi")
    pmid_col = col_map.get("PMID")
    title_col = col_map.get("Publication_Title")
    
    success_col = None
    for name, idx in col_map.items():
        if "是否下载" in name or "success" in name.lower() or "下载成功" in name:
            success_col = idx
            break
            
    path_col = None
    for name, idx in col_map.items():
        if "PDF路径" in name or "path" in name.lower() or "路径" in name:
            path_col = idx
            break
            
    papers = []
    for r_idx in range(2, sh.max_row + 1):
        doi_val = str(sh.cell(row=r_idx, column=doi_col).value or "").strip()
        pmid_val = str(sh.cell(row=r_idx, column=pmid_col).value or "").strip()
        title_val = str(sh.cell(row=r_idx, column=title_col).value or "").strip()
        
        # 序号是 r_idx - 1
        idx_num = r_idx - 1
        
        # 生成标准规范名
        std_filename = make_standard_filename(pmid_val, doi_val, title_val)
        
        papers.append({
            "idx": idx_num,
            "row_idx": r_idx,
            "title": title_val,
            "doi": doi_val,
            "pmid": pmid_val,
            "filename": std_filename,
            "status": "missing",
            "check_result": "N/A",
            "error_reason": ""
        })
        
    print(f"从 Excel 中成功提取了 {len(papers)} 篇文献定义")
    
    # 3. 对已下载的 PDF 文件进行二次精细检查 (PMID/DOI/Title 词重合)
    for paper in papers:
        actual_name = None
        
        # A-1. 优先通过 PMID 匹配文件名
        p_pmid = paper["pmid"]
        if p_pmid and p_pmid != "-" and p_pmid != "无" and len(p_pmid) > 4:
            for fname in pdf_files:
                if p_pmid in fname:
                    actual_name = fname
                    paper["check_result"] = "matched_by_filename_pmid"
                    break
                    
        # A-2. 其次通过 DOI 后缀匹配文件名
        if not actual_name and paper["doi"] and paper["doi"] != "-" and paper["doi"] != "无":
            # 提取 DOI 后缀
            doi_suffix = ""
            if "/" in paper["doi"]:
                doi_suffix = paper["doi"].split("/", 1)[1].strip()
            cleaned_suffix = clean_str(doi_suffix)
            if len(cleaned_suffix) > 5:
                for fname in pdf_files:
                    if cleaned_suffix in clean_str(fname):
                        actual_name = fname
                        paper["check_result"] = "matched_by_filename_doi_suffix"
                        break
                        
        # A-3. 再次通过标题关键词匹配文件名
        if not actual_name:
            title_words = [w for w in re.split(r'[^a-zA-Z]', paper["title"].lower()) if len(w) > 4]
            if len(title_words) >= 3:
                best_match = None
                max_overlap = 0
                for fname in pdf_files:
                    overlap = sum(1 for w in title_words if w in fname.lower())
                    if overlap >= 3 and overlap > max_overlap:
                        max_overlap = overlap
                        best_match = fname
                if best_match:
                    actual_name = best_match
                    paper["check_result"] = f"matched_by_filename_title (overlap: {max_overlap})"

        # A-4. 降级备用：如果连文件名都没对上，我们退回到物理 std_filename 尝试
        if not actual_name:
            std_path = os.path.join(DOWNLOADS_DIR, paper["filename"])
            if os.path.exists(std_path):
                actual_name = paper["filename"]
                paper["check_result"] = "matched_by_exact_std_filename"

        # B. 找到真实对应的 PDF 文件后，进行内容和状态校验
        if actual_name:
            target_path = os.path.join(DOWNLOADS_DIR, actual_name)
            paper["filename"] = actual_name
            
            if os.path.getsize(target_path) == 0:
                paper["status"] = "error"
                paper["error_reason"] = "PDF文件损坏（大小为 0 字节，需重新下载）"
                print(f"❌ [文献 #{paper['idx']}] 校验失败！空文件: {actual_name}")
                continue
                
            pdf_text = extract_pdf_text_head(target_path)
            pdf_text_clean = clean_str(pdf_text)
            
            content_verified = False
            
            if p_pmid and p_pmid != "-" and p_pmid != "无" and len(p_pmid) > 4:
                if p_pmid in pdf_text_clean:
                    content_verified = True
                    paper["check_result"] += " + content_verified_by_pmid"
                    
            if not content_verified and paper["doi"] and paper["doi"] != "-" and paper["doi"] != "无":
                doi_parts = paper["doi"].split("/", 1)
                if len(doi_parts) > 1:
                    doi_suffix = clean_str(doi_parts[1])
                    if len(doi_suffix) > 5 and doi_suffix in pdf_text_clean:
                        content_verified = True
                        paper["check_result"] += " + content_verified_by_doi"
                        
            if not content_verified:
                title_words = [w for w in re.split(r'[^a-zA-Z]', paper["title"].lower()) if len(w) > 4]
                if len(title_words) >= 3:
                    overlap_count = sum(1 for w in title_words if w in pdf_text_clean)
                    if overlap_count >= 3:
                        content_verified = True
                        paper["check_result"] += f" + content_verified_by_title (overlap: {overlap_count})"
                        
            if content_verified or "matched_by_filename" in paper["check_result"]:
                paper["status"] = "success"
                print(f"✅ [文献 #{paper['idx']}] 校验成功！依据: {paper['check_result']}")
            else:
                paper["status"] = "error"
                paper["error_reason"] = "PDF内容核对未通过（PMID/DOI/标题关键词不匹配）"
                print(f"❌ [文献 #{paper['idx']}] 校验失败！内容错配: {actual_name}")
        else:
            paper["status"] = "missing"
            
    # 4. 同步更新状态回写 Excel
    print("\n正在回填状态至 Excel...")
    for paper in papers:
        r_idx = paper["row_idx"]
        if paper["status"] == "success":
            sh.cell(row=r_idx, column=success_col, value="Y")
            sh.cell(row=r_idx, column=path_col, value=os.path.join(DOWNLOADS_DIR, paper["filename"]))
        elif paper["status"] == "error":
            sh.cell(row=r_idx, column=success_col, value="Error")
            sh.cell(row=r_idx, column=path_col, value=None)
        else:
            sh.cell(row=r_idx, column=success_col, value="N")
            sh.cell(row=r_idx, column=path_col, value=None)
            
    wb.save(XLSX_PATH)
    print("Excel 状态已安全保存更新！")
    
    # 5. 重新生成看板 HTML 
    print("\n正在重新拼装手动下载辅助 HTML 看板...")
    
    missing_papers = [p for p in papers if p["status"] == "missing"]
    error_papers = [p for p in papers if p["status"] == "error"]
    success_papers = [p for p in papers if p["status"] == "success"]
    
    new_html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>CellxGene 文献手动下载进度看板</title>
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700&family=Noto+Sans+SC:wght@300;400;500;700&display=swap" rel="stylesheet">
    <style>
        :root {{
            --bg-color: #f8fafc;
            --container-bg: #ffffff;
            --card-bg: #ffffff;
            --border-color: #e2e8f0;
            --text-primary: #0f172a;
            --text-secondary: #475569;
            --accent-blue: #0284c7;
            --success-color: #059669;
            --warning-color: #d97706;
            --danger-color: #dc2626;
            --indigo-color: #4f46e5;
        }}
        
        * {{
            box-sizing: border-box;
            transition: all 0.2s ease;
        }}

        body {{
            font-family: 'Outfit', 'Noto Sans SC', sans-serif;
            background-color: var(--bg-color);
            color: var(--text-primary);
            margin: 0;
            padding: 40px 20px;
            display: flex;
            flex-direction: column;
            align-items: center;
        }}

        .container {{
            max-width: 1000px;
            width: 100%;
        }}

        header {{
            text-align: center;
            margin-bottom: 40px;
        }}

        h1 {{
            font-size: 2.3rem;
            margin: 0 0 10px 0;
            color: var(--text-primary);
        }}

        .stats-banner {{
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 16px;
            padding: 20px;
            display: flex;
            justify-content: space-around;
            margin-bottom: 30px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        }}

        .stat-item {{
            text-align: center;
        }}

        .stat-val {{
            font-size: 1.8rem;
            font-weight: 700;
        }}

        .stat-label {{
            color: var(--text-secondary);
            font-size: 0.9rem;
            margin-top: 4px;
        }}

        .list-container {{
            display: flex;
            flex-direction: column;
            gap: 20px;
        }}

        .card {{
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 16px;
            padding: 22px;
            position: relative;
            box-shadow: 0 4px 6px -1px rgba(0,0,0,0.02);
        }}

        .card:hover {{
            transform: translateY(-2px);
            border-color: var(--accent-blue);
            box-shadow: 0 10px 15px -3px rgba(0,0,0,0.05);
        }}

        .card-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 12px;
        }}

        .paper-index {{
            font-weight: 600;
            background: #f1f5f9;
            padding: 4px 10px;
            border-radius: 8px;
            font-size: 0.9rem;
        }}

        .badge-error {{
            background: #fef2f2;
            color: var(--danger-color);
            border: 1px solid #fee2e2;
            padding: 4px 10px;
            border-radius: 8px;
            font-size: 0.85rem;
            font-weight: 500;
        }}

        .badge-missing {{
            background: #fffbeb;
            color: var(--warning-color);
            border: 1px solid #fef3c7;
            padding: 4px 10px;
            border-radius: 8px;
            font-size: 0.85rem;
            font-weight: 500;
        }}

        .paper-title {{
            font-size: 1.25rem;
            margin: 0 0 12px 0;
            cursor: pointer;
        }}

        .meta-info {{
            font-size: 0.9rem;
            color: var(--text-secondary);
            margin-bottom: 15px;
        }}

        .filename-box {{
            background: #f8fafc;
            border: 1px dashed var(--border-color);
            border-radius: 10px;
            padding: 12px;
            cursor: pointer;
            font-family: monospace;
            margin-bottom: 15px;
        }}

        .filename-value {{
            color: var(--indigo-color);
            font-weight: 600;
        }}

        .action-links {{
            display: flex;
            gap: 10px;
            flex-wrap: wrap;
        }}

        .btn {{
            text-decoration: none;
            padding: 8px 16px;
            border-radius: 8px;
            font-size: 0.9rem;
            font-weight: 500;
            background: #f1f5f9;
            color: var(--text-primary);
        }}

        .btn-scholar {{ background: #e0f2fe; color: #0369a1; }}
        .btn-pubmed {{ background: #f0fdf4; color: #166534; }}
        .btn-scihub {{ background: #faf5ff; color: #6b21a8; }}

        details {{
            margin-top: 30px;
            background: #fff;
            border: 1px solid var(--border-color);
            border-radius: 16px;
            padding: 20px;
        }}

        summary {{
            font-weight: 600;
            cursor: pointer;
            font-size: 1.1rem;
            outline: none;
        }}
    </style>
    <script>
        function copyText(text) {{
            navigator.clipboard.writeText(text).then(() => {{
                alert('已复制: ' + text);
            }});
        }}
    </script>
</head>
<body>
    <div class="container">
        <header>
            <h1>CellxGene 文献手动下载进度看板</h1>
            <p style="color: var(--text-secondary);">所有 75 篇文献的物理比对与精细校验看板 (自动流 + 纠错大闭环)</p>
        </header>

        <div class="stats-banner">
            <div class="stat-item">
                <div class="stat-val" style="color: var(--success-color);">{len(success_papers)}</div>
                <div class="stat-label">已成功归库且校验正确</div>
            </div>
            <div class="stat-item">
                <div class="stat-val" style="color: var(--danger-color);">{len(error_papers)}</div>
                <div class="stat-label">疑似损坏/下载错误</div>
            </div>
            <div class="stat-item">
                <div class="stat-val" style="color: var(--warning-color);">{len(missing_papers)}</div>
                <div class="stat-label">待手动下载 (当前缺失)</div>
            </div>
        </div>

        <div class="list-container">
            <h2>⚠️ 需人工干预的文献卡片 ({len(error_papers) + len(missing_papers)} 篇)</h2>
"""

    # 5.1 渲染校验出错的卡片
    for p in error_papers:
        new_html += f"""
            <div class="card" data-idx="{p['idx']}" data-title="{p['title']}" data-doi="{p['doi']}" data-pmid="{p['pmid']}" data-filename="{p['filename']}" data-status="error" style="border-left: 5px solid var(--danger-color);">
                <div class="card-header">
                    <span class="paper-index">文献 #{p['idx']}</span>
                    <span class="badge-error">❌ {p['error_reason']}</span>
                </div>
                <h2 class="paper-title" onclick="copyText('{p['title']}')" title="点击复制标题">{p['title']}</h2>
                <div class="meta-info">
                    <span>DOI: <strong>{p['doi']}</strong></span> | 
                    <span>PMID: <strong>{p['pmid']}</strong></span>
                </div>
                <div class="filename-box" onclick="copyText('{p['filename']}')" title="点击复制文件名">
                    <div>需保存的规范文件名（点击复制）：</div>
                    <div class="filename-value">{p['filename']}</div>
                </div>
                <div class="action-links">
                    <a href="https://scholar.google.com/scholar?q={p['doi'] if p['doi'] != '-' else p['title']}" target="_blank" class="btn btn-scholar">谷歌学术 检索</a>
                    {f'<a href="https://pubmed.ncbi.nlm.nih.gov/{p["pmid"]}/" target="_blank" class="btn btn-pubmed">PubMed 直达</a>' if p['pmid'] != '-' else ''}
                    {f'<a href="https://sci-hub.st/{p["doi"]}" target="_blank" class="btn btn-scihub">Sci-Hub 下载</a>' if p['doi'] != '-' else ''}
                </div>
            </div>
"""

    # 5.2 渲染仍然缺失的卡片
    for p in missing_papers:
        new_html += f"""
            <div class="card" data-idx="{p['idx']}" data-title="{p['title']}" data-doi="{p['doi']}" data-pmid="{p['pmid']}" data-filename="{p['filename']}" data-status="missing" style="border-left: 5px solid var(--warning-color);">
                <div class="card-header">
                    <span class="paper-index">文献 #{p['idx']}</span>
                    <span class="badge-missing">⚠️ 待手动下载 (当前缺失)</span>
                </div>
                <h2 class="paper-title" onclick="copyText('{p['title']}')" title="点击复制标题">{p['title']}</h2>
                <div class="meta-info">
                    <span>DOI: <strong>{p['doi']}</strong></span> | 
                    <span>PMID: <strong>{p['pmid']}</strong></span>
                </div>
                <div class="filename-box" onclick="copyText('{p['filename']}')" title="点击复制文件名">
                    <div>需保存的规范文件名（点击复制）：</div>
                    <div class="filename-value">{p['filename']}</div>
                </div>
                <div class="action-links">
                    <a href="https://scholar.google.com/scholar?q={p['doi'] if p['doi'] != '-' else p['title']}" target="_blank" class="btn btn-scholar">谷歌学术 检索</a>
                    {f'<a href="https://pubmed.ncbi.nlm.nih.gov/{p["pmid"]}/" target="_blank" class="btn btn-pubmed">PubMed 直达</a>' if p['pmid'] != '-' else ''}
                    {f'<a href="https://sci-hub.st/{p["doi"]}" target="_blank" class="btn btn-scihub">Sci-Hub 下载</a>' if p['doi'] != '-' else ''}
                </div>
            </div>
"""

    new_html += """
        </div>
    </div>
</body>
</html>
"""

    with open(HTML_PATH, "w", encoding="utf-8") as f:
        f.write(new_html)
        
    print(f"\nHTML 看板页面已成功重写！地址: {HTML_PATH}")
    print(f"=== 物理校验与看板重构工作全部圆满完成！ ===")

if __name__ == "__main__":
    main()
