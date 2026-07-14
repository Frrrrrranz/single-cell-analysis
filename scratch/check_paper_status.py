"""Check which papers have markers in the new markers sheet"""
import openpyxl, os, re

papers_dir = r"D:\OneDrive\Desktop\组\papers"
dir_papers = set()
for d in sorted(os.listdir(papers_dir)):
    dp = os.path.join(papers_dir, d)
    if os.path.isdir(dp):
        pdf = os.path.join(dp, f"{d}.full.pdf")
        if os.path.exists(pdf):
            dir_papers.add(d)

wb = openpyxl.load_workbook(r"D:\OneDrive\Desktop\组\db\pns-scrna.xlsx", data_only=True)
ws = wb["papers"]
h = [c.value for c in ws[1]]
col = {v:i for i,v in enumerate(h)}
db_papers = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    db_papers[row[0]] = {"title": row[col.get("title",1)], "status": row[col.get("status",12)]}

ws2 = wb["markers"]
h2 = [c.value for c in ws2[1]]
col2 = {v:i for i,v in enumerate(h2)}
marker_papers = set()
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=True):
    notes = str(row[col2.get("notes",9)] or "")
    m = re.search(r"paper=([^,)]+)", notes)
    if m:
        marker_papers.add(m.group(1))

print("=== Papers in directory ===")
for p in sorted(dir_papers):
    has_marker = " [MARKERS DONE]" if p in marker_papers else ""
    db_match = ""
    for pid, info in db_papers.items():
        if pid and (pid in p or p.startswith(pid)):
            db_match = f" [DB: {pid} | {info['status']}]"
    print(f"  {p}{has_marker}{db_match}")

print()
print("=== Database papers ===")
for pid, info in sorted(db_papers.items()):
    in_dir = any(pid in d for d in dir_papers)
    has_marker = " [MARKERS DONE]" if pid in marker_papers else ""
    print(f"  DB {pid}: {info['status']} | {str(info['title'])[:50]}{has_marker}")
    if not in_dir:
        print("    -> NOT in papers/ directory")

print()
print(f"Summary:")
print(f"  Papers in directory with PDF: {len(dir_papers)}")
print(f"  Papers in DB: {len(db_papers)}")
print(f"  Papers with markers: {len(marker_papers)}")
print(f"  In dir but NOT in DB:")
for p in sorted(dir_papers):
    if not any(pid in p for pid in db_papers):
        print(f"    {p}")
wb.close()
