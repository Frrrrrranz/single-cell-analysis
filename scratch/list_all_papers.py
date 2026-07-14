"""Wider search for our 9 papers in literature table"""
import openpyxl

wb = openpyxl.load_workbook(
    r"D:\OneDrive\Desktop\组\db\文献_细胞类型_工具一览表.xlsx",
    data_only=True
)
ws = wb.active

# Print ALL rows to find matching papers
print("ALL papers in literature table:")
for row_idx in range(2, ws.max_row + 1):
    title = str(ws.cell(row=row_idx, column=3).value or "")
    doi = str(ws.cell(row=row_idx, column=6).value or "")
    source = str(ws.cell(row=row_idx, column=1).value or "")
    cat = str(ws.cell(row=row_idx, column=2).value or "")
    pmid = str(ws.cell(row=row_idx, column=6).value or "")
    print(f"  Row {row_idx:2d} | {cat} | DOI={doi[:50]:50s} | {title[:70]}")

wb.close()
print()
print("=== Our 9 paper names ===")
pdfs = [
    "CELL-STEM-CELL.30.20.2023",
    "COMMUN-BIOL.5.1105.2022",
    "ENEURO.0066-20.2020",
    "FRONT-CELL-NEUROSCI.15.624826.2021",
    "GLIA.69.188.2020",
    "J-NEUROINFLAMM.22.205.2025",
    "NATURE.587.619.2020",
    "PNAS.117.9466.2020",
    "SCI-IMMUNOL.8.adf9988.2023",
]
for p in pdfs:
    print(f"  {p}")
